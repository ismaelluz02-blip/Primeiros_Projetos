import os
import glob
import re
import atexit
import ctypes
import sqlite3
import shutil
import sys
import unicodedata
import calendar
import time
from calendar import monthrange
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
import fitz
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

MESES = [
    "janeiro",
    "fevereiro",
    "marco",
    "abril",
    "maio",
    "junho",
    "julho",
    "agosto",
    "setembro",
    "outubro",
    "novembro",
    "dezembro",
]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
APP_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else SCRIPT_DIR
BASE_DIR = SCRIPT_DIR
PROJECT_DIR = APP_DIR if getattr(sys, "frozen", False) else os.path.dirname(BASE_DIR)
RELATORIOS_DIR = os.path.join(PROJECT_DIR, "RELATORIOS")
APP_DATA_DIR = os.path.join(
    os.environ.get("LOCALAPPDATA", BASE_DIR),
    "Horizonte Logistica",
    "Sistema de Faturamento",
)
DB_PATH = os.path.join(APP_DATA_DIR, "faturamento.db")
LOCK_PATH = os.path.join(APP_DATA_DIR, ".sistema_faturamento.lock")
LEGACY_DB_PATH = os.path.join(APP_DIR, "faturamento.db")
LOGO_PATH = os.path.join(APP_DIR, "logo.png")
APP_USER_MODEL_ID = "horizonte.logistica.sistema.faturamento"


def _processo_ativo(pid):
    try:
        os.kill(pid, 0)
        return True
    except:
        return False


def preparar_arquivos_aplicacao():
    os.makedirs(APP_DATA_DIR, exist_ok=True)

    def _contar_documentos(caminho_db):
        try:
            conn = sqlite3.connect(caminho_db)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM documentos")
            total = int(cursor.fetchone()[0])
            conn.close()
            return total
        except Exception:
            return None

    # Migra automaticamente o banco legado da pasta do executavel
    # para a pasta de dados do usuario na primeira execucao.
    if not os.path.exists(DB_PATH):
        if os.path.exists(LEGACY_DB_PATH):
            try:
                shutil.copy2(LEGACY_DB_PATH, DB_PATH)
            except OSError:
                pass
        return

    # Se o banco novo existir mas estiver vazio e o legado possuir dados,
    # reaproveita o legado para evitar "sumir" com historico em atualizacoes.
    if os.path.exists(LEGACY_DB_PATH):
        docs_novo = _contar_documentos(DB_PATH)
        docs_legado = _contar_documentos(LEGACY_DB_PATH)
        if docs_novo == 0 and isinstance(docs_legado, int) and docs_legado > 0:
            try:
                shutil.copy2(LEGACY_DB_PATH, DB_PATH)
            except OSError:
                pass


def adquirir_lock_instancia():
    pid_existente = None
    if os.path.exists(LOCK_PATH):
        try:
            with open(LOCK_PATH, "r", encoding="utf-8") as f:
                pid_txt = f.read().strip()
            pid_existente = int(pid_txt)
            if pid_existente != os.getpid() and _processo_ativo(pid_existente):
                return False, pid_existente
        except (ValueError, OSError):
            pass

        try:
            os.remove(LOCK_PATH)
        except OSError:
            return False, pid_existente

    try:
        with open(LOCK_PATH, "w", encoding="utf-8") as f:
            f.write(str(os.getpid()))
    except OSError:
        return False, pid_existente

    return True, None


def liberar_lock_instancia():
    try:
        if os.path.exists(LOCK_PATH):
            with open(LOCK_PATH, "r", encoding="utf-8") as f:
                pid_txt = f.read().strip()
            if pid_txt == str(os.getpid()):
                os.remove(LOCK_PATH)
    except OSError:
        pass


def alertar_instancia_em_execucao(pid_existente=None):
    msg = "O sistema de faturamento ja esta em execucao.\nFeche a janela atual antes de abrir novamente."
    if pid_existente:
        msg += f"\n\nPID em execucao: {pid_existente}"
    try:
        ctypes.windll.user32.MessageBoxW(0, msg, "Sistema ja em execucao", 0x30)
    except Exception:
        try:
            messagebox.showwarning("Sistema ja em execucao", msg)
        except Exception:
            print(msg)


# ------------------------
# BANCO
# ------------------------


def obter_conexao_banco():
    return sqlite3.connect(DB_PATH)


def obter_configuracao(chave, padrao=""):
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute("SELECT valor FROM configuracoes WHERE chave=?", (chave,))
    row = cursor.fetchone()
    conn.close()
    if row and row[0] is not None:
        return str(row[0])
    return padrao


def salvar_configuracao(chave, valor):
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO configuracoes (chave, valor)
        VALUES (?, ?)
        ON CONFLICT(chave) DO UPDATE SET valor=excluded.valor
        """,
        (chave, "" if valor is None else str(valor)),
    )
    conn.commit()
    conn.close()


def iniciar_banco():
    conn = obter_conexao_banco()
    cursor = conn.cursor()

    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS documentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero INTEGER,
            numero_original TEXT,
            tipo TEXT,
            data_emissao TEXT,
            valor_inicial REAL,
            valor_final REAL,
            frete TEXT,
            status TEXT,
            competencia TEXT
        )
        """
    )

    cursor.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_documentos_numero_tipo
        ON documentos (numero, tipo)
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS configuracoes (
            chave TEXT PRIMARY KEY,
            valor TEXT
        )
        """
    )

    colunas_existentes = {linha[1] for linha in cursor.execute("PRAGMA table_info(documentos)").fetchall()}
    if "valor_inicial_original" not in colunas_existentes:
        cursor.execute("ALTER TABLE documentos ADD COLUMN valor_inicial_original REAL")
    if "valor_final_original" not in colunas_existentes:
        cursor.execute("ALTER TABLE documentos ADD COLUMN valor_final_original REAL")
    if "status_original" not in colunas_existentes:
        cursor.execute("ALTER TABLE documentos ADD COLUMN status_original TEXT")
    if "cancelado_manual" not in colunas_existentes:
        cursor.execute("ALTER TABLE documentos ADD COLUMN cancelado_manual INTEGER DEFAULT 0")
    if "competencia_manual" not in colunas_existentes:
        cursor.execute("ALTER TABLE documentos ADD COLUMN competencia_manual INTEGER DEFAULT 0")

    conn.commit()
    conn.close()


# ------------------------
# VARIAVEIS
# ------------------------

relatorio_selecionado = ""
pasta_relatorios_saida = RELATORIOS_DIR

paginas_lidas = 0
docs_encontrados = 0


# ------------------------
# UTIL
# ------------------------

def valor_brasileiro(v):
    v = v.replace(".", "").replace(",", ".")
    return float(v)


def formatar_moeda_brl(valor):
    try:
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "R$ 0,00"


def parse_valor_monetario(valor_str):
    bruto = re.sub(r"[^\d,.\s]", "", str(valor_str)).replace(" ", "")
    if not bruto:
        return None

    if "," in bruto:
        try:
            return float(bruto.replace(".", "").replace(",", "."))
        except ValueError:
            return None

    if bruto.count(".") == 1:
        try:
            return float(bruto)
        except ValueError:
            return None

    if bruto.count(".") > 1:
        return None

    try:
        return float(bruto.replace(".", ""))
    except ValueError:
        return None


def normalizar_texto(texto):
    sem_acentos = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    return sem_acentos.upper()


def competencia_por_data(data):
    return f"{MESES[data.month - 1]}/{data.year}"


def periodo_padrao_mes_atual():
    hoje = datetime.now()
    primeiro_dia = datetime(hoje.year, hoje.month, 1)
    ultimo_dia = datetime(hoje.year, hoje.month, monthrange(hoje.year, hoje.month)[1])
    return primeiro_dia, ultimo_dia


def ler_data_filtro(texto_data, nome_campo):
    try:
        return datetime.strptime(texto_data.strip(), "%d/%m/%Y")
    except ValueError as exc:
        raise ValueError(f"{nome_campo} invalida. Use o formato DD/MM/AAAA.") from exc


def centralizar_janela(janela, largura, altura):
    janela.update_idletasks()
    screen_width = janela.winfo_screenwidth()
    screen_height = janela.winfo_screenheight()
    x = int((screen_width / 2) - (largura / 2))
    y = int((screen_height / 2) - (altura / 2))
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


def manter_interface_responsiva():
    try:
        if "app" in globals() and app.winfo_exists():
            app.update()
    except Exception:
        pass


def solicitar_atualizacao_dashboard(_event=None):
    try:
        if "atualizar_dashboard" in globals():
            atualizar_dashboard()
    except Exception:
        pass


def obter_pasta_saida_relatorios():
    global pasta_relatorios_saida
    if not pasta_relatorios_saida:
        pasta_relatorios_saida = RELATORIOS_DIR
    os.makedirs(pasta_relatorios_saida, exist_ok=True)
    return pasta_relatorios_saida


def atualizar_label_pasta_saida():
    if "pasta_saida_label" in globals():
        pasta_saida_label.configure(text=f"Pasta de saida: {obter_pasta_saida_relatorios()}")


def carregar_pasta_saida_relatorios():
    global pasta_relatorios_saida
    caminho_salvo = obter_configuracao("pasta_relatorios_saida", "").strip()
    if caminho_salvo:
        pasta_relatorios_saida = caminho_salvo
    obter_pasta_saida_relatorios()
    atualizar_label_pasta_saida()


def selecionar_pasta_saida_relatorios():
    global pasta_relatorios_saida
    caminho = filedialog.askdirectory(
        title="Selecionar pasta para salvar o relatorio",
        initialdir=obter_pasta_saida_relatorios(),
        mustexist=True,
    )
    if not caminho:
        return

    pasta_relatorios_saida = caminho
    salvar_configuracao("pasta_relatorios_saida", pasta_relatorios_saida)
    obter_pasta_saida_relatorios()
    atualizar_label_pasta_saida()
    messagebox.showinfo("Pasta de saida", f"Relatorios serao salvos em:\n{pasta_relatorios_saida}")


def aplicar_icone_aplicacao(janela):
    # Forca um AppUserModelID proprio para evitar agrupamento/icone do Python na barra.
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
    except Exception:
        pass

    try:
        icon_path = os.path.join(BASE_DIR, "logo.ico")
        if os.path.exists(icon_path):
            janela.iconbitmap(default=icon_path)
    except Exception as e:
        print(f"Aviso: nao foi possivel carregar o icone .ico - {e}")

    # Fallback para garantir icone em janelas Tk quando .ico falhar.
    try:
        if os.path.exists(LOGO_PATH):
            janela._app_icon_photo = tk.PhotoImage(file=LOGO_PATH)
            janela.iconphoto(True, janela._app_icon_photo)
    except Exception:
        pass


def abrir_seletor_data(entry_widget):
    texto_atual = entry_widget.get().strip()
    try:
        data_base = datetime.strptime(texto_atual, "%d/%m/%Y")
    except ValueError:
        data_base = datetime.now()

    picker = ctk.CTkToplevel(app)
    picker.title("Selecionar data")
    centralizar_janela(picker, 290, 280)
    picker.grab_set()

    estado = {"ano": data_base.year, "mes": data_base.month}
    semana_nomes = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]

    topo = ctk.CTkFrame(picker, fg_color="transparent")
    topo.pack(fill="x", padx=8, pady=(8, 4))

    def mudar_mes(delta):
        mes = estado["mes"] + delta
        ano = estado["ano"]
        if mes < 1:
            mes = 12
            ano -= 1
        elif mes > 12:
            mes = 1
            ano += 1
        estado["mes"] = mes
        estado["ano"] = ano
        renderizar()

    ctk.CTkButton(topo, text="<", width=30, command=lambda: mudar_mes(-1)).pack(side="left")
    titulo_mes = ctk.CTkLabel(topo, text="", font=ctk.CTkFont(weight="bold"))
    titulo_mes.pack(side="left", expand=True)
    ctk.CTkButton(topo, text=">", width=30, command=lambda: mudar_mes(1)).pack(side="right")

    corpo = ctk.CTkFrame(picker)
    corpo.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def selecionar_dia(dia):
        valor = datetime(estado["ano"], estado["mes"], dia).strftime("%d/%m/%Y")
        entry_widget.delete(0, "end")
        entry_widget.insert(0, valor)
        solicitar_atualizacao_dashboard()
        picker.destroy()

    def renderizar():
        for child in corpo.winfo_children():
            child.destroy()

        titulo_mes.configure(text=f"{MESES[estado['mes'] - 1].capitalize()}/{estado['ano']}")

        for c, nome in enumerate(semana_nomes):
            ctk.CTkLabel(corpo, text=nome).grid(row=0, column=c, padx=2, pady=(4, 2))

        semanas = calendar.monthcalendar(estado["ano"], estado["mes"])
        for r, semana in enumerate(semanas, start=1):
            for c, dia in enumerate(semana):
                if dia == 0:
                    ctk.CTkLabel(corpo, text="", width=34).grid(row=r, column=c, padx=2, pady=2)
                else:
                    ctk.CTkButton(
                        corpo,
                        text=str(dia),
                        width=34,
                        height=28,
                        command=lambda d=dia: selecionar_dia(d),
                    ).grid(row=r, column=c, padx=2, pady=2)

    renderizar()


# ------------------------
# PORTAL NFS-E
# ------------------------


# ------------------------
# EXTRAIR CTE
# ------------------------

# ------------------------
# EXTRAIR NF
# ------------------------

def _normalizar_mes_relatorio(mes_txt):
    mes_norm = normalizar_texto(mes_txt).lower()
    mapa = {normalizar_texto(m).lower(): m for m in MESES}
    return mapa.get(mes_norm, mes_norm)


def _extrair_docs_pagina_relatorio(texto_pagina, competencia_atual=None):
    docs = []
    linhas = [l.strip() for l in texto_pagina.splitlines() if l.strip()]

    header_re = re.compile(r"(?:C\.T\.R\.C\.|N\.F\.)\s*-\s*([A-Za-z?-?]+)\s*/\s*(\d{2,4})", re.IGNORECASE)
    money_re = re.compile(r"R\$\s*([0-9\.,]+)")
    date_re = re.compile(r"^\d{2}/\d{2}/\d{4}$")

    i = 0
    while i < len(linhas):
        linha = linhas[i]
        m_header = header_re.search(linha)
        if m_header:
            mes_txt = _normalizar_mes_relatorio(m_header.group(1))
            ano_txt = m_header.group(2)
            ano = int(ano_txt)
            if ano < 100:
                ano += 2000
            competencia_atual = f"{mes_txt}/{ano}"

        if linha.isdigit() and (i + 1) < len(linhas) and linhas[i + 1].isdigit():
            window_end = min(i + 35, len(linhas))
            tipo_idx = None
            tipo = None
            for j in range(i + 2, window_end):
                t = normalizar_texto(linhas[j])
                if t in {"CTRC", "NF"}:
                    tipo_idx = j
                    tipo = t
                    break

            if tipo_idx is not None:
                trecho = linhas[i:tipo_idx + 1]
                codigo = int(linha)

                valor = None
                for item in trecho:
                    mm = money_re.search(item)
                    if mm:
                        v = parse_valor_monetario(mm.group(1))
                        if v is not None and v > 0:
                            valor = v
                            break

                data = None
                for item in trecho:
                    if date_re.match(item):
                        try:
                            data = datetime.strptime(item, "%d/%m/%Y")
                            break
                        except ValueError:
                            pass

                if data and valor is not None:
                    if tipo == "NF":
                        numero = int(f"{data.year}{codigo:04d}")
                        valor_final = valor * 0.95
                    else:
                        numero = codigo
                        valor_final = valor

                    docs.append({
                        "numero": numero,
                        "numero_original": codigo,
                        "tipo": "NF" if tipo == "NF" else "CTE",
                        "data": data,
                        "valor_inicial": valor,
                        "valor_final": valor_final,
                        "frete": "FRANQUIA",
                        "status": "OK",
                        "competencia": competencia_atual.lower() if competencia_atual else competencia_por_data(data),
                    })

                i = tipo_idx + 1
                continue

        i += 1

    return docs, competencia_atual


def importar_relatorio_consolidado(caminho_pdf):
    global docs_encontrados, paginas_lidas

    docs_encontrados = 0
    paginas_lidas = 0

    vistos = {}
    competencia_ctx = None

    try:
        with fitz.open(caminho_pdf) as pdf:
            total_pag = pdf.page_count
            progress.set(0)
            for idx, pagina in enumerate(pdf, start=1):
                texto = pagina.get_text("text")
                docs, competencia_ctx = _extrair_docs_pagina_relatorio(texto, competencia_ctx)
                for doc in docs:
                    chave = (doc["tipo"], doc["numero"])
                    if chave in vistos:
                        continue
                    vistos.add(chave)
                    salvar_documento(doc)
                    docs_encontrados += 1

                paginas_lidas = idx
                progress.set(idx / total_pag)
                status_label.configure(text=f"Importando relatorio: pagina {idx}/{total_pag} | Docs:{docs_encontrados}")
                manter_interface_responsiva()
    except Exception as exc:
        messagebox.showerror("Relatorio", f"Falha ao importar relatorio.\n\n{exc}")
        return

    messagebox.showinfo("Relatorio", f"Importacao concluida. {docs_encontrados} documento(s) carregado(s).")

    atualizar_dashboard()


def _normalizar_coluna_relatorio(nome_coluna):
    base = normalizar_texto(str(nome_coluna))
    base = re.sub(r"[^A-Z0-9 ]", " ", base)
    return " ".join(base.split())


def _achar_coluna(colunas_norm, regras):
    for regra in regras:
        for col_original, col_norm in colunas_norm.items():
            if all(chave in col_norm for chave in regra):
                return col_original
    return None


def _achar_coluna_exata(colunas_norm, nomes_exatos):
    nomes = {n.upper() for n in nomes_exatos}
    for col_original, col_norm in colunas_norm.items():
        if col_norm in nomes:
            return col_original
    return None


def _parse_tipo_documento(valor_tipo):
    t = normalizar_texto(str(valor_tipo))
    if "CTE" in t or "CTRC" in t:
        return "CTE"
    if "NF" in t:
        return "NF"
    return None


def _mapear_colunas_planilha(df):
    colunas_norm = {col: _normalizar_coluna_relatorio(col) for col in df.columns}

    col_tipo = _achar_coluna(colunas_norm, [
        ("TIPO", "DOC"),
        ("TIPO",),
    ])
    col_serie = _achar_coluna_exata(colunas_norm, ["SERIE"]) or _achar_coluna(colunas_norm, [("SERIE",)])
    col_numero = (
        _achar_coluna_exata(colunas_norm, ["CODIGO", "COD"])
        or _achar_coluna(colunas_norm, [
            ("NUMERO", "DOC"),
            ("NUM", "DOC"),
            ("NRO", "DOC"),
            ("NR", "DOC"),
            ("DOC", "NUM"),
            ("CODIGO",),
            ("COD",),
            ("NUMERO",),
            ("DOCUMENTO",),
            ("NR",),
        ])
    )

    # Data de emissao: sempre prioriza coluna "Data" (sem "Ref").
    col_data = _achar_coluna_exata(colunas_norm, ["DATA"])
    if not col_data:
        for col_original, col_norm in colunas_norm.items():
            if (("DATA" in col_norm) or (col_norm == "DT") or ("EMISSAO" in col_norm)) and ("REF" not in col_norm):
                col_data = col_original
                break

    # Data de referencia: apenas fallback quando Data estiver vazia/invalida.
    col_data_ref = _achar_coluna_exata(colunas_norm, ["DATA REF", "DT REF"])
    if not col_data_ref:
        for col_original, col_norm in colunas_norm.items():
            if ("REF" in col_norm) and (("DATA" in col_norm) or ("DT" in col_norm)):
                col_data_ref = col_original
                break

    col_valor = _achar_coluna(colunas_norm, [
        ("FRETE",),
        ("VALOR", "DOCUMENTO"),
        ("VALOR", "TOTAL"),
        ("VLR", "DOCUMENTO"),
        ("VLR", "TOTAL"),
        ("VALOR", "FRETE"),
        ("VLR", "FRETE"),
        ("TOTAL", "FRETE"),
        ("VALOR", "RECEBER"),
        ("VALOR",),
        ("VLR",),
    ])
    col_frete = _achar_coluna(colunas_norm, [
        ("FRETE",),
    ])
    col_status = _achar_coluna(colunas_norm, [
        ("STATUS",),
        ("SITUACAO",),
    ])
    col_filial = _achar_coluna_exata(colunas_norm, ["FILIAL"]) or _achar_coluna(colunas_norm, [
        ("FILIAL",),
    ])
    col_pagador = _achar_coluna_exata(colunas_norm, ["PAGADOR"]) or _achar_coluna(colunas_norm, [
        ("PAGADOR",),
        ("CLIENTE",),
    ])

    faltando = []
    if not col_numero:
        faltando.append("numero")
    if not col_data:
        faltando.append("data_emissao")
    if not col_valor:
        faltando.append("valor")

    return {
        "tipo": col_tipo,
        "serie": col_serie,
        "numero": col_numero,
        "data": col_data,
        "data_ref": col_data_ref,
        "valor": col_valor,
        "frete": col_frete,
        "status": col_status,
        "filial": col_filial,
        "pagador": col_pagador,
        "faltando": faltando,
        "ordem_colunas": list(df.columns),
        "colunas_norm": colunas_norm,
    }



def _linha_valida_para_importacao(linha, mapa):
    def _normalizar_filial(valor_raw):
        if pd.isna(valor_raw):
            return ""
        txt = str(valor_raw).strip()
        # Excel pode trazer 88 como 88.0
        if re.fullmatch(r"\d+(?:\.0+)?", txt):
            try:
                return str(int(float(txt)))
            except ValueError:
                return ""
        return re.sub(r"\D", "", txt)

    # 1) Filial obrigatoriamente 88
    if not mapa.get("filial"):
        return False
    filial_txt = _normalizar_filial(linha.get(mapa["filial"], ""))
    if filial_txt != "88":
        return False

    # 2) Numero do documento obrigatorio no campo codigo/documento
    numero_txt = re.sub(r"\D", "", str(linha.get(mapa["numero"], "")))
    if not numero_txt:
        return False
    # Evita capturar datas como codigo (ex.: 27022026 / 20260227) e valores fora do padrao esperado.
    if len(numero_txt) > 6:
        return False
    if re.fullmatch(r"\d{8}", numero_txt):
        try:
            datetime.strptime(numero_txt, "%d%m%Y")
            return False
        except ValueError:
            pass
        try:
            datetime.strptime(numero_txt, "%Y%m%d")
            return False
        except ValueError:
            pass

    # 3) Data de emissao obrigatoria
    data_raw = linha.get(mapa["data"])
    data = pd.to_datetime(data_raw, dayfirst=True, errors="coerce")
    if pd.isna(data) and mapa.get("data_ref"):
        data = pd.to_datetime(linha.get(mapa["data_ref"]), dayfirst=True, errors="coerce")
    if pd.isna(data):
        return False

    # 4) Pagador obrigatoriamente Energisa
    if mapa.get("pagador"):
        pagador_txt = normalizar_texto(str(linha.get(mapa["pagador"], "")))
    else:
        pagador_txt = normalizar_texto(" ".join(str(v) for v in linha.tolist()))
    if "ENERGISA" not in pagador_txt:
        return False

    # 5) Ignora linhas de total/somatorio
    texto_linha = normalizar_texto(" ".join(str(v) for v in linha.tolist()))
    if any(chave in texto_linha for chave in [
        "TOTAL FRETE",
        "TOTAL FILIAL",
        "TOTAL C T R C",
        "TOTAL N F",
        "TOTAL FRETE N F",
        "TOTAL FRETE MARCO",
        "=>",
    ]):
        return False

    return True

def _inferir_tipo_documento_linha(linha, mapa):
    tipo_secao = normalizar_texto(str(linha.get("__secao_tipo", "")))
    if tipo_secao in {"NF", "CTE"}:
        return tipo_secao

    if mapa.get("serie"):
        serie_txt = re.sub(r"\D", "", str(linha.get(mapa["serie"], "")))
        if serie_txt == "1":
            return "NF"
        if serie_txt == "2":
            return "CTE"

    if mapa.get("tipo"):
        tipo = _parse_tipo_documento(linha.get(mapa["tipo"]))
        if tipo:
            return tipo

    texto_linha = normalizar_texto(" ".join(str(v) for v in linha.tolist()))
    if "NF" in texto_linha or "NOTA" in texto_linha:
        return "NF"
    if "CTE" in texto_linha or "CTRC" in texto_linha:
        return "CTE"

    return "CTE"


def _extrair_valor_frete_linha(linha, mapa):
    col_frete = mapa.get("frete")
    if not col_frete or col_frete not in linha.index:
        return None

    valor_raw = linha.get(col_frete)
    valor = parse_valor_monetario(valor_raw)
    if valor is None or valor <= 0:
        return None
    if valor > 100_000:
        return None
    return float(valor)


def _normalizar_nome_coluna_planilha(valor, idx_coluna):
    nome = str(valor).strip()
    if not nome or nome.lower() == "nan":
        return f"COL_{idx_coluna + 1}"
    return nome


def _linha_parece_cabecalho_planilha(row_vals):
    tokens = {_normalizar_coluna_relatorio(v) for v in row_vals}
    if "FILIAL" not in tokens:
        return False

    tem_codigo = any(t in {"CODIGO", "COD"} for t in tokens)
    tem_serie = "SERIE" in tokens
    tem_data = any(t.startswith("DATA") for t in tokens)
    tem_pagador = "PAGADOR" in tokens

    return tem_codigo and tem_serie and tem_data and tem_pagador


def _identificar_secao_planilha(row_vals):
    texto = normalizar_texto(" ".join(str(v) for v in row_vals))
    # Considera secao apenas na linha de titulo do bloco: "C.T.R.C. - janeiro / 26" ou "N.F. - janeiro / 26".
    # Aceita variacoes com/sem ponto final e com hifen simples ou longo.
    if not (("/" in texto) and (("-" in texto) or ("–" in texto))):
        return None
    if re.search(r"\bC\s*\.?\s*T\s*\.?\s*R\s*\.?\s*C\s*\.?\s*[-–]\s*[A-Z ]+\s*/\s*\d{2,4}", texto):
        return "CTE"
    if re.search(r"\bN\s*\.?\s*F\s*\.?\s*[-–]\s*[A-Z ]+\s*/\s*\d{2,4}", texto):
        return "NF"
    return None


def _linha_totalizadora_planilha(row_vals):
    texto = normalizar_texto(" ".join(str(v) for v in row_vals))
    return any(chave in texto for chave in [
        "TOTAL FRETE",
        "TOTAL FILIAL",
        "TOTAL C T R C",
        "TOTAL N F",
        "TOTAL FRETE N F",
        "TOTAL FRETE MARCO",
        "=>",
    ])


def _preparar_dataframe_planilha(df_raw):
    if df_raw is None or df_raw.empty:
        return pd.DataFrame()

    blocos = []
    i = 0
    total_linhas = len(df_raw.index)

    encontrou_secao = False
    while i < total_linhas:
        row_vals = df_raw.iloc[i].tolist()
        tipo_secao = _identificar_secao_planilha(row_vals)
        if not tipo_secao:
            i += 1
            continue
        encontrou_secao = True

        idx_cabecalho = None
        limite_busca = min(i + 8, total_linhas)
        for idx_tentativa in range(i + 1, limite_busca):
            if _linha_parece_cabecalho_planilha(df_raw.iloc[idx_tentativa].tolist()):
                idx_cabecalho = idx_tentativa
                break

        if idx_cabecalho is None:
            i += 1
            continue

        row_vals = df_raw.iloc[idx_cabecalho].tolist()
        headers = []
        usados = {}
        for idx_coluna, valor in enumerate(row_vals):
            nome_base = _normalizar_nome_coluna_planilha(valor, idx_coluna)
            if nome_base in usados:
                usados[nome_base] += 1
                nome = f"{nome_base}_{usados[nome_base]}"
            else:
                usados[nome_base] = 1
                nome = nome_base
            headers.append(nome)

        j = idx_cabecalho + 1
        linhas_bloco = []
        while j < total_linhas:
            prox_vals = df_raw.iloc[j].tolist()

            if _identificar_secao_planilha(prox_vals):
                break

            if _linha_totalizadora_planilha(prox_vals):
                break

            if not all(pd.isna(v) or str(v).strip() == "" for v in prox_vals):
                linhas_bloco.append(prox_vals)

            j += 1

        if linhas_bloco:
            df_bloco = pd.DataFrame(linhas_bloco, columns=headers)
            df_bloco = df_bloco.dropna(how="all")
            if not df_bloco.empty:
                df_bloco["__secao_tipo"] = tipo_secao
                blocos.append(df_bloco)

        i = j

    if blocos:
        return pd.concat(blocos, ignore_index=True, sort=False)

    # Fallback: se nao montar blocos por secao (ou secao estiver ausente),
    # usa blocos por cabecalho e para em totalizadores.
    if not blocos:
        i = 0
        while i < total_linhas:
            row_vals = df_raw.iloc[i].tolist()
            if not _linha_parece_cabecalho_planilha(row_vals):
                i += 1
                continue

            headers = []
            usados = {}
            for idx_coluna, valor in enumerate(row_vals):
                nome_base = _normalizar_nome_coluna_planilha(valor, idx_coluna)
                if nome_base in usados:
                    usados[nome_base] += 1
                    nome = f"{nome_base}_{usados[nome_base]}"
                else:
                    usados[nome_base] = 1
                    nome = nome_base
                headers.append(nome)

            j = i + 1
            linhas_bloco = []
            while j < total_linhas:
                prox_vals = df_raw.iloc[j].tolist()
                if _linha_parece_cabecalho_planilha(prox_vals):
                    break
                if _linha_totalizadora_planilha(prox_vals):
                    break
                if not all(pd.isna(v) or str(v).strip() == "" for v in prox_vals):
                    linhas_bloco.append(prox_vals)
                j += 1

            if linhas_bloco:
                df_bloco = pd.DataFrame(linhas_bloco, columns=headers)
                df_bloco = df_bloco.dropna(how="all")
                if not df_bloco.empty:
                    df_bloco["__secao_tipo"] = ""
                    blocos.append(df_bloco)

            i = j

        if blocos:
            return pd.concat(blocos, ignore_index=True, sort=False)

    return pd.DataFrame()
def importar_relatorio_planilha(caminho_planilha):
    global docs_encontrados, paginas_lidas

    docs_encontrados = 0
    paginas_lidas = 0

    try:
        planilhas = pd.read_excel(caminho_planilha, sheet_name=None, header=None)
    except ImportError as exc:
        messagebox.showerror(
            "Relatorio",
            "Falha ao abrir planilha. Para arquivo .xls, instale 'xlrd' ou exporte para .xlsx.\n\n" + str(exc),
        )
        return
    except Exception as exc:
        messagebox.showerror("Relatorio", f"Falha ao abrir planilha.\n\n{exc}")
        return

    if not planilhas:
        messagebox.showwarning("Relatorio", "A planilha nao possui abas para importar.")
        return

    # Limpa importacoes automaticas anteriores para evitar residuos no novo relatorio.
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute(
        """
        DELETE FROM documentos
        WHERE COALESCE(cancelado_manual,0)=0
          AND COALESCE(competencia_manual,0)=0
          AND UPPER(COALESCE(status,'')) NOT LIKE '%CANCELADO%'
          AND UPPER(COALESCE(status,'')) NOT LIKE '%SUBSTITUIDO%'
          AND UPPER(COALESCE(status,'')) NOT LIKE 'DOCUMENTO SUBSTITUINDO%'
        """
    )
    conn.commit()
    conn.close()

    vistos = {}

    conn = obter_conexao_banco()
    cursor = conn.cursor()

    erros = []
    total_linhas = 0
    planilhas_preparadas = {}
    for nome_aba, df in planilhas.items():
        df_prep = _preparar_dataframe_planilha(df)
        planilhas_preparadas[nome_aba] = df_prep
        if df_prep is not None and not df_prep.empty:
            total_linhas += len(df_prep.index)

    if total_linhas == 0:
        messagebox.showwarning("Relatorio", "Nao ha linhas com dados na planilha selecionada.")
        return

    processadas = 0
    total_abas = len(planilhas)

    for idx_aba, (nome_aba, df_aba) in enumerate(planilhas_preparadas.items(), start=1):
        paginas_lidas = idx_aba

        if df_aba is None or df_aba.empty:
            continue

        mapa = _mapear_colunas_planilha(df_aba)
        if mapa["faltando"]:
            preview_cols = ", ".join([str(c) for c in list(df_aba.columns)[:20]])
            erros.append(f"Aba '{nome_aba}' sem colunas obrigatorias: {', '.join(mapa['faltando'])}. Colunas detectadas: {preview_cols}")
            continue

        for _, linha in df_aba.iterrows():
            processadas += 1

            if not _linha_valida_para_importacao(linha, mapa):
                continue

            tipo = _inferir_tipo_documento_linha(linha, mapa)

            numero_txt = re.sub(r"\D", "", str(linha.get(mapa["numero"], "")))
            if not numero_txt:
                continue

            try:
                numero = int(numero_txt)
            except ValueError:
                continue

            data_raw = linha.get(mapa["data"])
            data = pd.to_datetime(data_raw, dayfirst=True, errors="coerce")
            if pd.isna(data) and mapa.get("data_ref"):
                data = pd.to_datetime(linha.get(mapa["data_ref"]), dayfirst=True, errors="coerce")
            if pd.isna(data):
                continue
            data = data.to_pydatetime()

            valor = _extrair_valor_frete_linha(linha, mapa)
            status = "OK"
            if valor is None:
                valor_inicial = 0.0
                status = "ERRO NO LANCAMENTO NO SISTEMA"
            else:
                valor_inicial = float(valor)

            if tipo == "NF":
                numero = int(f"{data.year}{int(numero_txt):04d}")
                valor_final = round(valor_inicial * 0.95, 2) if valor_inicial > 0 else 0.0
            else:
                numero = int(numero_txt)
                valor_final = valor_inicial

            frete = "FRANQUIA"
            if mapa["status"]:
                status_raw = str(linha.get(mapa["status"], "")).strip()
                if status == "OK" and status_raw and status_raw.lower() != "nan":
                    status = status_raw

            chave = (tipo, numero_txt) if tipo == "NF" else (tipo, numero)
            status_chave = "OK" if status == "OK" else "ERRO"
            if chave in vistos:
                status_existente = vistos[chave]
                # Mantem o registro mais confiavel: OK > ERRO.
                if status_existente == "OK":
                    continue
                if status_existente == "ERRO" and status_chave == "ERRO":
                    continue

            salvar_documento(
                {
                    "numero": numero,
                    "numero_original": numero_txt,
                    "tipo": tipo,
                    "data": data,
                    "valor_inicial": valor_inicial,
                    "valor_final": valor_final,
                    "frete": frete,
                    "status": status,
                    "competencia": competencia_por_data(data),
                },
                cursor
            )
            vistos[chave] = status_chave
            docs_encontrados += 1

            progress.set(processadas / total_linhas)
            status_label.configure(
                text=f"Importando planilha: linha {processadas}/{total_linhas} | Docs:{docs_encontrados}"
            )
            manter_interface_responsiva()

        progress.set(idx_aba / total_abas)
        status_label.configure(
            text=f"Importando planilha: aba {idx_aba}/{total_abas} | Docs:{docs_encontrados}"
        )
        manter_interface_responsiva()

    conn.commit()
    conn.close()

    resumo = f"Importacao concluida. {docs_encontrados} documento(s) carregado(s)."
    if erros:
        resumo += "\n\nAvisos:\n- " + "\n- ".join(erros[:5])
        if len(erros) > 5:
            resumo += f"\n- ... e mais {len(erros) - 5} aviso(s)."

    messagebox.showinfo("Relatorio", resumo)

    atualizar_dashboard()


def selecionar_relatorio():
    global relatorio_selecionado
    caminho = filedialog.askopenfilename(
        title="Selecionar relatorio consolidado",
        initialdir=RELATORIOS_DIR,
        filetypes=[
            ("Relatorios", "*.xlsx *.xls *.pdf"),
            ("Excel", "*.xlsx *.xls"),
            ("PDF", "*.pdf"),
            ("Todos os arquivos", "*.*"),
        ],
    )
    if caminho:
        relatorio_selecionado = caminho
        pasta_label.configure(text=relatorio_selecionado)


def importar_relatorio_ui():
    try:
        caminho = relatorio_selecionado
        if not caminho and "pasta_label" in globals():
            txt_label = str(pasta_label.cget("text") or "").strip()
            if txt_label and txt_label.lower() != "nenhum relatorio selecionado":
                caminho = txt_label

        if not caminho:
            messagebox.showwarning("Relatorio", "Selecione um relatorio antes de importar.")
            return

        if not os.path.exists(caminho):
            messagebox.showerror("Relatorio", "Arquivo selecionado nao foi encontrado.")
            return

        ext = os.path.splitext(caminho)[1].lower()
        if ext == ".pdf":
            importar_relatorio_consolidado(caminho)
            return

        if ext in {".xlsx", ".xls"}:
            importar_relatorio_planilha(caminho)
            return

        messagebox.showerror("Relatorio", "Formato de arquivo nao suportado. Use PDF, XLSX ou XLS.")
    except Exception as exc:
        try:
            status_label.configure(text="Falha ao importar relatorio.")
        except Exception:
            pass
        messagebox.showerror("Relatorio", f"Falha ao importar relatorio.\n\n{exc}")


# ------------------------
# SALVAR
# ------------------------

def salvar_documento(doc, cursor=None):
    conn_externo = cursor is not None
    if not conn_externo:
        conn = obter_conexao_banco()
        cursor = conn.cursor()

    try:
        tipo_doc = str(doc.get("tipo", "")).upper()
        if tipo_doc == "NF":
            numero_original = str(doc.get("numero_original", "")).strip()
            try:
                numero_original_int = int(re.sub(r"\D", "", numero_original))
            except ValueError:
                numero_original_int = None

            if numero_original:
                if numero_original_int is not None:
                    cursor.execute(
                        """
                        DELETE FROM documentos
                        WHERE tipo='NF' AND numero<>? AND (numero_original=? OR numero=?)
                        """,
                        (doc["numero"], numero_original, numero_original_int),
                    )
                else:
                    cursor.execute(
                        """
                        DELETE FROM documentos
                        WHERE tipo='NF' AND numero<>? AND numero_original=?
                        """,
                        (doc["numero"], numero_original),
                    )

        cursor.execute(
            """
            INSERT INTO documentos
            (
                numero,numero_original,tipo,data_emissao,valor_inicial,valor_final,frete,status,competencia,
                valor_inicial_original,valor_final_original,status_original,cancelado_manual,competencia_manual
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(numero,tipo) DO UPDATE SET
                numero_original=excluded.numero_original,
                data_emissao=excluded.data_emissao,
                frete=excluded.frete,
                competencia=CASE
                    WHEN documentos.competencia_manual=1 THEN documentos.competencia
                    ELSE excluded.competencia
                END,
                valor_inicial=CASE
                    WHEN documentos.cancelado_manual=1 THEN documentos.valor_inicial
                    ELSE excluded.valor_inicial
                END,
                valor_final=CASE
                    WHEN documentos.cancelado_manual=1 THEN documentos.valor_final
                    ELSE excluded.valor_final
                END,
                status=CASE
                    WHEN documentos.cancelado_manual=1 THEN documentos.status
                    ELSE excluded.status
                END,
                valor_inicial_original=CASE
                    WHEN documentos.cancelado_manual=1 THEN documentos.valor_inicial_original
                    ELSE NULL
                END,
                valor_final_original=CASE
                    WHEN documentos.cancelado_manual=1 THEN documentos.valor_final_original
                    ELSE NULL
                END,
                status_original=CASE
                    WHEN documentos.cancelado_manual=1 THEN documentos.status_original
                    ELSE NULL
                END
            """,
            (
                doc["numero"],
                doc["numero_original"],
                doc["tipo"],
                doc["data"].strftime("%d/%m/%Y"),
                doc["valor_inicial"],
                doc["valor_final"],
                doc["frete"],
                doc["status"],
                doc.get("competencia", competencia_por_data(doc["data"])),
                None,
                None,
                None,
                0,
                0,
            ),
        )
        if not conn_externo:
            conn.commit()
    finally:
        if not conn_externo:
            conn.close()


def alterar_competencia_documento(tipo, numero, mes_competencia, ano_competencia):
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    competencia = f"{mes_competencia}/{ano_competencia}"
    cursor.execute(
        "UPDATE documentos SET competencia=?, competencia_manual=1 WHERE tipo=? AND numero=?",
        (competencia, tipo, numero),
    )
    alterados = cursor.rowcount
    conn.commit()
    conn.close()

    atualizar_dashboard()

    return alterados


def declarar_documento_frete(tipo, numero, novo_frete):
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE documentos SET frete=? WHERE tipo=? AND numero=?",
        (novo_frete, tipo, numero),
    )
    alterados = cursor.rowcount
    conn.commit()
    conn.close()

    atualizar_dashboard()

    return alterados


def registrar_substituicao(tipo_antigo, numero_antigo, tipo_novo, numero_novo):
    conn = obter_conexao_banco()
    cursor = conn.cursor()

    status_novo = f"DOCUMENTO SUBSTITUINDO DOCUMENTO {numero_antigo} {tipo_antigo}"
    status_antigo = f"DOCUMENTO SUBSTITUIDO POR {numero_novo} {tipo_novo}"

    cursor.execute(
        """
        UPDATE documentos
        SET
            status_original=COALESCE(status_original, status),
            status=?
        WHERE tipo=? AND numero=?
        """,
        (status_novo, tipo_novo, numero_novo),
    )
    novo_alterado = cursor.rowcount

    cursor.execute(
        """
        UPDATE documentos
        SET
            valor_inicial_original=COALESCE(valor_inicial_original, valor_inicial),
            valor_final_original=COALESCE(valor_final_original, valor_final),
            status_original=COALESCE(status_original, status),
            valor_inicial=0,
            valor_final=0,
            status=?
        WHERE tipo=? AND numero=?
        """,
        (status_antigo, tipo_antigo, numero_antigo),
    )
    antigo_alterado = cursor.rowcount

    conn.commit()
    conn.close()

    atualizar_dashboard()

    return novo_alterado, antigo_alterado


def desfazer_substituicao(tipo_antigo, numero_antigo, tipo_novo, numero_novo):
    conn = obter_conexao_banco()
    cursor = conn.cursor()

    cursor.execute(
        """
        UPDATE documentos
        SET
            valor_inicial=COALESCE(valor_inicial_original, valor_inicial),
            valor_final=COALESCE(valor_final_original, valor_final),
            status=COALESCE(status_original, 'OK'),
            valor_inicial_original=NULL,
            valor_final_original=NULL,
            status_original=NULL
        WHERE tipo=? AND numero=? AND UPPER(status) LIKE 'DOCUMENTO SUBSTITUIDO POR%'
        """,
        (tipo_antigo, numero_antigo),
    )
    antigo_restaurado = cursor.rowcount

    cursor.execute(
        """
        UPDATE documentos
        SET
            status=COALESCE(status_original, 'OK'),
            status_original=NULL
        WHERE tipo=? AND numero=? AND UPPER(status) LIKE 'DOCUMENTO SUBSTITUINDO DOCUMENTO%'
        """,
        (tipo_novo, numero_novo),
    )
    novo_restaurado = cursor.rowcount

    conn.commit()
    conn.close()

    atualizar_dashboard()

    return antigo_restaurado, novo_restaurado


def cancelar_documento(tipo, numero):
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE documentos
        SET
            valor_inicial_original=COALESCE(valor_inicial_original, valor_inicial),
            valor_final_original=COALESCE(valor_final_original, valor_final),
            status_original=COALESCE(status_original, status),
            valor_inicial=0,
            valor_final=0,
            status='CANCELADO MANUALMENTE',
            cancelado_manual=1
        WHERE tipo=? AND numero=?
        """,
        (tipo, numero),
    )
    alterados = cursor.rowcount
    conn.commit()
    conn.close()

    atualizar_dashboard()

    return alterados


def desfazer_cancelamento_documento(tipo, numero):
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute(
        """
        UPDATE documentos
        SET
            valor_inicial=COALESCE(valor_inicial_original, valor_inicial),
            valor_final=COALESCE(valor_final_original, valor_final),
            status=COALESCE(status_original, 'OK'),
            valor_inicial_original=NULL,
            valor_final_original=NULL,
            status_original=NULL,
            cancelado_manual=0
        WHERE tipo=? AND numero=? AND UPPER(status)='CANCELADO MANUALMENTE'
        """,
        (tipo, numero),
    )
    alterados = cursor.rowcount
    conn.commit()
    conn.close()

    atualizar_dashboard()

    return alterados


def abrir_relatorio():
    nome = os.path.join(obter_pasta_saida_relatorios(), "Faturamento_AC.xlsx")
    if os.path.exists(nome):
        try:
            os.startfile(nome)
        except OSError as e:
            messagebox.showerror("Erro", f"Nao foi possivel abrir o relatorio: {e}")
    else:
        messagebox.showwarning("Relatorio", "Relatorio nao encontrado. Gere o relatorio primeiro.")


# ------------------------
# ATUALIZAR
# ------------------------

# ------------------------
# STATUS
# ------------------------

# ------------------------
# EXCEL
# ------------------------

def gerar_excel():
    conn = obter_conexao_banco()
    df = pd.read_sql_query("SELECT * FROM documentos", conn)
    conn.close()

    if df.empty:
        messagebox.showwarning("Excel", "Nenhum documento encontrado no banco.")
        return

    df["data_emissao"] = pd.to_datetime(df["data_emissao"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["data_emissao"])

    if df.empty:
        messagebox.showwarning("Excel", "Nao ha datas validas para gerar o relatorio.")
        return

    df["numero"] = pd.to_numeric(df["numero"], errors="coerce")
    df = df.dropna(subset=["numero"])

    if df.empty:
        messagebox.showwarning("Excel", "Nao ha numeros de documento validos para gerar o relatorio.")
        return

    df["numero"] = df["numero"].astype(int)

    try:
        data_inicial = ler_data_filtro(data_inicio_entry.get(), "Data inicial")
        data_final = ler_data_filtro(data_fim_entry.get(), "Data final")
    except ValueError as erro_data:
        messagebox.showwarning("Excel", str(erro_data))
        return

    if data_inicial > data_final:
        messagebox.showwarning("Excel", "A data inicial nao pode ser maior que a data final.")
        return

    # Converte competencia para data (primeiro dia do mês) para filtro
    def competencia_para_data(comp_str):
        try:
            partes = comp_str.lower().split("/")
            if len(partes) == 2:
                mes_nome = partes[0].strip()
                ano_str = partes[1].strip()
                ano = int(ano_str)
                mes_idx = MESES.index(mes_nome) + 1
                return datetime(ano, mes_idx, 1)
        except:
            pass
        return None

    df["data_competencia"] = df["competencia"].apply(competencia_para_data)
    df = df.dropna(subset=["data_competencia"])

    # Filtra por competência (período selecionado)
    df = df[(df["data_competencia"] >= data_inicial) & (df["data_competencia"] <= data_final)].copy()

    # Evita duplicidade de NF quando existir numero antigo (ex.: 20260092) e numero real (92).
    df["numero_original_num"] = pd.to_numeric(df.get("numero_original"), errors="coerce")
    df["chave_documento"] = df.apply(
        lambda r: (
            f"NF:{int(r['numero_original_num'])}"
            if str(r["tipo"]).upper() == "NF" and pd.notna(r["numero_original_num"])
            else f"{str(r['tipo']).upper()}:{int(r['numero'])}"
        ),
        axis=1,
    )
    df = df.sort_values(["id"]).drop_duplicates(subset=["chave_documento"], keep="last")

    if df.empty:
        messagebox.showwarning(
            "Excel",
            f"Nao ha documentos para o periodo selecionado ({data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}).",
        )
        return

    df = df.sort_values(["data_emissao", "numero"], ascending=[True, True])
    df["Concat"] = df["numero"].astype(str) + " " + df["tipo"]
    # Usa a data de competencia ja validada para evitar coluna "Mes Referencia" vazia no Excel.
    df["competencia_excel"] = df["data_competencia"]

    # Remove arquivo anterior se existir
    pasta_saida = obter_pasta_saida_relatorios()
    nome = os.path.join(pasta_saida, "Faturamento_AC.xlsx")
    for antigo in glob.glob(os.path.join(pasta_saida, "Faturamento_AC*.xlsx")):
        try:
            if os.path.exists(antigo):
                os.remove(antigo)
        except (OSError, PermissionError) as e:
            try:
                os.rename(antigo, antigo + ".bak")
            except:
                pass

    colunas_base = [
        "data_emissao",
        "competencia_excel",
        "numero",
        "numero_original_num",
        "tipo",
        "frete",
        "valor_inicial",
        "valor_final",
        "status",
    ]

    df_base = df[colunas_base].copy()

    def montar_df_relatorio(df_in, usar_numero_real_nf=False):
        dados = df_in.copy()
        if usar_numero_real_nf:
            dados["numero_doc"] = dados.apply(
                lambda r: int(r["numero_original_num"])
                if str(r["tipo"]).upper() == "NF" and pd.notna(r["numero_original_num"])
                else int(r["numero"]) if pd.notna(r["numero"]) else None,
                axis=1,
            )
        else:
            dados["numero_doc"] = dados["numero"]

        dados["numero_doc"] = pd.to_numeric(dados["numero_doc"], errors="coerce")
        dados = dados.dropna(subset=["numero_doc"]).copy()
        dados["numero_doc"] = dados["numero_doc"].astype(int)
        dados["Concat"] = dados["numero_doc"].astype(str) + " " + dados["tipo"].astype(str)

        dados = dados[
            [
                "data_emissao",
                "competencia_excel",
                "numero_doc",
                "tipo",
                "Concat",
                "frete",
                "valor_inicial",
                "valor_final",
                "status",
            ]
        ]

        dados.columns = [
            "Data Emissao",
            "Mes Referencia",
            "Numero Doc",
            "Tipo Doc",
            "Concat",
            "Frete",
            "Valor Inicial",
            "Valor Final",
            "Status",
        ]
        return dados

    df_relatorio_1 = montar_df_relatorio(df_base, usar_numero_real_nf=False)
    df_relatorio_2 = montar_df_relatorio(df_base, usar_numero_real_nf=True)

    try:
        with pd.ExcelWriter(nome, engine="openpyxl") as writer:
            nome_aba_1 = "Faturamento AC"
            nome_aba_2 = "Faturamento AC 2"
            df_relatorio_1.to_excel(writer, index=False, startcol=1, sheet_name=nome_aba_1)
            df_relatorio_2.to_excel(writer, index=False, startcol=1, sheet_name=nome_aba_2)

            def formatar_aba(nome_aba, df_aba):
                ws = writer.sheets[nome_aba]
                ws.sheet_view.showGridLines = False
                ws.freeze_panes = "B2"

                ultima_linha = len(df_aba) + 1
                linha_inicial = 1
                col_inicial = 2  # B
                col_final = 10   # J

                cor_cabecalho = PatternFill("solid", fgColor="1F4E78")
                fonte_cabecalho = Font(color="FFFFFF", bold=True)
                cor_cancelado = PatternFill("solid", fgColor="F8D7DA")
                cor_b_dados = PatternFill("solid", fgColor="E6E6E6")
                cor_f_dados = PatternFill("solid", fgColor="FFF2CC")
                borda_fina = Border(
                    left=Side(style="thin", color="D9D9D9"),
                    right=Side(style="thin", color="D9D9D9"),
                    top=Side(style="thin", color="D9D9D9"),
                    bottom=Side(style="thin", color="D9D9D9"),
                )

                for col in range(col_inicial, col_final + 1):
                    celula = ws.cell(row=1, column=col)
                    celula.fill = cor_cabecalho
                    celula.font = fonte_cabecalho
                    celula.alignment = Alignment(horizontal="center", vertical="center")

                for row in range(linha_inicial, ultima_linha + 1):
                    for col in range(col_inicial, col_final + 1):
                        ws.cell(row=row, column=col).border = borda_fina

                for row in range(2, ultima_linha + 1):
                    for col in range(col_inicial, col_final + 1):
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

                    ws[f"B{row}"].fill = cor_b_dados
                    ws[f"F{row}"].fill = cor_f_dados

                    status_valor = str(ws[f"J{row}"].value or "").upper()
                    if ("CANCELADO" in status_valor) or ("SUBSTITUIDO" in status_valor):
                        for col in range(col_inicial, col_final + 1):
                            ws.cell(row=row, column=col).fill = cor_cancelado

                    ws[f"B{row}"].number_format = "DD/MM/YYYY"
                    ws[f"C{row}"].number_format = '[$-pt-BR]mmmm/yyyy'
                    ws[f"D{row}"].number_format = "0"
                    ws[f"H{row}"].number_format = "R$ #,##0.00"
                    ws[f"I{row}"].number_format = "R$ #,##0.00"

                for col in range(col_inicial, col_final + 1):
                    letra_coluna = ws.cell(row=1, column=col).column_letter
                    maior = 0
                    for row in range(1, ultima_linha + 1):
                        valor = ws.cell(row=row, column=col).value
                        tamanho = len(str(valor)) if valor is not None else 0
                        if tamanho > maior:
                            maior = tamanho
                    ws.column_dimensions[letra_coluna].width = min(maior + 2, 45)

            formatar_aba(nome_aba_1, df_relatorio_1)
            formatar_aba(nome_aba_2, df_relatorio_2)

        messagebox.showinfo(
            "Excel",
            f"Relatorio gerado com {len(df_relatorio_1)} documento(s) no periodo {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}\n\nAbas: Faturamento AC e Faturamento AC 2\nArquivo: {nome}",
        )
    except Exception as e:
        messagebox.showerror("Excel", f"Erro ao gerar o relatorio: {e}")


# ------------------------
# INTERFACE
# ------------------------

ctk.set_appearance_mode("light")

preparar_arquivos_aplicacao()

lock_ok, pid_existente = adquirir_lock_instancia()
if not lock_ok:
    alertar_instancia_em_execucao(pid_existente)
    raise SystemExit(0)

atexit.register(liberar_lock_instancia)

app = ctk.CTk()

largura_tela = app.winfo_screenwidth()
altura_tela = app.winfo_screenheight()

largura = max(340, min(500, largura_tela - 60))
altura = max(500, min(720, altura_tela - 110))
largura = min(largura, largura_tela - 10)
altura = min(altura, altura_tela - 10)

centralizar_janela(app, largura, altura)
app.title("Sistema de Faturamento - Horizonte Logistica")
app.resizable(True, True)

aplicar_icone_aplicacao(app)


def _encerrar_aplicacao():
    liberar_lock_instancia()
    app.destroy()


app.protocol("WM_DELETE_WINDOW", _encerrar_aplicacao)

iniciar_banco()
carregar_pasta_saida_relatorios()

container_scroll = ctk.CTkScrollableFrame(app, corner_radius=0, fg_color="transparent")
container_scroll.pack(fill="both", expand=True)

if os.path.exists(LOGO_PATH):
    try:
        logo = ctk.CTkImage(Image.open(LOGO_PATH), size=(240, 128))
        ctk.CTkLabel(container_scroll, image=logo, text="").pack(pady=(8, 4))
    except Exception as e:
        print(f"Aviso: nao foi possivel carregar a logo - {e}")

main_frame = ctk.CTkFrame(container_scroll, corner_radius=10)
main_frame.pack(fill="x", padx=12, pady=(0, 8))

primeiro_dia_padrao, ultimo_dia_padrao = periodo_padrao_mes_atual()

filtro_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
filtro_frame.pack(fill="x", padx=10, pady=(6, 3))

ctk.CTkLabel(filtro_frame, text="Periodo de emissao (DD/MM/AAAA)").pack(anchor="center")

periodo_linha = ctk.CTkFrame(filtro_frame, fg_color="transparent")
periodo_linha.pack(anchor="center", pady=(2, 0))

data_inicio_entry = ctk.CTkEntry(periodo_linha, width=112)
data_inicio_entry.insert(0, primeiro_dia_padrao.strftime("%d/%m/%Y"))
data_inicio_entry.pack(side="left")
data_inicio_entry.bind("<Button-1>", lambda _e: abrir_seletor_data(data_inicio_entry))
data_inicio_entry.bind("<FocusOut>", solicitar_atualizacao_dashboard)
data_inicio_entry.bind("<Return>", solicitar_atualizacao_dashboard)

ctk.CTkLabel(periodo_linha, text="ate").pack(side="left", padx=6)

data_fim_entry = ctk.CTkEntry(periodo_linha, width=112)
data_fim_entry.insert(0, ultimo_dia_padrao.strftime("%d/%m/%Y"))
data_fim_entry.pack(side="left")
data_fim_entry.bind("<Button-1>", lambda _e: abrir_seletor_data(data_fim_entry))
data_fim_entry.bind("<FocusOut>", solicitar_atualizacao_dashboard)
data_fim_entry.bind("<Return>", solicitar_atualizacao_dashboard)


def abrir_dialogo_alterar_competencia():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Alterar competencia de documento")
    centralizar_janela(dialog, 620, 270)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Tipo de documento").grid(row=0, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Numero do documento").grid(row=0, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=260)
    tipo_combo.set("NF")
    tipo_combo.grid(row=1, column=0, sticky="ew", padx=6, pady=(0, 10))

    numero_entry = ctk.CTkEntry(form, width=260, placeholder_text="Ex.: 20260089 ou 2390")
    numero_entry.grid(row=1, column=1, sticky="ew", padx=6, pady=(0, 10))

    ctk.CTkLabel(form, text="Novo mes de competencia").grid(row=2, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Ano da competencia").grid(row=2, column=1, sticky="w", padx=6, pady=(0, 4))

    mes_combo = ctk.CTkComboBox(form, values=MESES, width=260)
    mes_combo.set(MESES[datetime.now().month - 1])
    mes_combo.grid(row=3, column=0, sticky="ew", padx=6)

    anos_competencia = [str(ano) for ano in range(datetime.now().year - 3, datetime.now().year + 5)]
    ano_combo = ctk.CTkComboBox(form, values=anos_competencia, width=260)
    ano_combo.set(str(datetime.now().year))
    ano_combo.grid(row=3, column=1, sticky="ew", padx=6)

    def salvar_alteracao():
        tipo = tipo_combo.get().strip().upper()
        mes_novo = mes_combo.get().strip().lower()
        ano_novo = ano_combo.get().strip()
        numero_texto = numero_entry.get().strip()

        if tipo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento invalido.")
            return

        if mes_novo not in MESES:
            messagebox.showwarning("Aviso", "Selecione um mes valido.")
            return

        if not numero_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe um numero de documento valido.")
            return

        if not ano_novo.isdigit():
            messagebox.showwarning("Aviso", "Informe um ano valido.")
            return

        numero = int(numero_texto)
        alterados = alterar_competencia_documento(tipo, numero, mes_novo, int(ano_novo))
        if alterados == 0:
            messagebox.showwarning("Aviso", "Documento nao encontrado para alterar competencia.")
            return

        messagebox.showinfo("Sucesso", "Competencia atualizada com sucesso.")
        dialog.destroy()

    ctk.CTkButton(form, text="Salvar alteracao", command=salvar_alteracao, width=200).grid(row=4, column=0, columnspan=2, pady=(14, 0))


def _abrir_dialogo_declarar_frete(rotulo_frete):
    dialog = ctk.CTkToplevel(app)
    dialog.title(rotulo_frete)
    centralizar_janela(dialog, 620, 220)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Tipo de documento").grid(row=0, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Numero do documento").grid(row=0, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=260)
    tipo_combo.set("NF")
    tipo_combo.grid(row=1, column=0, sticky="ew", padx=6)

    numero_entry = ctk.CTkEntry(form, width=260, placeholder_text="Ex.: 20260092 ou 2390")
    numero_entry.grid(row=1, column=1, sticky="ew", padx=6)

    desfazer_var = ctk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        form,
        text=f"Des{rotulo_frete.lower()}",
        variable=desfazer_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=2, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 0))

    def confirmar():
        tipo = tipo_combo.get().strip().upper()
        numero_texto = numero_entry.get().strip()

        if tipo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento invalido.")
            return
        if not numero_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe um numero de documento valido.")
            return

        if desfazer_var.get():
            # Desdeclarar - volta para FRANQUIA
            alterados = declarar_documento_frete(tipo, int(numero_texto), "FRANQUIA")
            if alterados == 0:
                messagebox.showwarning("Aviso", f"Documento nao encontrado para des{rotulo_frete.lower()}.")
                return
            messagebox.showinfo("Sucesso", f"Documento des{rotulo_frete.lower()} com sucesso.")
        else:
            # Declarar
            alterados = declarar_documento_frete(tipo, int(numero_texto), rotulo_frete.upper())
            if alterados == 0:
                messagebox.showwarning("Aviso", f"Documento nao encontrado para {rotulo_frete.lower()}.")
                return
            messagebox.showinfo("Sucesso", f"Documento declarado como {rotulo_frete}.")
        dialog.destroy()

    ctk.CTkButton(form, text="Confirmar", command=confirmar, width=220).grid(row=3, column=0, columnspan=2, pady=(14, 0))


def abrir_dialogo_declarar_intercompany():
    _abrir_dialogo_declarar_frete("Intercompany")


def abrir_dialogo_declarar_delta():
    _abrir_dialogo_declarar_frete("Delta")


def abrir_dialogo_cancelar_documento():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Cancelar documento")
    centralizar_janela(dialog, 620, 220)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Tipo de documento").grid(row=0, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Numero do documento").grid(row=0, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=260)
    tipo_combo.set("NF")
    tipo_combo.grid(row=1, column=0, sticky="ew", padx=6)

    numero_entry = ctk.CTkEntry(form, width=260, placeholder_text="Ex.: 20260089 ou 2390")
    numero_entry.grid(row=1, column=1, sticky="ew", padx=6)

    desfazer_var = ctk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        form,
        text="Desfazer cancelamento",
        variable=desfazer_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=2, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 0))

    def confirmar_cancelamento():
        tipo = tipo_combo.get().strip().upper()
        numero_texto = numero_entry.get().strip()

        if tipo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento invalido.")
            return

        if not numero_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe um numero de documento valido.")
            return

        numero = int(numero_texto)
        if desfazer_var.get():
            alterados = desfazer_cancelamento_documento(tipo, numero)
            if alterados == 0:
                messagebox.showwarning("Aviso", "Documento nao encontrado ou nao esta cancelado.")
                return
            messagebox.showinfo("Sucesso", "Cancelamento desfeito com sucesso.")
        else:
            alterados = cancelar_documento(tipo, numero)
            if alterados == 0:
                messagebox.showwarning("Aviso", "Documento nao encontrado.")
                return
            messagebox.showinfo("Sucesso", "Documento cancelado com sucesso.")
        dialog.destroy()

    ctk.CTkButton(form, text="Confirmar", command=confirmar_cancelamento, width=200).grid(row=3, column=0, columnspan=2, pady=(14, 0))





def abrir_dialogo_substituir_documento():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Substituir documento")
    centralizar_janela(dialog, 700, 320)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Documento antigo", font=ctk.CTkFont(weight="bold")).grid(
        row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 4)
    )
    ctk.CTkLabel(form, text="Tipo").grid(row=1, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Numero").grid(row=1, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_antigo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=320)
    tipo_antigo_combo.set("NF")
    tipo_antigo_combo.grid(row=2, column=0, sticky="ew", padx=6, pady=(0, 10))

    numero_antigo_entry = ctk.CTkEntry(form, width=320, placeholder_text="Ex.: 20260089 ou 2390")
    numero_antigo_entry.grid(row=2, column=1, sticky="ew", padx=6, pady=(0, 10))

    ctk.CTkLabel(form, text="Documento substituto", font=ctk.CTkFont(weight="bold")).grid(
        row=3, column=0, columnspan=2, sticky="w", padx=6, pady=(2, 4)
    )
    ctk.CTkLabel(form, text="Tipo").grid(row=4, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Numero").grid(row=4, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_novo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=320)
    tipo_novo_combo.set("NF")
    tipo_novo_combo.grid(row=5, column=0, sticky="ew", padx=6)

    numero_novo_entry = ctk.CTkEntry(form, width=320, placeholder_text="Ex.: 20260090 ou 2391")
    numero_novo_entry.grid(row=5, column=1, sticky="ew", padx=6)

    desfazer_var = ctk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        form,
        text="Desfazer substituicao",
        variable=desfazer_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=6, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 0))

    def confirmar_substituicao():
        tipo_antigo = tipo_antigo_combo.get().strip().upper()
        tipo_novo = tipo_novo_combo.get().strip().upper()
        numero_antigo_texto = numero_antigo_entry.get().strip()
        numero_novo_texto = numero_novo_entry.get().strip()

        if tipo_antigo not in {"NF", "CTE"} or tipo_novo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento invalido.")
            return

        if not numero_antigo_texto.isdigit() or not numero_novo_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe numeros de documento validos.")
            return

        numero_antigo = int(numero_antigo_texto)
        numero_novo = int(numero_novo_texto)

        if tipo_antigo == tipo_novo and numero_antigo == numero_novo:
            messagebox.showwarning("Aviso", "Documento antigo e substituto nao podem ser iguais.")
            return

        if desfazer_var.get():
            antigo_restaurado, novo_restaurado = desfazer_substituicao(
                tipo_antigo, numero_antigo, tipo_novo, numero_novo
            )
            if antigo_restaurado == 0:
                messagebox.showwarning("Aviso", "Documento antigo nao encontrado ou nao esta substituido.")
                return
            if novo_restaurado == 0:
                messagebox.showwarning("Aviso", "Documento substituto nao encontrado ou nao esta como substituto.")
                return
            messagebox.showinfo("Sucesso", "Substituicao desfeita com sucesso.")
        else:
            novo_alterado, antigo_alterado = registrar_substituicao(
                tipo_antigo, numero_antigo, tipo_novo, numero_novo
            )
            if novo_alterado == 0:
                messagebox.showwarning("Aviso", "Documento substituto nao encontrado.")
                return
            if antigo_alterado == 0:
                messagebox.showwarning("Aviso", "Documento antigo nao encontrado.")
                return
            messagebox.showinfo("Sucesso", "Substituicao registrada com sucesso.")
        dialog.destroy()

    ctk.CTkButton(form, text="Confirmar", command=confirmar_substituicao, width=220).grid(
        row=7, column=0, columnspan=2, pady=(14, 0)
    )


def abrir_relatorio_cancelados():
    try:
        data_inicial = ler_data_filtro(data_inicio_entry.get(), "Data inicial")
        data_final = ler_data_filtro(data_fim_entry.get(), "Data final")
    except ValueError as exc:
        messagebox.showwarning("Filtro de periodo", str(exc))
        return

    if data_inicial > data_final:
        messagebox.showwarning("Filtro de periodo", "Data inicial nao pode ser maior que a data final.")
        return

    conn = obter_conexao_banco()
    df = pd.read_sql_query(
        """
        SELECT tipo, numero, data_emissao, valor_final, status
        FROM documentos
        WHERE UPPER(status) LIKE '%CANCELADO%'
        """,
        conn,
    )
    conn.close()

    df["data_emissao_dt"] = pd.to_datetime(df["data_emissao"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["data_emissao_dt"])
    df = df[(df["data_emissao_dt"] >= data_inicial) & (df["data_emissao_dt"] <= data_final)]

    if df.empty:
        messagebox.showinfo(
            "Relatorio",
            (
                "Nenhum documento cancelado encontrado no periodo selecionado.\n\n"
                f"Periodo: {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}"
            ),
        )
        return

    df = df.sort_values(["data_emissao_dt", "numero"], ascending=[True, True])

    janela = ctk.CTkToplevel(app)
    janela.title("Relatorio de documentos cancelados")
    centralizar_janela(janela, 860, 520)
    janela.grab_set()

    ctk.CTkLabel(
        janela,
        text=(
            "Documentos Cancelados\n"
            f"{data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}"
        ),
        font=ctk.CTkFont(size=16, weight="bold"),
    ).pack(pady=(12, 8))

    container = ctk.CTkFrame(janela)
    container.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    style = ttk.Style()
    style.configure("Cancelados.Treeview", rowheight=28, borderwidth=1, relief="solid")
    style.configure(
        "Cancelados.Treeview.Heading",
        font=("Segoe UI", 10, "bold"),
        relief="solid",
        borderwidth=1,
    )

    colunas = ("tipo", "numero", "data_emissao", "valor", "status")
    tabela = ttk.Treeview(container, columns=colunas, show="headings", style="Cancelados.Treeview")

    tabela.heading("tipo", text="Tipo Doc")
    tabela.heading("numero", text="Numero Doc")
    tabela.heading("data_emissao", text="Data Emissao")
    tabela.heading("valor", text="Valor")
    tabela.heading("status", text="Status")

    tabela.column("tipo", width=110, anchor="center")
    tabela.column("numero", width=170, anchor="center")
    tabela.column("data_emissao", width=170, anchor="center")
    tabela.column("valor", width=180, anchor="e")
    tabela.column("status", width=240, anchor="center")

    scroll_y = ttk.Scrollbar(container, orient="vertical", command=tabela.yview)
    scroll_x = ttk.Scrollbar(container, orient="horizontal", command=tabela.xview)
    tabela.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    tabela.grid(row=0, column=0, sticky="nsew")
    scroll_y.grid(row=0, column=1, sticky="ns")
    scroll_x.grid(row=1, column=0, sticky="ew")

    container.grid_rowconfigure(0, weight=1)
    container.grid_columnconfigure(0, weight=1)

    for _, linha in df.iterrows():
        data_txt = (
            linha["data_emissao_dt"].strftime("%d/%m/%Y")
            if pd.notna(linha["data_emissao_dt"])
            else str(linha["data_emissao"])
        )
        valor = float(linha["valor_final"]) if pd.notna(linha["valor_final"]) else 0.0
        tabela.insert(
            "",
            "end",
            values=(
                str(linha["tipo"]),
                int(linha["numero"]),
                data_txt,
                formatar_moeda_brl(valor),
                str(linha["status"]),
            ),
        )




def abrir_busca_documentos():

    dialog = ctk.CTkToplevel(app)
    dialog.title("Buscar documentos")
    centralizar_janela(dialog, 750, 500)
    dialog.grab_set()

    ctk.CTkLabel(
        dialog,
        text="Buscar documento",
        font=ctk.CTkFont(size=16, weight="bold"),
    ).pack(pady=(12, 8))

    busca_entry = ctk.CTkEntry(dialog, width=300, placeholder_text="Digite o numero do documento")
    busca_entry.pack(pady=(0, 10))

    container = ctk.CTkFrame(dialog)
    container.pack(fill="both", expand=True, padx=10, pady=10)

    tabela = ttk.Treeview(
        container,
        columns=("tipo","numero","data","valor","status"),
        show="headings"
    )

    tabela.heading("tipo", text="Tipo")
    tabela.heading("numero", text="Numero")
    tabela.heading("data", text="Data")
    tabela.heading("valor", text="Valor")
    tabela.heading("status", text="Status")

    tabela.column("tipo", width=80, anchor="center")
    tabela.column("numero", width=120, anchor="center")
    tabela.column("data", width=120, anchor="center")
    tabela.column("valor", width=120, anchor="e")
    tabela.column("status", width=200, anchor="center")

    scroll = ttk.Scrollbar(container, orient="vertical", command=tabela.yview)
    tabela.configure(yscrollcommand=scroll.set)

    tabela.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    def buscar():

        termo = busca_entry.get().strip()

        for item in tabela.get_children():
            tabela.delete(item)

        conn = obter_conexao_banco()

        df = pd.read_sql_query(
            """
            SELECT tipo, numero, data_emissao, valor_final, status
            FROM documentos
            WHERE numero LIKE ?
            ORDER BY data_emissao
            """,
            conn,
            params=(f"%{termo}%",),
        )

        conn.close()

        for _, row in df.iterrows():

            tabela.insert(
                "",
                "end",
                values=(
                    row["tipo"],
                    row["numero"],
                    row["data_emissao"],
                    formatar_moeda_brl(row["valor_final"]),
                    row["status"],
                ),
            )

    ctk.CTkButton(dialog, text="Buscar", command=buscar, width=200).pack(pady=6)


def abrir_grafico_faturamento():
    try:
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.ticker import FuncFormatter
    except Exception as exc:
        messagebox.showerror("Grafico", f"Falha ao carregar bibliotecas do grafico.\n\n{exc}")
        return

    conn = obter_conexao_banco()

    df = pd.read_sql_query(
        """
        SELECT data_emissao, valor_final
        FROM documentos
        WHERE status NOT LIKE '%CANCELADO%'
        """,
        conn,
    )

    conn.close()

    if df.empty:
        messagebox.showwarning("Grafico", "Nao ha dados para gerar grafico.")
        return

    df["data_emissao"] = pd.to_datetime(df["data_emissao"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["data_emissao"])
    df["valor_final"] = pd.to_numeric(df["valor_final"], errors="coerce").fillna(0)

    try:
        data_inicial = datetime.strptime(data_inicio_entry.get().strip(), "%d/%m/%Y")
        data_final = datetime.strptime(data_fim_entry.get().strip(), "%d/%m/%Y")
        if data_inicial <= data_final:
            df = df[(df["data_emissao"] >= data_inicial) & (df["data_emissao"] <= data_final)]
    except ValueError:
        pass

    if df.empty:
        messagebox.showwarning("Grafico", "Nao ha dados no periodo selecionado para gerar grafico.")
        return

    df["mes"] = df["data_emissao"].dt.to_period("M")

    resumo = df.groupby("mes")["valor_final"].sum().reset_index().sort_values("mes")
    resumo["mes_dt"] = resumo["mes"].dt.to_timestamp()
    resumo["mes_label"] = resumo["mes_dt"].dt.strftime("%m/%Y")
    resumo["idx"] = range(len(resumo))

    if resumo.empty:
        messagebox.showwarning("Grafico", "Nao ha dados para gerar grafico.")
        return

    janela = ctk.CTkToplevel(app)
    janela.title("Faturamento mensal")
    centralizar_janela(janela, 980, 560)
    janela.grab_set()

    fig, ax = plt.subplots(figsize=(9.0, 4.8), dpi=110)
    fig.patch.set_facecolor("#F4F6FA")
    ax.set_facecolor("#FFFFFF")

    valores = resumo["valor_final"].tolist()
    max_valor = max(valores) if valores else 0
    cores = ["#5AA9E6" if v < max_valor else "#1D5D9B" for v in valores]

    barras = ax.bar(
        resumo["idx"],
        valores,
        color=cores,
        edgecolor="#1D4F80",
        linewidth=0.7,
        width=0.62,
        zorder=2,
    )

    ax.plot(
        resumo["idx"],
        valores,
        color="#0B3C6F",
        linewidth=1.8,
        marker="o",
        markersize=4.5,
        markerfacecolor="#FFFFFF",
        markeredgewidth=1.2,
        zorder=3,
    )

    def _fmt_brl_eixo(v, _pos):
        return f"R$ {v:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _fmt_brl_curto(v):
        if v >= 1_000_000:
            return f"R$ {v/1_000_000:.1f} mi".replace(".", ",")
        if v >= 1_000:
            return f"R$ {v/1_000:.1f} mil".replace(".", ",")
        return _fmt_brl_eixo(v, None)

    ax.set_title("Faturamento Mensal", fontsize=15, fontweight="bold", color="#0B3C6F", pad=12)
    ax.set_ylabel("Valor faturado", fontsize=10, color="#2E3A4E")
    ax.set_xlabel("Mes de emissao", fontsize=10, color="#2E3A4E", labelpad=8)
    ax.yaxis.set_major_formatter(FuncFormatter(_fmt_brl_eixo))
    ax.grid(axis="y", linestyle="--", linewidth=0.8, alpha=0.25, zorder=1)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#BFC9D9")
    ax.spines["bottom"].set_color("#BFC9D9")
    ax.tick_params(axis="x", colors="#253245")
    ax.tick_params(axis="y", colors="#253245")
    ax.set_xticks(resumo["idx"])
    ax.set_xticklabels(resumo["mes_label"], rotation=35, ha="right")

    for barra in barras:
        valor = barra.get_height()
        ax.text(
            barra.get_x() + barra.get_width() / 2,
            valor + (max_valor * 0.012 if max_valor else 0),
            _fmt_brl_curto(valor),
            ha="center",
            va="bottom",
            fontsize=8.5,
            color="#1E2A38",
        )

    periodo_ini = resumo["mes_dt"].min().strftime("%m/%Y")
    periodo_fim = resumo["mes_dt"].max().strftime("%m/%Y")
    resumo_txt = f"Periodo exibido: {periodo_ini} a {periodo_fim}"
    ax.text(
        0.01,
        1.02,
        resumo_txt,
        transform=ax.transAxes,
        fontsize=9,
        color="#5A6575",
        ha="left",
        va="bottom",
    )

    fig.tight_layout(pad=1.4)

    canvas = FigureCanvasTkAgg(fig, master=janela)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)


def atualizar_dashboard():
    def atualizar_resumo(total_inicial=0.0, total_final=0.0, total_docs=0, total_nf=0, total_cte=0, total_cancelados=0):
        total_inicial_valor_label.configure(text=formatar_moeda_brl(total_inicial))
        total_final_valor_label.configure(text=formatar_moeda_brl(total_final))
        total_documentos_label.configure(text=f"{total_docs}")
        nf_label.configure(text=f"{total_nf}")
        cte_label.configure(text=f"{total_cte}")
        cancelados_label.configure(text=f"{total_cancelados}")

        diferenca = total_inicial - total_final
        if diferenca > 0:
            diferenca_label.configure(
                text=f"Diferenca referente a Impostos de NFS-e: {formatar_moeda_brl(diferenca)}",
                text_color="#1E6A31",
            )
        elif diferenca < 0:
            diferenca_label.configure(
                text=f"Diferenca referente a Impostos de NFS-e: +{formatar_moeda_brl(abs(diferenca))}",
                text_color="#9F1D2B",
            )
        else:
            diferenca_label.configure(text="Diferenca referente a Impostos de NFS-e: R$ 0,00", text_color="#3A4654")

        if "cancelados_card" in globals():
            if total_cancelados > 0:
                cancelados_card.configure(fg_color="#F9E1E3")
                cancelados_label.configure(text_color="#9F1D2B")
            else:
                cancelados_card.configure(fg_color="#E5F3E7")
                cancelados_label.configure(text_color="#1F6E2F")

    # O dashboard agora segue a mesma regra do Excel:
    # filtro por competencia + deduplicacao de NF pela chave_documento.
    try:
        data_inicial = datetime.strptime(data_inicio_entry.get().strip(), "%d/%m/%Y")
        data_final = datetime.strptime(data_fim_entry.get().strip(), "%d/%m/%Y")
        if data_inicial > data_final:
            atualizar_resumo()
            return
    except ValueError:
        # Enquanto o usuario edita parcialmente a data, nao recalcula o card.
        return

    conn = obter_conexao_banco()
    df = pd.read_sql_query(
        """
        SELECT id, tipo, numero, numero_original, valor_inicial, valor_final, status, competencia
        FROM documentos
        """,
        conn,
    )
    conn.close()

    if df.empty:
        atualizar_resumo()
        return

    def competencia_para_data(comp_str):
        try:
            partes = str(comp_str).lower().split("/")
            if len(partes) == 2:
                mes_nome = normalizar_texto(partes[0].strip()).lower()
                ano = int(partes[1].strip())
                mes_idx = MESES.index(mes_nome) + 1
                return datetime(ano, mes_idx, 1)
        except Exception:
            return None
        return None

    df["data_competencia"] = df["competencia"].apply(competencia_para_data)
    df = df.dropna(subset=["data_competencia"])
    df = df[(df["data_competencia"] >= data_inicial) & (df["data_competencia"] <= data_final)].copy()

    if df.empty:
        atualizar_resumo()
        return

    df["numero"] = pd.to_numeric(df["numero"], errors="coerce")
    df["numero_original_num"] = pd.to_numeric(df.get("numero_original"), errors="coerce")
    df["valor_inicial"] = pd.to_numeric(df["valor_inicial"], errors="coerce")
    df["valor_final"] = pd.to_numeric(df["valor_final"], errors="coerce")
    df["tipo"] = df["tipo"].astype(str).str.upper()
    df["status"] = df["status"].astype(str)

    df["chave_documento"] = df.apply(
        lambda r: (
            f"NF:{int(r['numero_original_num'])}"
            if r["tipo"] == "NF" and pd.notna(r["numero_original_num"])
            else f"{r['tipo']}:{int(r['numero'])}" if pd.notna(r["numero"]) else f"{r['tipo']}:SEMNUM"
        ),
        axis=1,
    )
    df = df.sort_values(["id"]).drop_duplicates(subset=["chave_documento"], keep="last")

    total_inicial = df["valor_inicial"].fillna(0).sum()
    total_final = df["valor_final"].fillna(0).sum()
    total_docs = len(df)
    total_nf = int((df["tipo"] == "NF").sum())
    total_cte = int((df["tipo"] == "CTE").sum())
    total_cancelados = int(df["status"].str.upper().str.contains("CANCELADO", na=False).sum())

    atualizar_resumo(
        float(total_inicial),
        float(total_final),
        int(total_docs),
        total_nf,
        total_cte,
        total_cancelados,
    )


botoes_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
botoes_frame.pack(fill="x", padx=9, pady=(8, 3))
botoes_frame.grid_columnconfigure(0, weight=1)
botoes_frame.grid_columnconfigure(1, weight=1)
botoes_frame.grid_columnconfigure(2, weight=1)

acoes_principais = [
    ("Selecionar Relatorio", selecionar_relatorio),
    ("Importar Relatorio", importar_relatorio_ui),
    ("Gerar Faturamento", gerar_excel),
    ("Abrir Relatorio", abrir_relatorio),
]

acoes_menu_relatorios = [
    ("Pasta do relatorio", selecionar_pasta_saida_relatorios),
    ("Relatorio Cancelados", abrir_relatorio_cancelados),
    ("Buscar documento", abrir_busca_documentos),
    ("Grafico faturamento", abrir_grafico_faturamento),
]

acoes_menu_alteracoes = [
    ("Alterar competencia", abrir_dialogo_alterar_competencia),
    ("Substituir documento", abrir_dialogo_substituir_documento),
    ("Cancelar documento", abrir_dialogo_cancelar_documento),
    ("Declarar intercompany", abrir_dialogo_declarar_intercompany),
    ("Declarar delta", abrir_dialogo_declarar_delta),
]

def abrir_menu_dropdown(botao, opcoes):
    menu = tk.Menu(app, tearoff=0)
    for titulo, comando in opcoes:
        menu.add_command(label=titulo, command=comando)
    try:
        menu.tk_popup(botao.winfo_rootx(), botao.winfo_rooty() + botao.winfo_height())
    finally:
        menu.grab_release()

menus_dropdown = [
    ("Principal", acoes_principais, {}),
    ("Relatorios", acoes_menu_relatorios, {}),
    (
        "Alteracoes",
        acoes_menu_alteracoes,
        {},
    ),
]

for idx_menu, (titulo_menu, opcoes_menu, estilo) in enumerate(menus_dropdown):
    linha = idx_menu // 3
    coluna = idx_menu % 3
    botao_menu = ctk.CTkButton(
        botoes_frame,
        text=titulo_menu,
        width=140,
        **estilo,
    )
    botao_menu.configure(command=lambda b=botao_menu, op=opcoes_menu: abrir_menu_dropdown(b, op))
    botao_menu.grid(row=linha, column=coluna, padx=3, pady=4, sticky="ew")

pasta_saida_label = ctk.CTkLabel(main_frame, text="", anchor="w", wraplength=420, justify="left")
pasta_saida_label.pack(fill="x", padx=10, pady=(2, 1))
atualizar_label_pasta_saida()

pasta_label = ctk.CTkLabel(main_frame, text="Nenhum relatorio selecionado", anchor="w", wraplength=420, justify="left")
pasta_label.pack(fill="x", padx=10, pady=(3, 2))

progress = ctk.CTkProgressBar(main_frame, width=420)
progress.pack(pady=(4, 6))
progress.set(0)

status_label = ctk.CTkLabel(main_frame, text="Relatorio:0 pagina(s) | Docs:0")
status_label.pack(pady=(0, 7))

dashboard_frame = ctk.CTkFrame(main_frame, corner_radius=8, fg_color="#C9C9C9")
dashboard_frame.pack(fill="x", padx=10, pady=(1, 6))

linha1 = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
linha1.pack(fill="x", padx=7, pady=(5, 3))
linha1.grid_columnconfigure(0, weight=1)
linha1.grid_columnconfigure(1, weight=1)

ctk.CTkLabel(
    linha1,
    text="Resumo do periodo",
    font=ctk.CTkFont(size=12, weight="bold"),
    text_color="#24313F",
).grid(row=0, column=0, columnspan=2, pady=(0, 4))

total_inicial_card = ctk.CTkFrame(linha1, fg_color="#EAF3FC", corner_radius=7)
total_inicial_card.grid(row=1, column=0, padx=(0, 3), pady=(0, 3), sticky="nsew")
ctk.CTkLabel(total_inicial_card, text="Valor inicial", font=ctk.CTkFont(size=10, weight="bold"), text_color="#1C4E80").pack(pady=(4, 0))
total_inicial_valor_label = ctk.CTkLabel(total_inicial_card, text="R$ 0,00", font=ctk.CTkFont(size=13, weight="bold"), text_color="#16395E")
total_inicial_valor_label.pack(pady=(0, 4))

total_final_card = ctk.CTkFrame(linha1, fg_color="#EAF7ED", corner_radius=7)
total_final_card.grid(row=1, column=1, padx=(3, 0), pady=(0, 3), sticky="nsew")
ctk.CTkLabel(total_final_card, text="Valor final", font=ctk.CTkFont(size=10, weight="bold"), text_color="#1D6A32").pack(pady=(4, 0))
total_final_valor_label = ctk.CTkLabel(total_final_card, text="R$ 0,00", font=ctk.CTkFont(size=13, weight="bold"), text_color="#165726")
total_final_valor_label.pack(pady=(0, 4))

diferenca_label = ctk.CTkLabel(linha1, text="Diferenca: R$ 0,00", font=ctk.CTkFont(size=10, weight="bold"), text_color="#3A4654")
diferenca_label.grid(row=2, column=0, columnspan=2, pady=(0, 2))

linha2 = ctk.CTkFrame(dashboard_frame, fg_color="transparent")
linha2.pack(fill="x", padx=7, pady=(1, 6))
linha2.grid_columnconfigure(0, weight=1)
linha2.grid_columnconfigure(1, weight=1)

total_documentos_card = ctk.CTkFrame(linha2, fg_color="#E8F1FB", corner_radius=7)
total_documentos_card.grid(row=0, column=0, padx=3, pady=3, sticky="nsew")
ctk.CTkLabel(total_documentos_card, text="Total de documentos", font=ctk.CTkFont(size=10, weight="bold"), text_color="#1E4A7A").pack(pady=(4, 0))
total_documentos_label = ctk.CTkLabel(total_documentos_card, text="0", font=ctk.CTkFont(size=16, weight="bold"), text_color="#16395E")
total_documentos_label.pack(pady=(0, 4))

nf_card = ctk.CTkFrame(linha2, fg_color="#EAF7ED", corner_radius=7)
nf_card.grid(row=0, column=1, padx=3, pady=3, sticky="nsew")
ctk.CTkLabel(nf_card, text="NF", font=ctk.CTkFont(size=10, weight="bold"), text_color="#1F6E2F").pack(pady=(4, 0))
nf_label = ctk.CTkLabel(nf_card, text="0", font=ctk.CTkFont(size=16, weight="bold"), text_color="#185725")
nf_label.pack(pady=(0, 4))

cte_card = ctk.CTkFrame(linha2, fg_color="#EDECF9", corner_radius=7)
cte_card.grid(row=1, column=0, padx=3, pady=3, sticky="nsew")
ctk.CTkLabel(cte_card, text="CTE", font=ctk.CTkFont(size=10, weight="bold"), text_color="#39328A").pack(pady=(4, 0))
cte_label = ctk.CTkLabel(cte_card, text="0", font=ctk.CTkFont(size=16, weight="bold"), text_color="#2C266C")
cte_label.pack(pady=(0, 4))

cancelados_card = ctk.CTkFrame(linha2, fg_color="#E5F3E7", corner_radius=7)
cancelados_card.grid(row=1, column=1, padx=3, pady=3, sticky="nsew")
ctk.CTkLabel(cancelados_card, text="Cancelados", font=ctk.CTkFont(size=10, weight="bold"), text_color="#1F6E2F").pack(pady=(4, 0))
cancelados_label = ctk.CTkLabel(cancelados_card, text="0", font=ctk.CTkFont(size=16, weight="bold"), text_color="#1F6E2F")
cancelados_label.pack(pady=(0, 4))

# Ajusta a janela ao conteudo para reduzir area vazia inicial.
app.update_idletasks()
largura_ideal = min(max(420, app.winfo_reqwidth()), largura_tela - 10)
altura_ideal = min(max(500, app.winfo_reqheight()), altura_tela - 10)
centralizar_janela(app, largura_ideal, altura_ideal)

app.after(80, atualizar_dashboard)

try:
    app.mainloop()
finally:
    liberar_lock_instancia()


