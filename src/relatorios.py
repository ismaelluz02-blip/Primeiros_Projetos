"""
Funções puras de filtragem e montagem de DataFrames para relatórios.

Sem dependências de UI. Os orquestradores com messagebox/filedialog
(gerar_excel, exportar_relatorio_filtrado, abrir_relatorio, abrir_relatorio_filtrado)
permanecem em sistema_faturamento.py.
"""

from datetime import datetime

import pandas as pd

from src.banco import obter_conexao_banco
from src.logger import get_logger
from src.utils import MESES, _numero_documento_exibicao, _chave_documento_compativel

logger = get_logger(__name__)


# ─────────────────────────────────────────────
#  Filtragem de documentos por período
# ─────────────────────────────────────────────

def _obter_dataframe_relatorio_filtrado(data_inicial, data_final, docs_df_base=None):
    if isinstance(docs_df_base, pd.DataFrame):
        df = docs_df_base.copy()
    else:
        conn = obter_conexao_banco()
        df = pd.read_sql_query("SELECT * FROM documentos", conn)
        conn.close()

    if df.empty:
        return pd.DataFrame(), "Nenhum documento encontrado no banco."

    df["data_emissao"] = pd.to_datetime(df["data_emissao"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["data_emissao"])
    if df.empty:
        return pd.DataFrame(), "Não há datas válidas para gerar o relatório."

    df["numero"] = pd.to_numeric(df["numero"], errors="coerce")
    df = df.dropna(subset=["numero"])
    if df.empty:
        return pd.DataFrame(), "Não há números de documento válidos para gerar o relatório."
    df["numero"] = df["numero"].astype(int)

    def competencia_para_data(comp_str):
        try:
            partes = str(comp_str).lower().split("/")
            if len(partes) == 2:
                mes_nome = partes[0].strip()
                ano_str = partes[1].strip()
                ano = int(ano_str)
                mes_idx = MESES.index(mes_nome) + 1
                return datetime(ano, mes_idx, 1)
        except Exception:
            pass
        return None

    df["data_competencia"] = df["competencia"].apply(competencia_para_data)
    df = df.dropna(subset=["data_competencia"])
    df = df[(df["data_competencia"] >= data_inicial) & (df["data_competencia"] <= data_final)].copy()

    if df.empty:
        return pd.DataFrame(), ""

    df["numero_original_num"] = pd.to_numeric(df.get("numero_original"), errors="coerce")
    df["numero_exibicao"] = df.apply(
        lambda r: _numero_documento_exibicao(
            r["tipo"], r["numero"], r.get("numero_original", ""), r.get("data_emissao", "")
        ),
        axis=1,
    )
    df["chave_documento"] = df.apply(
        lambda r: _chave_documento_compativel(r["tipo"], r["numero"], r.get("numero_original", "")),
        axis=1,
    )
    df = df.sort_values(["id"]).drop_duplicates(subset=["chave_documento"], keep="last")
    return df, ""


# ─────────────────────────────────────────────
#  Montagem do DataFrame de exportação
# ─────────────────────────────────────────────

def _montar_dataframe_exportacao_periodo(df_filtrado):
    if df_filtrado is None or df_filtrado.empty:
        return pd.DataFrame()

    dados = df_filtrado.copy()
    dados["numero_doc"] = dados.apply(
        lambda r: _numero_documento_exibicao(
            r["tipo"], r["numero"], r.get("numero_original", ""), r.get("data_emissao", "")
        ),
        axis=1,
    )
    dados["numero_doc"] = dados["numero_doc"].astype(str).str.strip()
    dados = dados[dados["numero_doc"] != ""].copy()
    if dados.empty:
        return pd.DataFrame()

    dados["numero_doc_ordem"] = pd.to_numeric(dados["numero_doc"], errors="coerce")
    # Evita alerta do Excel "numero armazenado como texto" para documentos numericos.
    dados["numero_doc"] = dados.apply(
        lambda r: int(r["numero_doc_ordem"]) if pd.notna(r["numero_doc_ordem"]) else r["numero_doc"],
        axis=1,
    )
    dados["numero_ordem"] = pd.to_numeric(dados["numero"], errors="coerce")
    dados["numero_doc_ordem"] = dados["numero_doc_ordem"].fillna(dados["numero_ordem"])
    dados["tipo"] = dados["tipo"].astype(str).str.upper()
    dados["frete"] = dados["frete"].astype(str)
    dados["status"] = dados["status"].astype(str)
    dados["mes_referencia"] = pd.to_datetime(dados["data_competencia"], errors="coerce")
    dados = dados.sort_values(
        ["data_emissao", "numero_doc_ordem", "numero_doc"], ascending=[True, True, True]
    )

    export_df = dados[
        [
            "data_emissao",
            "mes_referencia",
            "numero_doc",
            "tipo",
            "frete",
            "valor_inicial",
            "valor_final",
            "status",
        ]
    ].copy()
    export_df.columns = [
        "Data Emissao",
        "Mes Referencia",
        "Numero Doc",
        "Tipo Doc",
        "Frete",
        "Valor Inicial",
        "Valor Final",
        "Status",
    ]
    return export_df

# ─────────────────────────────────────────────
#  Escrita do Excel de faturamento
# ─────────────────────────────────────────────

def _montar_df_relatorio_excel(df_in):
    """Transforma o df filtrado no formato de colunas para exportação Excel."""
    import pandas as pd
    dados = df_in.copy()
    dados["numero_doc"] = dados.apply(
        lambda r: _numero_documento_exibicao(
            r["tipo"], r["numero"], r.get("numero_original", ""), r.get("data_emissao", "")
        ),
        axis=1,
    )
    dados["numero_doc"] = dados["numero_doc"].astype(str).str.strip()
    dados = dados[dados["numero_doc"] != ""].copy()
    dados["numero_doc_ordem"] = pd.to_numeric(dados["numero_doc"], errors="coerce")
    dados["numero_doc"] = dados.apply(
        lambda r: int(r["numero_doc_ordem"]) if pd.notna(r["numero_doc_ordem"]) else r["numero_doc"],
        axis=1,
    )
    dados["numero_ordem"] = pd.to_numeric(dados["numero"], errors="coerce")
    dados["numero_doc_ordem"] = dados["numero_doc_ordem"].fillna(dados["numero_ordem"])
    dados["Concat"] = dados["numero_doc"].astype(str) + " " + dados["tipo"].astype(str)
    dados = dados.sort_values(
        ["data_emissao", "numero_doc_ordem", "numero_doc"], ascending=[True, True, True]
    )
    dados = dados[[
        "data_emissao", "competencia_excel", "numero_doc",
        "tipo", "Concat", "frete", "valor_inicial", "valor_final", "status",
    ]]
    dados.columns = [
        "Data Emissao", "Mes Referencia", "Numero Doc",
        "Tipo Doc", "Concat", "Frete", "Valor Inicial", "Valor Final", "Status",
    ]
    return dados


def escrever_excel_faturamento(df, nome_arquivo):
    """
    Recebe o df já filtrado (com data_competencia/competencia_excel) e escreve
    o arquivo xlsx de faturamento com duas abas formatadas.
    Retorna {"ok": True/False, "total_documentos": int, "arquivo": str, "erro": str}.
    """
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    try:
        df_relatorio = _montar_df_relatorio_excel(df)

        def _formatar_aba(ws, df_aba):
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "B2"

            ultima_linha = len(df_aba) + 1
            col_inicial, col_final = 2, 10

            cor_cabecalho = PatternFill("solid", fgColor="1F4E78")
            fonte_cabecalho = Font(color="FFFFFF", bold=True)
            cor_cancelado = PatternFill("solid", fgColor="F8D7DA")
            cor_b_dados = PatternFill("solid", fgColor="E6E6E6")
            cor_f_dados = PatternFill("solid", fgColor="FFF2CC")
            borda_fina = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )

            for col in range(col_inicial, col_final + 1):
                celula = ws.cell(row=1, column=col)
                celula.fill = cor_cabecalho
                celula.font = fonte_cabecalho
                celula.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(1, ultima_linha + 1):
                for col in range(col_inicial, col_final + 1):
                    ws.cell(row=row, column=col).border = borda_fina

            for row in range(2, ultima_linha + 1):
                for col in range(col_inicial, col_final + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                ws[f"B{row}"].fill = cor_b_dados
                ws[f"F{row}"].fill = cor_f_dados
                ws[f"G{row}"].alignment = Alignment(horizontal="left", vertical="center")

                status_valor = str(ws[f"J{row}"].value or "").upper()
                if "CANCELADO" in status_valor or "SUBSTITUIDO" in status_valor:
                    for col in range(col_inicial, col_final + 1):
                        ws.cell(row=row, column=col).fill = cor_cancelado

                ws[f"B{row}"].number_format = "DD/MM/YYYY"
                ws[f"C{row}"].number_format = '[$-pt-BR]mmmm/yyyy'
                ws[f"D{row}"].number_format = "0"
                ws[f"H{row}"].number_format = "R$ #,##0.00"
                ws[f"I{row}"].number_format = "R$ #,##0.00"

            for col in range(col_inicial, col_final + 1):
                letra = ws.cell(row=1, column=col).column_letter
                maior = max(
                    (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ultima_linha + 1)),
                    default=0,
                )
                ws.column_dimensions[letra].width = min(maior + 2, 45)
            ws.column_dimensions["G"].width = max(float(ws.column_dimensions["G"].width or 0), 18.0)

        with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
            df_relatorio.to_excel(writer, index=False, startcol=1, sheet_name="Faturamento AC")
            df_relatorio.to_excel(writer, index=False, startcol=1, sheet_name="Faturamento AC 2")
            _formatar_aba(writer.sheets["Faturamento AC"], df_relatorio)
            _formatar_aba(writer.sheets["Faturamento AC 2"], df_relatorio)

        return {"ok": True, "total_documentos": int(len(df_relatorio)), "arquivo": nome_arquivo, "erro": ""}

    except Exception as exc:
        logger.error("escrever_faturamento falhou para %r: %s", nome_arquivo, exc, exc_info=True)
        return {"ok": False, "total_documentos": 0, "arquivo": nome_arquivo, "erro": str(exc)}
