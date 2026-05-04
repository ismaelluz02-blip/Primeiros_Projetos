import os
import glob
import re
import atexit
import ctypes
import sqlite3
import shutil
import sys
import unicodedata
import calendar
import time
import json
import getpass
import socket
from calendar import monthrange
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

import customtkinter as ctk
import fitz
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageFilter

from src.utils import (
    MESES,
    valor_brasileiro,
    formatar_moeda_brl,
    formatar_moeda_brl_exata,
    parse_valor_monetario,
    normalizar_texto,
    _numero_para_texto,
    _extrair_ano_data_emissao,
    _normalizar_numero_original_nf,
    _coletar_numero_original_para_match,
    _numero_documento_exibicao,
    _chave_documento_compativel,
    competencia_por_data,
    periodo_padrao_mes_atual,
    obter_periodo_padrao_dashboard,
    obter_periodo_padrao_relatorios,
    ler_data_filtro,
    _obter_periodo_por_entries,
    _normalizar_hex_cor,
    _hex_para_rgb,
    _rgb_para_hex,
    _interpolar_cor,
    _competencia_para_data,
)

import src.config as _cfg
from src.medicao.auditor import run_audit as _medicao_run_audit
from src.medicao.report import generate_report as _medicao_generate_report
from src.medicao.scaffold import create_competencia_structure as _medicao_scaffold
from src.medicao.organizer_dialog import OrganizerDialog as _MedicaoOrganizerDialog
from src.banco import (
    obter_conexao_banco,
    obter_configuracao,
    salvar_configuracao,
    iniciar_banco,
)
from src.documentos import (
    salvar_documento,
    alterar_competencia_documento,
    _normalizar_modalidade_frete,
    _coletar_ids_documentos_por_numero,
    _coletar_ids_documentos_para_frete,
    atualizar_modalidade_frete_documento,
    declarar_documento_frete,
    salvar_alteracao_frete_manual,
    declarar_intercompany,
    declarar_delta,
    declarar_spot,
    registrar_substituicao,
    desfazer_substituicao,
    cancelar_documento,
    desfazer_cancelamento_documento,
    _buscar_documento_existente_sync,
    register_on_change as _register_doc_on_change,
)
from src.dashboard import (
    obter_dataframe_dashboard,
    criar_figura_faturamento_periodo,
    criar_figura_comparativo_tipos,
)
from src.relatorios import (
    _obter_dataframe_relatorio_filtrado,
    _montar_dataframe_exportacao_periodo,
    escrever_excel_faturamento,
)
from src.importacao import (
    _normalizar_mes_relatorio,
    _extrair_docs_pagina_relatorio,
    _normalizar_coluna_relatorio,
    _achar_coluna,
    _achar_coluna_exata,
    _parse_tipo_documento,
    _mapear_colunas_planilha,
    _linha_valida_para_importacao,
    _inferir_tipo_documento_linha,
    _extrair_valor_frete_linha,
    _normalizar_nome_coluna_planilha,
    _linha_parece_cabecalho_planilha,
    _identificar_secao_planilha,
    _linha_totalizadora_planilha,
    _preparar_dataframe_planilha,
)
from src.sync import (
    exportar_configuracoes_json,
    exportar_configuracoes_repo,
    importar_configuracoes_json,
    importar_configuracoes_repo_se_existir,
    _listar_documentos_alterados_para_sync,
)
from src.seguros import (
    STATUS_SEGURO,
    adicionar_seguro,
    inativar_seguro,
    listar_controle_competencia,
    resumo_competencia,
    atualizar_status_seguro,
    atualizar_observacao_seguro,
)
from src.tarefas import (
    STATUS_TAREFA,
    STATUS_LABELS as TAREFA_STATUS_LABELS,
    PRIORIDADE_LABELS as TAREFA_PRIORIDADE_LABELS,
    adicionar_categoria,
    classificar_prazo,
    criar_tarefa,
    excluir_tarefa,
    formatar_prazo_br,
    listar_categorias,
    listar_tarefas,
    mover_tarefa,
    resumo_tarefas,
    atualizar_tarefa,
)
from src.operacional_sync import (
    exportar_estado_operacional,
    importar_estado_operacional_se_existir,
)
from src.backup import (
    criar_backup_automatico_se_necessario,
    criar_backup_local,
    listar_backups,
    restaurar_backup,
)
from src.busca_global import buscar_global
from src.auditoria_consistencia import auditar_consistencia
from src.cache import doc_cache
from src.logger import get_logger as _get_logger

_logger = _get_logger('sistema_faturamento')

# -----------------------------------------------
#  Constantes de ambiente - lidas de src.config
#  (configurar_diretorio_dados ja rodou no import)
# -----------------------------------------------

SCRIPT_DIR               = _cfg.SCRIPT_DIR
APP_DIR                  = _cfg.APP_DIR
BASE_DIR                 = _cfg.BASE_DIR
PROJECT_DIR              = _cfg.PROJECT_DIR
RELATORIOS_DIR           = _cfg.RELATORIOS_DIR
DEFAULT_APP_DATA_DIR     = _cfg.DEFAULT_APP_DATA_DIR
FALLBACK_APP_DATA_DIR    = _cfg.FALLBACK_APP_DATA_DIR
LEGACY_DB_PATH           = _cfg.LEGACY_DB_PATH
LOGO_PATH                = _cfg.LOGO_PATH
APP_USER_MODEL_ID        = _cfg.APP_USER_MODEL_ID
SYNC_CONFIG_SCHEMA_VERSION = _cfg.SYNC_CONFIG_SCHEMA_VERSION
SYNC_DOCUMENT_FIELDS     = _cfg.SYNC_DOCUMENT_FIELDS

# Mutaveis: atualizados por configurar_diretorio_dados()
APP_DATA_DIR = _cfg.APP_DATA_DIR
DB_PATH      = _cfg.DB_PATH
LOCK_PATH    = _cfg.LOCK_PATH


def _diretorio_gravavel(caminho_dir):
    """Delega para src.config."""
    return _cfg._diretorio_gravavel(caminho_dir)


def configurar_diretorio_dados():
    """Delega para src.config e sincroniza os globals deste modulo."""
    global APP_DATA_DIR, DB_PATH, LOCK_PATH
    _cfg.configurar_diretorio_dados()
    APP_DATA_DIR = _cfg.APP_DATA_DIR
    DB_PATH      = _cfg.DB_PATH
    LOCK_PATH    = _cfg.LOCK_PATH


configurar_diretorio_dados()


def configurar_cache_matplotlib():
    """Delega para src.config."""
    _cfg.configurar_cache_matplotlib()


configurar_cache_matplotlib()


def _processo_ativo(pid):
    try:
        os.kill(pid, 0)
        return True
    except:
        return False


def preparar_arquivos_aplicacao():
    try:
        os.makedirs(APP_DATA_DIR, exist_ok=True)
    except OSError:
        configurar_diretorio_dados()
        os.makedirs(APP_DATA_DIR, exist_ok=True)

    def _contar_documentos(caminho_db):
        try:
            conn = sqlite3.connect(caminho_db)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM documentos")
            total = int(cursor.fetchone()[0])
            conn.close()
            return total
        except Exception:
            return None

    fontes_banco = []
    for origem in (
        os.path.join(DEFAULT_APP_DATA_DIR, "faturamento.db"),
        LEGACY_DB_PATH,
    ):
        origem_abs = os.path.abspath(origem)
        if origem_abs == os.path.abspath(DB_PATH):
            continue
        if origem_abs not in fontes_banco:
            fontes_banco.append(origem_abs)

    # Migra automaticamente um banco saudável para a pasta atual na primeira execução
    # ou quando o fallback local estiver vazio.
    if not os.path.exists(DB_PATH):
        for origem_db in fontes_banco:
            if not os.path.exists(origem_db):
                continue
            try:
                shutil.copy2(origem_db, DB_PATH)
                break
            except OSError:
                continue
        return

    docs_novo = _contar_documentos(DB_PATH)
    if docs_novo == 0:
        for origem_db in fontes_banco:
            docs_origem = _contar_documentos(origem_db)
            if isinstance(docs_origem, int) and docs_origem > 0:
                try:
                    shutil.copy2(origem_db, DB_PATH)
                    break
                except OSError:
                    continue


def adquirir_lock_instancia():
    pid_existente = None
    if os.path.exists(LOCK_PATH):
        try:
            with open(LOCK_PATH, "r", encoding="utf-8") as f:
                pid_txt = f.read().strip()
            pid_existente = int(pid_txt)
            if pid_existente != os.getpid() and _processo_ativo(pid_existente):
                return False, pid_existente
        except (ValueError, OSError):
            pass

        try:
            os.remove(LOCK_PATH)
        except OSError:
            # Se nao houver permissao para manipular o lock, nao bloqueia inicializacao.
            return True, None

    try:
        with open(LOCK_PATH, "w", encoding="utf-8") as f:
            f.write(str(os.getpid()))
    except OSError:
        # Em ambientes com pasta protegida, continua sem lock para nao impedir o uso.
        return True, None

    return True, None


def liberar_lock_instancia():
    try:
        if os.path.exists(LOCK_PATH):
            with open(LOCK_PATH, "r", encoding="utf-8") as f:
                pid_txt = f.read().strip()
            if pid_txt == str(os.getpid()):
                os.remove(LOCK_PATH)
    except OSError:
        pass


def alertar_instancia_em_execucao(pid_existente=None):
    msg = "O sistema de faturamento ja esta em execucao.\nFeche a janela atual antes de abrir novamente."
    if pid_existente:
        msg += f"\n\nPID em execucao: {pid_existente}"
    try:
        ctypes.windll.user32.MessageBoxW(0, msg, "Sistema ja em execucao", 0x30)
    except Exception:
        try:
            messagebox.showwarning("Sistema ja em execucao", msg)
        except Exception:
            print(msg)


# ------------------------
# BANCO  →  src/banco.py
# ------------------------


# _sqlite_db_valido / _tentar_recuperar_banco / obter_conexao_banco
# obter_configuracao / salvar_configuracao / iniciar_banco  →  src/banco.py


# ------------------------
# VARIAVEIS
# ------------------------

relatorio_selecionado = ""
pasta_relatorios_saida = RELATORIOS_DIR
caminho_relatorio_atual = ""
dados_importados = None
relatorio_carregado = False
relatorios_total_intervalo = None
dashboard_data_inicio = ""
dashboard_data_fim = ""
relatorio_data_inicio = ""
relatorio_data_fim = ""

paginas_lidas = 0
docs_encontrados = 0
dashboard_update_after_id = None
dashboard_update_running = False
scroll_refresh_after_id = None
scroll_dragging = False
watermark_hidden_by_scroll = False
ui_animations_paused = False

WATERMARK_POS = {
    "relx": 0.5,
    "rely": 0.52,
    "anchor": "center",
}


# ------------------------
# UTIL  →  src/utils.py
# ------------------------


def obter_periodo_dashboard(silencioso=False):
    global dashboard_data_inicio, dashboard_data_fim
    data_inicial, data_final = _obter_periodo_por_entries(
        globals().get("dashboard_data_inicio_entry"),
        globals().get("dashboard_data_fim_entry"),
        contexto="Filtro do Dashboard",
        silencioso=silencioso,
    )
    if data_inicial and data_final:
        dashboard_data_inicio = data_inicial.strftime("%d/%m/%Y")
        dashboard_data_fim = data_final.strftime("%d/%m/%Y")
    return data_inicial, data_final


def obter_periodo_relatorios(silencioso=False):
    global relatorio_data_inicio, relatorio_data_fim
    data_inicial, data_final = _obter_periodo_por_entries(
        globals().get("relatorio_data_inicio_entry"),
        globals().get("relatorio_data_fim_entry"),
        contexto="Filtro de Relatórios",
        silencioso=silencioso,
    )
    if data_inicial and data_final:
        relatorio_data_inicio = data_inicial.strftime("%d/%m/%Y")
        relatorio_data_fim = data_final.strftime("%d/%m/%Y")
    return data_inicial, data_final


def inicializar_filtros_dashboard():
    global dashboard_data_inicio, dashboard_data_fim
    data_inicial, data_final = obter_periodo_padrao_dashboard()
    dashboard_data_inicio = data_inicial.strftime("%d/%m/%Y")
    dashboard_data_fim = data_final.strftime("%d/%m/%Y")

    entry_inicio = globals().get("dashboard_data_inicio_entry")
    entry_fim = globals().get("dashboard_data_fim_entry")
    if entry_inicio is not None and entry_fim is not None:
        entry_inicio.delete(0, "end")
        entry_inicio.insert(0, dashboard_data_inicio)
        entry_fim.delete(0, "end")
        entry_fim.insert(0, dashboard_data_fim)


def inicializar_filtros_relatorios():
    global relatorio_data_inicio, relatorio_data_fim
    data_inicial, data_final = obter_periodo_padrao_relatorios()
    relatorio_data_inicio = data_inicial.strftime("%d/%m/%Y")
    relatorio_data_fim = data_final.strftime("%d/%m/%Y")

    entry_inicio = globals().get("relatorio_data_inicio_entry")
    entry_fim = globals().get("relatorio_data_fim_entry")
    if entry_inicio is not None and entry_fim is not None:
        entry_inicio.delete(0, "end")
        entry_inicio.insert(0, relatorio_data_inicio)
        entry_fim.delete(0, "end")
        entry_fim.insert(0, relatorio_data_fim)


def centralizar_janela(janela, largura, altura):
    janela.update_idletasks()
    screen_width = janela.winfo_screenwidth()
    screen_height = janela.winfo_screenheight()
    x = int((screen_width / 2) - (largura / 2))
    y = int((screen_height / 2) - (altura / 2))
    janela.geometry(f"{largura}x{altura}+{x}+{y}")


# ─── Suporte a múltiplos monitores ─────────────────────────────────────────

def _obter_area_util_monitor(cx, cy):
    """Retorna (left, top, largura, altura) da area util do monitor em (cx, cy)."""
    try:
        import ctypes

        class _POINT(ctypes.Structure):
            _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]

        class _RECT(ctypes.Structure):
            _fields_ = [("left", ctypes.c_long), ("top", ctypes.c_long),
                        ("right", ctypes.c_long), ("bottom", ctypes.c_long)]

        class _MONITORINFO(ctypes.Structure):
            _fields_ = [("cbSize", ctypes.c_ulong), ("rcMonitor", _RECT),
                        ("rcWork", _RECT), ("dwFlags", ctypes.c_ulong)]

        pt = _POINT(int(cx), int(cy))
        hmon = ctypes.windll.user32.MonitorFromPoint(pt, 2)  # MONITOR_DEFAULTTONEAREST
        info = _MONITORINFO()
        info.cbSize = ctypes.sizeof(_MONITORINFO)
        ctypes.windll.user32.GetMonitorInfoW(hmon, ctypes.byref(info))
        w = info.rcWork.right - info.rcWork.left
        h = info.rcWork.bottom - info.rcWork.top
        return info.rcWork.left, info.rcWork.top, w, h
    except Exception:
        return None, None, None, None


_monitor_anterior = {}
_ajuste_pendente_id = None


def _aplicar_ajuste_monitor():
    """Chamado com debounce — verifica se janela cabe no monitor atual."""
    global _monitor_anterior, _ajuste_pendente_id
    _ajuste_pendente_id = None
    try:
        app.update_idletasks()
        wx = app.winfo_x()
        wy = app.winfo_y()
        ww = app.winfo_width()
        wh = app.winfo_height()
        cx = wx + ww // 2
        cy = wy + wh // 2

        ml, mt, mw, mh = _obter_area_util_monitor(cx, cy)
        if mw is None:
            return

        chave = (ml, mt, mw, mh)
        if chave == _monitor_anterior.get("chave"):
            return
        _monitor_anterior["chave"] = chave

        # Redimensiona apenas se a janela não cabe no monitor de destino
        nova_w = max(860, min(1120, mw - 80))
        nova_h = max(650, min(860, mh - 90))
        nova_w = min(nova_w, mw - 10)
        nova_h = min(nova_h, mh - 10)

        precisa_resize = ww > mw - 8 or wh > mh - 8
        if not precisa_resize:
            return

        # Centraliza no novo monitor
        novo_x = ml + (mw - nova_w) // 2
        novo_y = mt + (mh - nova_h) // 2
        app.geometry(f"{nova_w}x{nova_h}+{novo_x}+{novo_y}")
    except Exception:
        pass


def _on_janela_configure(event):
    """Debounce de 400 ms no evento Configure para detectar troca de monitor."""
    global _ajuste_pendente_id
    try:
        if event.widget is not app:
            return
        if _ajuste_pendente_id is not None:
            app.after_cancel(_ajuste_pendente_id)
        _ajuste_pendente_id = app.after(400, _aplicar_ajuste_monitor)
    except Exception:
        pass


def manter_interface_responsiva():
    try:
        if "app" in globals() and app.winfo_exists():
            app.update()
    except Exception:
        pass


def solicitar_atualizacao_dashboard(_event=None, delay_ms=180):
    global dashboard_update_after_id

    def _executar():
        global dashboard_update_after_id
        dashboard_update_after_id = None
        try:
            if "atualizar_dashboard" in globals():
                atualizar_dashboard()
            _atualizar_status_relatorios_ui()
        except Exception:
            pass

    try:
        if "app" in globals() and app.winfo_exists():
            if dashboard_update_after_id:
                try:
                    app.after_cancel(dashboard_update_after_id)
                except Exception:
                    pass
                dashboard_update_after_id = None
            if delay_ms and delay_ms > 0:
                dashboard_update_after_id = app.after(int(delay_ms), _executar)
            else:
                _executar()
        else:
            _executar()
    except Exception:
        pass


def aplicar_filtro_dashboard(_event=None):
    try:
        data_inicial, data_final = obter_periodo_dashboard(silencioso=False)
    except ValueError as exc:
        messagebox.showwarning("Filtro do Dashboard", str(exc))
        return

    if data_inicial is None or data_final is None:
        messagebox.showwarning("Filtro do Dashboard", "Preencha as datas do Dashboard.")
        return

    if data_inicial > data_final:
        messagebox.showwarning("Filtro do Dashboard", "A data inicial não pode ser maior que a data final.")
        return

    atualizar_dashboard()


def aplicar_filtro_relatorios(_event=None, mensagem_sucesso=True):
    global relatorios_total_intervalo
    try:
        data_inicial, data_final = obter_periodo_relatorios(silencioso=False)
    except ValueError as exc:
        _atualizar_status_relatorios_ui(str(exc), tipo="erro", total_registros=None)
        messagebox.showwarning("Filtro de Relatórios", str(exc))
        return False

    if data_inicial is None or data_final is None:
        mensagem = "Preencha as datas da aba Relatórios."
        _atualizar_status_relatorios_ui(mensagem, tipo="erro", total_registros=None)
        messagebox.showwarning("Filtro de Relatórios", mensagem)
        return False

    if data_inicial > data_final:
        mensagem = "A data inicial não pode ser maior que a data final."
        _atualizar_status_relatorios_ui(mensagem, tipo="erro", total_registros=None)
        messagebox.showwarning("Filtro de Relatórios", mensagem)
        return False

    try:
        df_filtrado, msg_erro = _obter_dataframe_relatorio_filtrado(
            data_inicial,
            data_final,
            docs_df_base=_obter_documentos_em_memoria(force=False),
        )
    except Exception:
        df_filtrado, msg_erro = pd.DataFrame(), ""

    if msg_erro:
        _atualizar_status_relatorios_ui(msg_erro, tipo="erro", total_registros=0)
        if mensagem_sucesso:
            messagebox.showwarning("Filtro de Relatórios", msg_erro)
        return False

    total_registros = int(len(df_filtrado)) if isinstance(df_filtrado, pd.DataFrame) else 0
    relatorios_total_intervalo = total_registros
    if mensagem_sucesso:
        _atualizar_status_relatorios_ui(
            "Período de Relatórios atualizado",
            tipo="ok",
            total_registros=total_registros,
        )
    else:
        _atualizar_status_relatorios_ui(total_registros=total_registros)
    return True


def _forcar_redesenho_pos_scroll():
    # Mitiga artefatos visuais durante/apos arraste da barra lateral no Windows.
    canvas_scroll = globals().get("ui_refs", {}).get("scroll_canvas")
    if canvas_scroll is None:
        scroll_frame = globals().get("container_scroll")
        canvas_scroll = getattr(scroll_frame, "_parent_canvas", None) if scroll_frame is not None else None
    if canvas_scroll is not None:
        try:
            canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all"))
            canvas_scroll.update_idletasks()
        except Exception:
            pass

    for cfg in globals().get("dashboard_chart_widgets", {}).values():
        canvas = cfg.get("canvas")
        if canvas is None:
            continue
        try:
            canvas.draw_idle()
        except Exception:
            try:
                canvas.draw()
            except Exception:
                pass
    try:
        if "app" in globals() and app.winfo_exists():
            app.update_idletasks()
    except Exception:
        pass


def _ocultar_watermark_durante_scroll():
    global watermark_hidden_by_scroll

    lbl = globals().get("ui_refs", {}).get("logo_watermark_label")
    if lbl is None or not lbl.winfo_exists():
        return
    if watermark_hidden_by_scroll:
        return
    try:
        lbl.place_forget()
        watermark_hidden_by_scroll = True
    except Exception:
        pass


def _agendar_restauro_visual_scroll(delay_ms=110):
    global scroll_refresh_after_id

    try:
        if "app" not in globals() or not app.winfo_exists():
            return
        if scroll_refresh_after_id:
            try:
                app.after_cancel(scroll_refresh_after_id)
            except Exception:
                pass
            scroll_refresh_after_id = None
        scroll_refresh_after_id = app.after(max(40, int(delay_ms)), _restaurar_visual_apos_scroll)
    except Exception:
        pass


def _restaurar_visual_apos_scroll():
    global scroll_refresh_after_id, watermark_hidden_by_scroll, ui_animations_paused

    scroll_refresh_after_id = None
    if scroll_dragging:
        _agendar_restauro_visual_scroll(140)
        return

    if watermark_hidden_by_scroll:
        container = globals().get("ui_refs", {}).get("main_frame")
        if container is not None and container.winfo_exists():
            try:
                _aplicar_logo_watermark(container)
            except Exception:
                pass
        watermark_hidden_by_scroll = False

    ui_animations_paused = False
    _forcar_redesenho_pos_scroll()


def _on_scrollbar_press(_event=None):
    global scroll_dragging, ui_animations_paused
    scroll_dragging = True
    ui_animations_paused = True
    _ocultar_watermark_durante_scroll()
    _agendar_restauro_visual_scroll(180)


def _on_scrollbar_drag(_event=None):
    _ocultar_watermark_durante_scroll()
    _agendar_restauro_visual_scroll(180)


def _on_scrollbar_release(_event=None):
    global scroll_dragging
    scroll_dragging = False
    _agendar_restauro_visual_scroll(70)


def _configurar_mitigacao_artefato_scroll(scroll_frame):
    if scroll_frame is None:
        return
    if getattr(scroll_frame, "_artefato_scroll_configurado", False):
        return

    barra = getattr(scroll_frame, "_scrollbar", None)
    canvas = getattr(scroll_frame, "_parent_canvas", None)
    if canvas is not None and canvas.winfo_exists():
        try:
            # Menos steps por arraste reduz "ghosting" no Windows.
            canvas.configure(yscrollincrement=5)
        except Exception:
            pass

    if barra is not None and barra.winfo_exists():
        comando_original = getattr(canvas, "yview", None) if canvas is not None else None
        if callable(comando_original):
            def _comando_scrollbar(*args):
                _on_scrollbar_press()
                try:
                    comando_original(*args)
                finally:
                    _on_scrollbar_drag()
                    _forcar_redesenho_pos_scroll()

            try:
                barra.configure(command=_comando_scrollbar)
            except Exception:
                pass

        barra.bind("<ButtonPress-1>", _on_scrollbar_press, add="+")
        barra.bind("<B1-Motion>", _on_scrollbar_drag, add="+")
        barra.bind("<ButtonRelease-1>", _on_scrollbar_release, add="+")

    if canvas is not None and canvas.winfo_exists():
        canvas.bind("<ButtonRelease-1>", _on_scrollbar_release, add="+")
    try:
        if "app" in globals() and app.winfo_exists():
            app.bind_all("<ButtonRelease-1>", _on_scrollbar_release, add="+")
    except Exception:
        pass

    setattr(scroll_frame, "_artefato_scroll_configurado", True)


def obter_pasta_saida_relatorios():
    global pasta_relatorios_saida
    if not pasta_relatorios_saida:
        pasta_relatorios_saida = RELATORIOS_DIR
    os.makedirs(pasta_relatorios_saida, exist_ok=True)
    return pasta_relatorios_saida


def atualizar_label_pasta_saida():
    pasta_txt = f"Pasta de saída: {obter_pasta_saida_relatorios()}"
    if "pasta_saida_label" in globals():
        pasta_saida_label.configure(text=pasta_txt)
    lbl_saida = None
    if "ui_refs" in globals():
        lbl_saida = ui_refs.get("relatorios_saida_label")
    try:
        if lbl_saida is not None and lbl_saida.winfo_exists():
            if "UI_THEME" in globals():
                lbl_saida.configure(text=pasta_txt, text_color=UI_THEME["text_secondary"])
            else:
                lbl_saida.configure(text=pasta_txt)
    except Exception:
        pass


def carregar_pasta_saida_relatorios():
    global pasta_relatorios_saida
    caminho_salvo = obter_configuracao("pasta_relatorios_saida", "").strip()
    if caminho_salvo:
        pasta_relatorios_saida = caminho_salvo
    obter_pasta_saida_relatorios()
    atualizar_label_pasta_saida()


def selecionar_pasta_saida_relatorios():
    global pasta_relatorios_saida
    caminho = filedialog.askdirectory(
        title="Selecionar pasta para salvar o relatório",
        initialdir=obter_pasta_saida_relatorios(),
        mustexist=True,
    )
    if not caminho:
        return

    pasta_relatorios_saida = caminho
    salvar_configuracao("pasta_relatorios_saida", pasta_relatorios_saida)
    obter_pasta_saida_relatorios()
    atualizar_label_pasta_saida()
    messagebox.showinfo("Pasta de saída", f"Relatórios serão salvos em:\n{pasta_relatorios_saida}")


def _resolver_pasta_arquivo_relatorio():
    pasta_saida = obter_pasta_saida_relatorios()
    candidatos = []
    if isinstance(dados_importados, dict):
        candidatos.append(dados_importados.get("ultimo_export_periodo", ""))
        candidatos.append(dados_importados.get("ultimo_relatorio_filtrado", ""))

    for candidato in candidatos:
        caminho = str(candidato or "").strip()
        if not caminho:
            continue
        if os.path.isfile(caminho):
            pasta = os.path.dirname(os.path.abspath(caminho))
            if os.path.normcase(os.path.abspath(pasta)) == os.path.normcase(os.path.abspath(pasta_saida)):
                return pasta
        if os.path.isdir(caminho):
            pasta = os.path.abspath(caminho)
            if os.path.normcase(pasta) == os.path.normcase(os.path.abspath(pasta_saida)):
                return pasta

    return pasta_saida


def abrir_pasta_arquivo_relatorio():
    pasta = _resolver_pasta_arquivo_relatorio()
    if not pasta:
        messagebox.showwarning("Abrir pasta", "Nenhuma pasta de relatório foi encontrada.")
        return

    try:
        os.makedirs(pasta, exist_ok=True)
        os.startfile(pasta)
    except Exception as exc:
        messagebox.showerror("Abrir pasta", f"Não foi possível abrir a pasta:\n{pasta}\n\n{exc}")


def salvar_ultimo_relatorio(caminho):
    if not caminho:
        return
    caminho_abs = os.path.abspath(caminho)
    salvar_configuracao("ultimo_relatorio_caminho", caminho_abs)
    salvar_configuracao("ultimo_relatorio_nome", os.path.basename(caminho_abs))
    salvar_configuracao("ultimo_relatorio_diretorio", os.path.dirname(caminho_abs))


def _resolver_ultimo_relatorio_salvo():
    caminho_salvo = obter_configuracao("ultimo_relatorio_caminho", "").strip()
    if caminho_salvo and os.path.exists(caminho_salvo):
        return caminho_salvo

    diretorio_salvo = obter_configuracao("ultimo_relatorio_diretorio", "").strip()
    nome_salvo = obter_configuracao("ultimo_relatorio_nome", "").strip()
    if diretorio_salvo and nome_salvo:
        candidato = os.path.join(diretorio_salvo, nome_salvo)
        if os.path.exists(candidato):
            salvar_ultimo_relatorio(candidato)
            return candidato

    return ""


def atualizar_label_relatorio():
    if "pasta_label" in globals() and pasta_label is not None:
        if relatorio_selecionado:
            nome_arquivo = os.path.basename(relatorio_selecionado)
            pasta_label.configure(text=f"Arquivo: {nome_arquivo}")
        else:
            pasta_label.configure(text="Arquivo: Nenhum selecionado")
    _atualizar_status_relatorios_ui()


def definir_relatorio_selecionado(caminho, persistir=False):
    global relatorio_selecionado, caminho_relatorio_atual
    relatorio_selecionado = os.path.normpath(caminho) if caminho else ""
    caminho_relatorio_atual = relatorio_selecionado
    if persistir and relatorio_selecionado:
        salvar_ultimo_relatorio(relatorio_selecionado)
    atualizar_label_relatorio()
    return relatorio_selecionado


def _periodo_dashboard_relatorio_texto():
    try:
        ini, fim = obter_periodo_relatorios(silencioso=True)
        if ini is None and relatorio_data_inicio and relatorio_data_fim:
            ini = ler_data_filtro(relatorio_data_inicio, "Data inicial")
            fim = ler_data_filtro(relatorio_data_fim, "Data final")
        if ini and fim:
            meses_abrev = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            ini_txt = f"{meses_abrev[ini.month - 1]}/{ini.strftime('%y')}"
            fim_txt = f"{meses_abrev[fim.month - 1]}/{fim.strftime('%y')}"
            return f"{ini_txt} a {fim_txt}"
    except Exception:
        pass
    return "Período indefinido"


def _atualizar_status_relatorios_ui(mensagem=None, tipo="ok", total_registros=None):
    global relatorios_total_intervalo
    if "UI_THEME" not in globals() or "ui_refs" not in globals():
        return

    lbl_status = ui_refs.get("relatorios_status_label")
    lbl_arquivo = ui_refs.get("relatorios_arquivo_label")
    lbl_periodo = ui_refs.get("relatorios_periodo_label")
    lbl_registros = ui_refs.get("relatorios_registros_label")
    lbl_saida = ui_refs.get("relatorios_saida_label")

    if relatorio_selecionado and relatorio_carregado:
        nome_arquivo = os.path.basename(relatorio_selecionado)
        arquivo_txt = f"Arquivo: {nome_arquivo}"
        status_padrao = "Relatório carregado com sucesso"
    elif relatorio_selecionado:
        nome_arquivo = os.path.basename(relatorio_selecionado)
        arquivo_txt = f"Arquivo: {nome_arquivo}"
        status_padrao = "Arquivo selecionado"
    else:
        arquivo_txt = "Arquivo: Nenhum selecionado"
        status_padrao = "Nenhum relatório carregado"
        if mensagem is None:
            tipo = "neutral"

    status_txt = mensagem or status_padrao
    periodo_txt = f"Período aplicado: {_periodo_dashboard_relatorio_texto()}"
    if total_registros is not None:
        relatorios_total_intervalo = int(total_registros)
    registros_txt = (
        f"Registros no período: {relatorios_total_intervalo}"
        if isinstance(relatorios_total_intervalo, int)
        else "Registros no período: -"
    )

    cor_status = UI_THEME["success_text"] if tipo == "ok" else (
        UI_THEME["danger_text"] if tipo == "erro" else UI_THEME["text_secondary"]
    )

    _safe_config(lbl_status, text=status_txt, text_color=cor_status)
    _safe_config(lbl_arquivo, text=arquivo_txt, text_color=UI_THEME["text_primary"])
    _safe_config(lbl_saida, text=f"Pasta de saída: {obter_pasta_saida_relatorios()}", text_color=UI_THEME["text_secondary"])
    _safe_config(lbl_periodo, text=periodo_txt, text_color=UI_THEME["text_secondary"])
    _safe_config(lbl_registros, text=registros_txt, text_color=UI_THEME["text_secondary"])


def carregar_ultimo_relatorio():
    global relatorio_carregado, dados_importados, caminho_relatorio_atual
    caminho = _resolver_ultimo_relatorio_salvo()
    if caminho:
        definir_relatorio_selecionado(caminho, persistir=False)
    caminho_relatorio_atual = relatorio_selecionado
    try:
        conn = obter_conexao_banco()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM documentos")
        relatorio_carregado = int(cursor.fetchone()[0] or 0) > 0
        conn.close()
    except Exception:
        relatorio_carregado = False
    if relatorio_carregado:
        dados_importados = {"origem": "banco_local", "carregado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S")}
        try:
            _obter_documentos_em_memoria(force=True)
        except Exception:
            pass


def _carregar_documentos_para_memoria():
    conn = obter_conexao_banco()
    try:
        df = pd.read_sql_query("SELECT * FROM documentos", conn)
    finally:
        conn.close()
    return df


def _obter_documentos_em_memoria(force=False):
    global dados_importados

    if not force and doc_cache.valido:
        return doc_cache.get()

    df_mem = _carregar_documentos_para_memoria()
    doc_cache.set(df_mem)

    # Mantém compatibilidade com dados_importados legado
    if not isinstance(dados_importados, dict):
        dados_importados = {}
    dados_importados.update(doc_cache.para_dict_legado())

    return doc_cache.get()


def _atualizar_cache_documentos_pos_alteracao():
    global dados_importados, relatorio_carregado
    try:
        df_mem = _obter_documentos_em_memoria(force=True)
        relatorio_carregado = not df_mem.empty
        if not isinstance(dados_importados, dict):
            dados_importados = {}
        dados_importados["total_documentos"] = doc_cache.total
        dados_importados["memoria_atualizada_em"] = doc_cache.atualizado_em
    except Exception as exc:
        _logger.warning("_atualizar_cache_documentos_pos_alteracao falhou: %s", exc)


def aplicar_icone_aplicacao(janela):
    # Forca um AppUserModelID proprio para evitar agrupamento/icone do Python na barra.
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
    except Exception:
        pass

    try:
        icon_path = os.path.join(BASE_DIR, "logo.ico")
        if os.path.exists(icon_path):
            janela.iconbitmap(default=icon_path)
    except Exception as e:
        print(f"Aviso: não foi possível carregar o ícone .ico - {e}")

    # Fallback para garantir icone em janelas Tk quando .ico falhar.
    try:
        if os.path.exists(LOGO_PATH):
            janela._app_icon_photo = tk.PhotoImage(file=LOGO_PATH)
            janela.iconphoto(True, janela._app_icon_photo)
    except Exception:
        pass


def abrir_seletor_data(entry_widget):
    texto_atual = entry_widget.get().strip()
    try:
        data_base = datetime.strptime(texto_atual, "%d/%m/%Y")
    except ValueError:
        data_base = datetime.now()

    picker = ctk.CTkToplevel(app)
    picker.title("Selecionar data")
    centralizar_janela(picker, 290, 280)
    picker.grab_set()

    estado = {"ano": data_base.year, "mes": data_base.month}
    semana_nomes = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]

    topo = ctk.CTkFrame(picker, fg_color="transparent")
    topo.pack(fill="x", padx=8, pady=(8, 4))

    def mudar_mes(delta):
        mes = estado["mes"] + delta
        ano = estado["ano"]
        if mes < 1:
            mes = 12
            ano -= 1
        elif mes > 12:
            mes = 1
            ano += 1
        estado["mes"] = mes
        estado["ano"] = ano
        renderizar()

    ctk.CTkButton(topo, text="<", width=30, command=lambda: mudar_mes(-1)).pack(side="left")
    titulo_mes = ctk.CTkLabel(topo, text="", font=ctk.CTkFont(weight="bold"))
    titulo_mes.pack(side="left", expand=True)
    ctk.CTkButton(topo, text=">", width=30, command=lambda: mudar_mes(1)).pack(side="right")

    corpo = ctk.CTkFrame(picker)
    corpo.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def selecionar_dia(dia):
        valor = datetime(estado["ano"], estado["mes"], dia).strftime("%d/%m/%Y")
        entry_widget.delete(0, "end")
        entry_widget.insert(0, valor)
        solicitar_atualizacao_dashboard(delay_ms=0)
        picker.destroy()

    def renderizar():
        for child in corpo.winfo_children():
            child.destroy()

        titulo_mes.configure(text=f"{MESES[estado['mes'] - 1].capitalize()}/{estado['ano']}")

        for c, nome in enumerate(semana_nomes):
            ctk.CTkLabel(corpo, text=nome).grid(row=0, column=c, padx=2, pady=(4, 2))

        semanas = calendar.monthcalendar(estado["ano"], estado["mes"])
        for r, semana in enumerate(semanas, start=1):
            for c, dia in enumerate(semana):
                if dia == 0:
                    ctk.CTkLabel(corpo, text="", width=34).grid(row=r, column=c, padx=2, pady=2)
                else:
                    ctk.CTkButton(
                        corpo,
                        text=str(dia),
                        width=34,
                        height=28,
                        command=lambda d=dia: selecionar_dia(d),
                    ).grid(row=r, column=c, padx=2, pady=2)

    renderizar()


# ------------------------
# PORTAL NFS-E
# ------------------------


# ------------------------
# EXTRAIR CTE
# ------------------------

# ------------------------
# EXTRAIR NF
# ------------------------

# ─── src/importacao.py (parsers PDF) ────────────────────────────────────────

def importar_relatorio_consolidado(caminho_pdf):
    global docs_encontrados, paginas_lidas

    docs_encontrados = 0
    paginas_lidas = 0

    vistos = set()
    competencia_ctx = None

    try:
        with fitz.open(caminho_pdf) as pdf:
            total_pag = pdf.page_count
            progress.set(0)
            for idx, pagina in enumerate(pdf, start=1):
                texto = pagina.get_text("text")
                docs, competencia_ctx = _extrair_docs_pagina_relatorio(texto, competencia_ctx)
                for doc in docs:
                    chave = (doc["tipo"], doc["numero"])
                    if chave in vistos:
                        continue
                    vistos.add(chave)
                    salvar_documento(doc)
                    docs_encontrados += 1

                paginas_lidas = idx
                progress.set(idx / total_pag)
                status_label.configure(text=f"Importando relatório: página {idx}/{total_pag} | Docs: {docs_encontrados}")
                manter_interface_responsiva()
    except Exception as exc:
        messagebox.showerror("Relatório", f"Falha ao importar relatório.\n\n{exc}")
        return

    messagebox.showinfo("Relatório", f"Importação concluída. {docs_encontrados} documento(s) carregado(s).")

    atualizar_dashboard()
    _atualizar_status_relatorios_ui("Relatório carregado com sucesso", tipo="ok")


# ─── src/importacao.py (parsers planilha) ───────────────────────────────────

def importar_relatorio_planilha(caminho_planilha):
    global docs_encontrados, paginas_lidas

    docs_encontrados = 0
    paginas_lidas = 0

    try:
        planilhas = pd.read_excel(caminho_planilha, sheet_name=None, header=None)
    except ImportError as exc:
        messagebox.showerror(
            "Relatório",
            "Falha ao abrir planilha. Para arquivo .xls, instale 'xlrd' ou exporte para .xlsx.\n\n" + str(exc),
        )
        return
    except Exception as exc:
        messagebox.showerror("Relatório", f"Falha ao abrir planilha.\n\n{exc}")
        return

    if not planilhas:
        messagebox.showwarning("Relatório", "A planilha não possui abas para importar.")
        return

    # Limpa importacoes automaticas anteriores para evitar residuos no novo relatorio.
    conn = obter_conexao_banco()
    cursor = conn.cursor()
    cursor.execute(
        """
        DELETE FROM documentos
        WHERE COALESCE(cancelado_manual,0)=0
          AND COALESCE(competencia_manual,0)=0
          AND COALESCE(frete_manual,0)=0
          AND UPPER(COALESCE(status,'')) NOT LIKE '%CANCELADO%'
          AND UPPER(COALESCE(status,'')) NOT LIKE '%SUBSTITUIDO%'
          AND UPPER(COALESCE(status,'')) NOT LIKE 'DOCUMENTO SUBSTITUINDO%'
        """
    )
    conn.commit()
    conn.close()

    vistos = {}

    conn = obter_conexao_banco()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    erros = []
    total_linhas = 0
    planilhas_preparadas = {}
    for nome_aba, df in planilhas.items():
        df_prep = _preparar_dataframe_planilha(df)
        planilhas_preparadas[nome_aba] = df_prep
        if df_prep is not None and not df_prep.empty:
            total_linhas += len(df_prep.index)

    if total_linhas == 0:
        messagebox.showwarning("Relatório", "Não há linhas com dados na planilha selecionada.")
        return

    processadas = 0
    total_abas = len(planilhas)

    for idx_aba, (nome_aba, df_aba) in enumerate(planilhas_preparadas.items(), start=1):
        paginas_lidas = idx_aba

        if df_aba is None or df_aba.empty:
            continue

        mapa = _mapear_colunas_planilha(df_aba)
        if mapa["faltando"]:
            preview_cols = ", ".join([str(c) for c in list(df_aba.columns)[:20]])
            erros.append(f"Aba '{nome_aba}' sem colunas obrigatorias: {', '.join(mapa['faltando'])}. Colunas detectadas: {preview_cols}")
            continue

        for _, linha in df_aba.iterrows():
            processadas += 1

            if not _linha_valida_para_importacao(linha, mapa):
                continue

            tipo = _inferir_tipo_documento_linha(linha, mapa)

            numero_txt = re.sub(r"\D", "", str(linha.get(mapa["numero"], "")))
            if not numero_txt:
                continue

            try:
                numero = int(numero_txt)
            except ValueError:
                continue

            data_raw = linha.get(mapa["data"])
            data = pd.to_datetime(data_raw, dayfirst=True, errors="coerce")
            if pd.isna(data) and mapa.get("data_ref"):
                data = pd.to_datetime(linha.get(mapa["data_ref"]), dayfirst=True, errors="coerce")
            if pd.isna(data):
                continue
            data = data.to_pydatetime()

            valor = _extrair_valor_frete_linha(linha, mapa)
            status = "OK"
            if valor is None:
                valor_inicial = 0.0
                status = "ERRO NO LANCAMENTO NO SISTEMA"
            else:
                valor_inicial = float(valor)

            if tipo == "NF":
                numero = int(numero_txt)
                valor_final = round(valor_inicial * 0.95, 2) if valor_inicial > 0 else 0.0
            else:
                numero = int(numero_txt)
                valor_final = valor_inicial

            frete = "FRANQUIA"
            if mapa["status"]:
                status_raw = str(linha.get(mapa["status"], "")).strip()
                if status == "OK" and status_raw and status_raw.lower() != "nan":
                    status = status_raw

            chave = (tipo, numero_txt) if tipo == "NF" else (tipo, numero)
            status_chave = "OK" if status == "OK" else "ERRO"
            if chave in vistos:
                status_existente = vistos[chave]
                # Mantem o registro mais confiavel: OK > ERRO.
                if status_existente == "OK":
                    continue
                if status_existente == "ERRO" and status_chave == "ERRO":
                    continue

            salvar_documento(
                {
                    "numero": numero,
                    "numero_original": numero_txt,
                    "tipo": tipo,
                    "data": data,
                    "valor_inicial": valor_inicial,
                    "valor_final": valor_final,
                    "frete": frete,
                    "status": status,
                    "competencia": competencia_por_data(data),
                },
                cursor
            )
            vistos[chave] = status_chave
            docs_encontrados += 1

            progress.set(processadas / total_linhas)
            status_label.configure(
                text=f"Importando planilha: linha {processadas}/{total_linhas} | Docs:{docs_encontrados}"
            )
            manter_interface_responsiva()

        progress.set(idx_aba / total_abas)
        status_label.configure(
            text=f"Importando planilha: aba {idx_aba}/{total_abas} | Docs:{docs_encontrados}"
        )
        manter_interface_responsiva()

    conn.commit()
    conn.close()

    resumo = f"Importação concluída. {docs_encontrados} documento(s) carregado(s)."
    if erros:
        resumo += "\n\nAvisos:\n- " + "\n- ".join(erros[:5])
        if len(erros) > 5:
            resumo += f"\n- ... e mais {len(erros) - 5} aviso(s)."

    messagebox.showinfo("Relatório", resumo)

    atualizar_dashboard()
    _atualizar_status_relatorios_ui("Relatório carregado com sucesso", tipo="ok")


def _importar_relatorio_por_caminho(caminho):
    global relatorio_carregado, dados_importados, caminho_relatorio_atual, relatorios_total_intervalo
    ext = os.path.splitext(caminho)[1].lower()
    try:
        if ext == ".pdf":
            importar_relatorio_consolidado(caminho)
        elif ext in {".xlsx", ".xls"}:
            importar_relatorio_planilha(caminho)
        else:
            messagebox.showerror("Relatório", "Formato de arquivo não suportado. Use PDF, XLSX ou XLS.")
            _atualizar_status_relatorios_ui("Falha ao carregar relatório", tipo="erro")
            return False
    except Exception as exc:
        relatorio_carregado = False
        dados_importados = None
        relatorios_total_intervalo = None
        _atualizar_status_relatorios_ui("Falha ao carregar relatório", tipo="erro")
        messagebox.showerror("Relatório", f"Falha ao importar relatório.\n\n{exc}")
        return False

    caminho_relatorio_atual = os.path.normpath(caminho)
    relatorio_carregado = True
    relatorios_total_intervalo = None
    dados_importados = {
        "caminho": caminho_relatorio_atual,
        "fonte": ext,
        "importado_em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }
    try:
        dados_importados["df_documentos"] = _obter_documentos_em_memoria(force=True)
        dados_importados["total_documentos"] = int(len(dados_importados["df_documentos"]))
    except Exception:
        pass
    _atualizar_status_relatorios_ui("Relatório carregado com sucesso", tipo="ok")
    return True


def importar_relatorio_automaticamente(caminho):
    return _importar_relatorio_por_caminho(caminho)


def selecionar_relatorio():
    diretorio_inicial = (
        os.path.dirname(relatorio_selecionado)
        if relatorio_selecionado
        else obter_configuracao("ultimo_relatorio_diretorio", RELATORIOS_DIR).strip() or RELATORIOS_DIR
    )
    caminho = filedialog.askopenfilename(
        title="Selecionar relatório consolidado",
        initialdir=diretorio_inicial,
        filetypes=[
            ("Relatórios", "*.xlsx *.xls *.pdf"),
            ("Excel", "*.xlsx *.xls"),
            ("PDF", "*.pdf"),
            ("Todos os arquivos", "*.*"),
        ],
    )
    if caminho:
        caminho = definir_relatorio_selecionado(caminho, persistir=True)
        importar_relatorio_automaticamente(caminho)
        return caminho
    return ""


def importar_relatorio_ui():
    try:
        caminho = relatorio_selecionado

        # 1) Tenta relatorio salvo em memoria/configuracao.
        if not caminho:
            caminho = _resolver_ultimo_relatorio_salvo()
            if caminho:
                definir_relatorio_selecionado(caminho, persistir=True)

        # 2) Se o caminho salvo nao existir mais, tenta recuperar pelo ultimo diretorio + nome.
        if not os.path.exists(caminho):
            nome_salvo = os.path.basename(caminho) if caminho else obter_configuracao("ultimo_relatorio_nome", "").strip()
            diretorio_salvo = obter_configuracao("ultimo_relatorio_diretorio", "").strip()
            if nome_salvo and diretorio_salvo:
                candidato = os.path.join(diretorio_salvo, nome_salvo)
                if os.path.exists(candidato):
                    caminho = definir_relatorio_selecionado(candidato, persistir=True)

        # 3) Fallback: pede selecao manual somente se necessario.
        if not caminho or not os.path.exists(caminho):
            caminho = selecionar_relatorio()
            if not caminho:
                messagebox.showwarning("Relatório", "Nenhum relatório selecionado para importação.")
                return

        definir_relatorio_selecionado(caminho, persistir=True)
        _importar_relatorio_por_caminho(caminho)
    except Exception as exc:
        try:
            status_label.configure(text="Falha ao importar relatório.")
        except Exception:
            pass
        _atualizar_status_relatorios_ui("Falha ao carregar relatório", tipo="erro")
        messagebox.showerror("Relatório", f"Falha ao importar relatório.\n\n{exc}")


# ─── src/sync.py ────────────────────────────────────────────────────────────
# _documento_possui_alteracao_manual, _listar_documentos_alterados_para_sync,
# exportar_configuracoes_json, _extrair_documentos_payload_sync,
# importar_configuracoes_json,
# _to_float, _to_optional_float, _to_manual_flag, _normalizar_data_emissao_sync
# ─────────────────────────────────────────────────────────────────────────────

# _coletar_numero_original_para_match  →  src/utils.py


def exportar_configuracoes_ui():
    try:
        docs = _listar_documentos_alterados_para_sync()
        if not docs:
            messagebox.showwarning("Configurações", "Não há alterações manuais para exportar.")
            return

        diretorio_inicial = (
            obter_configuracao("ultimo_sync_diretorio", obter_pasta_saida_relatorios()).strip()
            or obter_pasta_saida_relatorios()
        )
        nome_padrao = f"configuracoes_faturamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        caminho = filedialog.asksaveasfilename(
            title="Exportar configurações",
            defaultextension=".json",
            initialdir=diretorio_inicial,
            initialfile=nome_padrao,
            filetypes=[("Arquivo JSON", "*.json"), ("Todos os arquivos", "*.*")],
        )
        if not caminho:
            return

        total = exportar_configuracoes_json(caminho)
        salvar_configuracao("ultimo_sync_diretorio", os.path.dirname(caminho))
        messagebox.showinfo(
            "Configurações",
            "Configurações exportadas com sucesso.\n\n"
            f"Documentos exportados: {total}\n"
            f"Arquivo: {caminho}",
        )
    except Exception as exc:
        messagebox.showerror("Configurações", f"Falha ao exportar configurações.\n\n{exc}")


def importar_configuracoes_ui():
    try:
        diretorio_inicial = (
            obter_configuracao("ultimo_sync_diretorio", obter_pasta_saida_relatorios()).strip()
            or obter_pasta_saida_relatorios()
        )
        caminho = filedialog.askopenfilename(
            title="Importar configurações",
            initialdir=diretorio_inicial,
            filetypes=[("Arquivo JSON", "*.json"), ("Todos os arquivos", "*.*")],
        )
        if not caminho:
            return

        salvar_configuracao("ultimo_sync_diretorio", os.path.dirname(caminho))
        resumo = importar_configuracoes_json(caminho)
        _atualizar_cache_documentos_pos_alteracao()
        atualizar_dashboard()
        _atualizar_status_relatorios_ui("Configurações importadas com sucesso", tipo="ok")

        mensagem = (
            "Importação concluída com sucesso.\n\n"
            f"Inseridos: {resumo['inseridos']}\n"
            f"Atualizados: {resumo['atualizados']}\n"
            f"Ignorados: {resumo['ignorados']}\n"
            f"Erros: {len(resumo['erros'])}"
        )
        if resumo["erros"]:
            mensagem += "\n\nPrimeiros erros:\n- " + "\n- ".join(resumo["erros"][:5])
            if len(resumo["erros"]) > 5:
                mensagem += f"\n- ... e mais {len(resumo['erros']) - 5} erro(s)."

        messagebox.showinfo("Configurações", mensagem)
    except Exception as exc:
        messagebox.showerror("Configurações", f"Falha ao importar configurações.\n\n{exc}")


def importar_configuracoes_repo_inicial():
    try:
        resumo = importar_configuracoes_repo_se_existir()
        if resumo:
            _logger.info("Configurações manuais do repositório importadas: %s", resumo)
    except Exception as exc:
        _logger.warning("Falha ao importar configurações manuais do repositório: %s", exc, exc_info=True)


def exportar_configuracoes_repo_silencioso():
    try:
        total = exportar_configuracoes_repo()
        _logger.info("Configurações manuais do repositório exportadas: %s documento(s)", total)
    except Exception as exc:
        _logger.warning("Falha ao exportar configurações manuais do repositório: %s", exc, exc_info=True)


def importar_estado_operacional_inicial():
    try:
        resumo = importar_estado_operacional_se_existir()
        if resumo:
            _logger.info("Estado operacional do repositório importado: %s", resumo)
    except Exception as exc:
        _logger.warning("Falha ao importar estado operacional do repositório: %s", exc, exc_info=True)


def exportar_estado_operacional_silencioso():
    try:
        resumo = exportar_estado_operacional()
        _logger.info("Estado operacional do repositório exportado: %s", resumo)
    except Exception as exc:
        _logger.warning("Falha ao exportar estado operacional do repositório: %s", exc, exc_info=True)


def criar_backup_automatico_silencioso(acionado_por="automatico", forcar=False):
    try:
        exportar_configuracoes_repo()
        exportar_estado_operacional()
        if forcar:
            resultado = criar_backup_local(acionado_por=acionado_por)
        else:
            resultado = criar_backup_automatico_se_necessario(acionado_por=acionado_por)
        _logger.info("Backup local: %s", resultado)
        return resultado
    except Exception as exc:
        _logger.warning("Falha ao criar backup local: %s", exc, exc_info=True)
        return None


def criar_backup_manual_ui():
    resultado = criar_backup_automatico_silencioso("manual", forcar=True)
    if not resultado:
        messagebox.showerror("Backup", "Não foi possível criar o backup agora. Veja os logs do sistema.")
        return
    messagebox.showinfo(
        "Backup criado",
        "Backup local criado com sucesso.\n\n"
        f"Pasta: {resultado.get('pasta')}\n"
        f"Arquivos salvos: {resultado.get('arquivos')}",
    )
    _recriar_tela_hoje()


def _caminho_log_principal():
    return os.path.join(_cfg.APP_DATA_DIR, "faturamento.log")


def _ler_ultimas_linhas(caminho, limite=160):
    if not caminho or not os.path.exists(caminho):
        return []
    try:
        with open(caminho, "r", encoding="utf-8", errors="replace") as f:
            linhas = f.readlines()
        return [linha.rstrip("\n") for linha in linhas[-int(limite):]]
    except OSError:
        return []


def _abrir_pasta_no_explorer(caminho):
    try:
        if not caminho:
            return
        os.makedirs(caminho, exist_ok=True)
        os.startfile(caminho)
    except Exception as exc:
        messagebox.showerror("Abrir pasta", str(exc))


def _abrir_dialogo_logs():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Logs e erros do sistema")
    centralizar_janela(dialog, 920, 620)
    dialog.grab_set()

    frame = ctk.CTkFrame(dialog, fg_color=UI_THEME["app_bg"])
    frame.pack(fill="both", expand=True, padx=14, pady=14)
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(1, weight=1)

    topo = ctk.CTkFrame(frame, fg_color="transparent")
    topo.grid(row=0, column=0, sticky="ew", pady=(0, 10))
    topo.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        topo,
        text="Ultimos registros do sistema",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")

    texto = ctk.CTkTextbox(frame, corner_radius=12, font=ctk.CTkFont(family="Consolas", size=11))
    texto.grid(row=1, column=0, sticky="nsew")

    def _carregar():
        caminhos = [
            ("Sistema", _caminho_log_principal()),
            ("Auditoria", os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs", "medicao_audit.log")),
            ("Diagnostico", os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs", "medicao_diagnostico.txt")),
        ]
        blocos = []
        for titulo, caminho in caminhos:
            linhas = _ler_ultimas_linhas(caminho, 120)
            blocos.append(f"===== {titulo}: {caminho} =====")
            blocos.extend(linhas if linhas else ["Sem registros encontrados."])
            blocos.append("")
        texto.configure(state="normal")
        texto.delete("1.0", "end")
        texto.insert("1.0", "\n".join(blocos))
        texto.configure(state="disabled")

    botoes = ctk.CTkFrame(frame, fg_color="transparent")
    botoes.grid(row=2, column=0, sticky="ew", pady=(10, 0))
    botoes.grid_columnconfigure((0, 1, 2), weight=1)
    ctk.CTkButton(botoes, text="Atualizar", height=34, command=_carregar).grid(row=0, column=0, sticky="ew", padx=(0, 5))
    ctk.CTkButton(
        botoes,
        text="Abrir pasta de logs",
        height=34,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        command=lambda: _abrir_pasta_no_explorer(_cfg.APP_DATA_DIR),
    ).grid(row=0, column=1, sticky="ew", padx=5)
    ctk.CTkButton(
        botoes,
        text="Fechar",
        height=34,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        command=dialog.destroy,
    ).grid(row=0, column=2, sticky="ew", padx=(5, 0))
    _carregar()


def _abrir_dialogo_busca_global():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Busca Global")
    centralizar_janela(dialog, 860, 620)
    dialog.grab_set()

    frame = ctk.CTkFrame(dialog, fg_color=UI_THEME["app_bg"])
    frame.pack(fill="both", expand=True, padx=16, pady=16)
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(2, weight=1)

    ctk.CTkLabel(
        frame,
        text="Busca Global",
        font=ctk.CTkFont(family="Segoe UI", size=21, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(
        frame,
        text="Procure NF, CTE, tarefas, seguros e alteracoes manuais em um lugar so.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=1, column=0, sticky="w", pady=(2, 10))

    topo = ctk.CTkFrame(frame, fg_color="transparent")
    topo.grid(row=0, column=1, rowspan=2, sticky="e")
    termo_entry = ctk.CTkEntry(topo, width=260, height=36, placeholder_text="Digite ao menos 2 caracteres")
    termo_entry.pack(side="left", padx=(0, 8))

    lista = ctk.CTkScrollableFrame(frame, fg_color="transparent")
    lista.grid(row=2, column=0, columnspan=2, sticky="nsew")

    def _destino_por_origem(origem):
        return {
            "documentos": "relatorios",
            "tarefas": "tarefas",
            "seguros": "seguros",
            "alteracoes": "alteracoes",
        }.get(origem, "dashboard")

    def _renderizar(resultados, termo):
        for widget in lista.winfo_children():
            widget.destroy()
        if not termo or len(termo.strip()) < 2:
            msg = "Digite ao menos 2 caracteres para pesquisar."
        elif not resultados:
            msg = "Nenhum resultado encontrado."
        else:
            msg = ""
        if msg:
            ctk.CTkLabel(lista, text=msg, font=ctk.CTkFont(family="Segoe UI", size=12), text_color=UI_THEME["text_secondary"]).pack(anchor="w", padx=8, pady=12)
            return
        for item in resultados:
            origem = item.get("origem")
            row = ctk.CTkFrame(lista, fg_color=UI_THEME["surface_alt"], corner_radius=13, border_width=1, border_color=UI_THEME["border"])
            row.pack(fill="x", padx=4, pady=5)
            row.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(row, text=item.get("tipo") or "-", fg_color=UI_THEME["accent"], corner_radius=10, text_color=UI_THEME["on_accent"], width=82, font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold")).grid(row=0, column=0, rowspan=2, padx=10, pady=10)
            ctk.CTkLabel(row, text=item.get("titulo") or "-", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color=UI_THEME["text_primary"], anchor="w").grid(row=0, column=1, sticky="ew", pady=(10, 1))
            ctk.CTkLabel(row, text=item.get("detalhe") or "-", font=ctk.CTkFont(family="Segoe UI", size=10), text_color=UI_THEME["text_secondary"], anchor="w").grid(row=1, column=1, sticky="ew", pady=(0, 10))
            ctk.CTkButton(
                row,
                text="Abrir",
                width=70,
                height=30,
                fg_color=UI_THEME["surface"],
                hover_color=UI_THEME["tab_hover"],
                border_width=1,
                border_color=UI_THEME["border"],
                text_color=UI_THEME["text_primary"],
                command=lambda d=_destino_por_origem(origem): (dialog.destroy(), app.mostrar_tela(d)),
            ).grid(row=0, column=2, rowspan=2, padx=10, pady=10)

    def _buscar(_event=None):
        termo = termo_entry.get().strip()
        try:
            resultados = buscar_global(termo)
        except Exception as exc:
            messagebox.showerror("Busca Global", str(exc))
            return
        _renderizar(resultados, termo)

    ctk.CTkButton(topo, text="Buscar", width=90, height=36, fg_color=UI_THEME["accent"], hover_color=UI_THEME["accent_hover"], command=_buscar).pack(side="left")
    termo_entry.bind("<Return>", _buscar)
    termo_entry.focus_set()
    _renderizar([], "")


def _abrir_dialogo_auditoria_consistencia():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Auditoria de Consistencia")
    centralizar_janela(dialog, 960, 680)
    dialog.grab_set()

    frame = ctk.CTkFrame(dialog, fg_color=UI_THEME["app_bg"])
    frame.pack(fill="both", expand=True, padx=16, pady=16)
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(2, weight=1)

    ctk.CTkLabel(frame, text="Auditoria de Consistencia", font=ctk.CTkFont(family="Segoe UI", size=21, weight="bold"), text_color=UI_THEME["text_primary"]).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(frame, text="Verifica dados suspeitos antes de exportar ou analisar relatorios.", font=ctk.CTkFont(family="Segoe UI", size=11), text_color=UI_THEME["text_secondary"]).grid(row=1, column=0, sticky="w", pady=(2, 10))

    lista = ctk.CTkScrollableFrame(frame, fg_color="transparent")
    lista.grid(row=2, column=0, sticky="nsew")

    def _cor_severidade(sev):
        return {"alta": "#D85B6A", "media": "#E0A422", "baixa": "#2F80D0"}.get(str(sev), "#65758B")

    def _renderizar():
        for widget in lista.winfo_children():
            widget.destroy()
        try:
            resultado = auditar_consistencia()
        except Exception as exc:
            messagebox.showerror("Auditoria de Consistencia", str(exc))
            return
        resumo = resultado.get("resumo", {})
        resumo_card = _criar_card(lista, corner_radius=15)
        resumo_card.pack(fill="x", padx=4, pady=(0, 8))
        resumo_card.grid_columnconfigure((0, 1, 2, 3), weight=1)
        for idx, (titulo, chave, cor) in enumerate((("Total", "total", UI_THEME["text_primary"]), ("Alta", "alta", "#D85B6A"), ("Media", "media", "#E0A422"), ("Baixa", "baixa", "#2F80D0"))):
            ctk.CTkLabel(resumo_card, text=titulo, font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), text_color=UI_THEME["text_secondary"]).grid(row=0, column=idx, pady=(12, 0))
            ctk.CTkLabel(resumo_card, text=str(resumo.get(chave, 0)), font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"), text_color=cor).grid(row=1, column=idx, pady=(0, 12))
        problemas = resultado.get("problemas", [])
        if not problemas:
            ctk.CTkLabel(lista, text="Nenhum problema encontrado nessa auditoria rapida.", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color="#1FAE66").pack(anchor="w", padx=8, pady=12)
            return
        for item in problemas[:120]:
            cor = _cor_severidade(item.get("severidade"))
            row = ctk.CTkFrame(lista, fg_color=UI_THEME["surface_alt"], corner_radius=13, border_width=1, border_color=UI_THEME["border"])
            row.pack(fill="x", padx=4, pady=5)
            row.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(row, text=str(item.get("severidade") or "-").upper(), fg_color=cor, corner_radius=10, text_color="#FFFFFF", width=70, font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold")).grid(row=0, column=0, rowspan=3, padx=10, pady=10)
            ctk.CTkLabel(row, text=f"{item.get('titulo')} | {item.get('documento')}", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color=UI_THEME["text_primary"], anchor="w").grid(row=0, column=1, sticky="ew", pady=(9, 0))
            ctk.CTkLabel(row, text=item.get("detalhe") or "-", font=ctk.CTkFont(family="Segoe UI", size=10), text_color=UI_THEME["text_secondary"], anchor="w").grid(row=1, column=1, sticky="ew", pady=(1, 0))
            ctk.CTkLabel(row, text=item.get("sugestao") or "-", font=ctk.CTkFont(family="Segoe UI", size=10), text_color=UI_THEME["text_secondary"], anchor="w").grid(row=2, column=1, sticky="ew", pady=(1, 9))

    botoes = ctk.CTkFrame(frame, fg_color="transparent")
    botoes.grid(row=3, column=0, sticky="ew", pady=(10, 0))
    botoes.grid_columnconfigure((0, 1, 2), weight=1)
    ctk.CTkButton(botoes, text="Atualizar auditoria", height=34, command=_renderizar).grid(row=0, column=0, sticky="ew", padx=(0, 5))
    ctk.CTkButton(botoes, text="Abrir Alteracoes", height=34, fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], command=lambda: (dialog.destroy(), app.mostrar_tela("alteracoes"))).grid(row=0, column=1, sticky="ew", padx=5)
    ctk.CTkButton(botoes, text="Fechar", height=34, fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], command=dialog.destroy).grid(row=0, column=2, sticky="ew", padx=(5, 0))
    _renderizar()


def _abrir_dialogo_restaurar_backup():
    backups = listar_backups()
    if not backups:
        messagebox.showinfo("Restaurar backup", "Nenhum backup local encontrado.")
        return

    dialog = ctk.CTkToplevel(app)
    dialog.title("Restaurar Backup")
    centralizar_janela(dialog, 760, 560)
    dialog.grab_set()

    frame = ctk.CTkFrame(dialog, fg_color=UI_THEME["app_bg"])
    frame.pack(fill="both", expand=True, padx=16, pady=16)
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(1, weight=1)

    ctk.CTkLabel(
        frame,
        text="Restaurar backup",
        font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w", pady=(0, 4))
    ctk.CTkLabel(
        frame,
        text="A restauracao sobrescreve o banco e os arquivos de sincronizacao. Um backup de seguranca sera criado antes.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=0, column=0, sticky="w", pady=(30, 10))

    lista = ctk.CTkScrollableFrame(frame, fg_color="transparent")
    lista.grid(row=1, column=0, sticky="nsew", pady=(8, 10))
    selecionado = {"pasta": backups[0]["pasta"]}

    def _selecionar(pasta):
        selecionado["pasta"] = pasta
        for widget in lista.winfo_children():
            try:
                ativo = getattr(widget, "_backup_pasta", "") == pasta
                widget.configure(border_color=UI_THEME["accent"] if ativo else UI_THEME["border"], border_width=2 if ativo else 1)
            except Exception:
                pass

    for item in backups[:20]:
        card = ctk.CTkFrame(lista, fg_color=UI_THEME["surface_alt"], corner_radius=13, border_width=1, border_color=UI_THEME["border"])
        card._backup_pasta = item["pasta"]
        card.pack(fill="x", padx=4, pady=5)
        card.grid_columnconfigure(0, weight=1)
        criado = str(item.get("criado_em") or "").replace("T", " ")
        qtd = len(item.get("arquivos") or [])
        ctk.CTkLabel(card, text=item.get("nome") or "Backup", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color=UI_THEME["text_primary"]).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 1))
        ctk.CTkLabel(card, text=f"{criado} | {qtd} arquivo(s) | {item.get('acionado_por') or '-'}", font=ctk.CTkFont(family="Segoe UI", size=10), text_color=UI_THEME["text_secondary"]).grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))
        ctk.CTkButton(card, text="Selecionar", width=86, height=30, command=lambda p=item["pasta"]: _selecionar(p)).grid(row=0, column=1, rowspan=2, padx=12, pady=10)
    _selecionar(selecionado["pasta"])

    def _restaurar():
        texto = "RESTAURAR"
        confirm = simpledialog.askstring(
            "Confirmar restauracao",
            f"Digite {texto} para confirmar.\n\nO app vai restaurar o backup selecionado e recomenda reiniciar em seguida.",
            parent=dialog,
        )
        if str(confirm or "").strip().upper() != texto:
            return
        try:
            resultado = restaurar_backup(selecionado["pasta"], criar_pre_restore=True)
        except Exception as exc:
            messagebox.showerror("Restaurar backup", str(exc))
            return
        _atualizar_cache_documentos_pos_alteracao()
        atualizar_dashboard()
        dialog.destroy()
        messagebox.showinfo(
            "Backup restaurado",
            "Backup restaurado com sucesso.\n\n"
            f"Arquivos restaurados: {len(resultado.get('restaurados') or [])}\n\n"
            "Feche e abra o aplicativo para garantir que todas as telas recarreguem com o banco restaurado.",
        )
        _recriar_tela_hoje()

    acoes = ctk.CTkFrame(frame, fg_color="transparent")
    acoes.grid(row=2, column=0, sticky="ew")
    acoes.grid_columnconfigure((0, 1, 2), weight=1)
    ctk.CTkButton(acoes, text="Abrir pasta de backups", height=36, fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], command=lambda: _abrir_pasta_no_explorer(_cfg.BACKUP_DIR)).grid(row=0, column=0, sticky="ew", padx=(0, 6))
    ctk.CTkButton(acoes, text="Restaurar selecionado", height=36, fg_color="#D85B6A", hover_color="#B94755", command=_restaurar).grid(row=0, column=1, sticky="ew", padx=6)
    ctk.CTkButton(acoes, text="Cancelar", height=36, fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], command=dialog.destroy).grid(row=0, column=2, sticky="ew", padx=(6, 0))



# ─── src/documentos.py ──────────────────────────────────────────────────────
# _buscar_documento_existente_sync, salvar_documento, alterar_competencia_documento,
# _normalizar_modalidade_frete, _coletar_ids_documentos_*, atualizar_modalidade_frete_documento,
# declarar_intercompany/delta/spot, registrar_substituicao, desfazer_substituicao,
# cancelar_documento, desfazer_cancelamento_documento
# ─────────────────────────────────────────────────────────────────────────────


# ─── src/relatorios.py ─────────────────────────────────────────────────────
# _obter_dataframe_relatorio_filtrado, _montar_dataframe_exportacao_periodo
# ─────────────────────────────────────────────────────────────────────────────

def abrir_relatorio():
    global relatorio_carregado, dados_importados

    base_disponivel = bool(relatorio_carregado)
    if not relatorio_selecionado:
        caminho_salvo = _resolver_ultimo_relatorio_salvo()
        if caminho_salvo:
            definir_relatorio_selecionado(caminho_salvo, persistir=True)

    if not relatorio_selecionado and not base_disponivel:
        mensagem = "Selecione um relatório para importar automaticamente antes de abrir."
        _atualizar_status_relatorios_ui(mensagem, tipo="neutral", total_registros=None)
        messagebox.showwarning("Relatório", mensagem)
        return

    try:
        data_inicial, data_final = obter_periodo_relatorios(silencioso=False)
    except ValueError as exc:
        _atualizar_status_relatorios_ui(str(exc), tipo="erro", total_registros=None)
        messagebox.showwarning("Filtro de Relatórios", str(exc))
        return

    if data_inicial is None or data_final is None:
        mensagem = "Preencha o período da aba Relatórios para abrir o arquivo filtrado."
        _atualizar_status_relatorios_ui(mensagem, tipo="erro", total_registros=None)
        messagebox.showwarning("Filtro de Relatórios", mensagem)
        return

    if data_inicial > data_final:
        mensagem = "A data inicial não pode ser maior que a data final."
        _atualizar_status_relatorios_ui(mensagem, tipo="erro", total_registros=None)
        messagebox.showwarning("Filtro de Relatórios", mensagem)
        return

    resultado = gerar_excel(data_inicial=data_inicial, data_final=data_final, exibir_mensagem=False)
    if not resultado.get("ok"):
        _atualizar_status_relatorios_ui(
            resultado.get("mensagem", "Não foi possível abrir o relatório."),
            tipo="erro",
            total_registros=0,
        )
        messagebox.showwarning("Relatório", resultado.get("mensagem", "Não foi possível abrir o relatório."))
        return

    nome = resultado.get("arquivo", "")
    try:
        os.startfile(nome)
    except OSError as e:
        messagebox.showerror("Erro", f"Não foi possível abrir o relatório: {e}")
        return

    relatorio_carregado = True
    if not isinstance(dados_importados, dict):
        dados_importados = {}
    dados_importados["ultimo_relatorio_filtrado"] = nome
    dados_importados["ultimo_periodo_relatorios"] = (
        data_inicial.strftime("%d/%m/%Y"),
        data_final.strftime("%d/%m/%Y"),
    )
    dados_importados["ultimo_total_registros"] = int(resultado.get("total_documentos", 0))
    _atualizar_status_relatorios_ui(
        f"Relatório aberto com {resultado.get('total_documentos', 0)} registro(s)",
        tipo="ok",
        total_registros=resultado.get("total_documentos", 0),
    )


def abrir_relatorio_filtrado():
    abrir_relatorio()



def exportar_relatorio_filtrado():
    global relatorio_carregado, dados_importados

    if not relatorio_carregado:
        mensagem = "Selecione e importe um relatório antes de exportar."
        _atualizar_status_relatorios_ui(mensagem, tipo="neutral", total_registros=None)
        messagebox.showwarning("Exportação", mensagem)
        return

    try:
        data_inicial, data_final = obter_periodo_relatorios(silencioso=False)
    except ValueError as exc:
        _atualizar_status_relatorios_ui(str(exc), tipo="erro", total_registros=None)
        messagebox.showwarning("Exportação", str(exc))
        return

    if data_inicial is None or data_final is None:
        mensagem = "Preencha o período da aba Relatórios antes de exportar."
        _atualizar_status_relatorios_ui(mensagem, tipo="erro", total_registros=None)
        messagebox.showwarning("Exportação", mensagem)
        return

    if data_inicial > data_final:
        mensagem = "A data inicial não pode ser maior que a data final."
        _atualizar_status_relatorios_ui(mensagem, tipo="erro", total_registros=None)
        messagebox.showwarning("Exportação", mensagem)
        return

    df_memoria = _obter_documentos_em_memoria(force=False)
    df_filtrado, msg_erro = _obter_dataframe_relatorio_filtrado(
        data_inicial,
        data_final,
        docs_df_base=df_memoria,
    )
    if msg_erro:
        _atualizar_status_relatorios_ui(msg_erro, tipo="erro", total_registros=0)
        messagebox.showwarning("Exportação", msg_erro)
        return
    if df_filtrado.empty:
        mensagem = "Nenhum registro encontrado para o período selecionado."
        _atualizar_status_relatorios_ui(mensagem, tipo="neutral", total_registros=0)
        messagebox.showinfo("Exportação", mensagem)
        return

    df_export = _montar_dataframe_exportacao_periodo(df_filtrado)
    if df_export.empty:
        mensagem = "Nenhum registro válido para exportação no período selecionado."
        _atualizar_status_relatorios_ui(mensagem, tipo="neutral", total_registros=0)
        messagebox.showinfo("Exportação", mensagem)
        return

    pasta_saida = obter_pasta_saida_relatorios()
    nome_arquivo = "Faturamento_AC.xlsx"
    caminho_saida = os.path.join(pasta_saida, nome_arquivo)

    try:
        with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Relatório Filtrado")
            ws = writer.sheets["Relatório Filtrado"]
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            for idx, coluna in enumerate(df_export.columns, start=1):
                maior = max(
                    len(str(coluna)),
                    int(df_export[coluna].astype(str).str.len().max()) if not df_export.empty else 0,
                )
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(maior + 2, 42)
            # Coluna E = Frete. Força largura mínima para INTERCOMPANY sem corte visual.
            ws.column_dimensions["E"].width = max(float(ws.column_dimensions["E"].width or 0), 18.0)
            for linha in range(2, len(df_export) + 2):
                ws[f"A{linha}"].number_format = "DD/MM/YYYY"
                ws[f"B{linha}"].number_format = '[$-pt-BR]mmmm/yyyy'
                ws[f"C{linha}"].number_format = "0"
                ws[f"E{linha}"].alignment = Alignment(horizontal="left", vertical="center")
                ws[f"F{linha}"].number_format = "R$ #,##0.00"
                ws[f"G{linha}"].number_format = "R$ #,##0.00"
    except Exception as exc:
        _atualizar_status_relatorios_ui("Falha ao exportar relatório", tipo="erro", total_registros=len(df_export))
        messagebox.showerror("Exportação", f"Erro ao exportar relatório do período.\n\n{exc}")
        return

    if not isinstance(dados_importados, dict):
        dados_importados = {}
    dados_importados["ultimo_export_periodo"] = caminho_saida
    dados_importados["ultimo_total_registros"] = int(len(df_export))

    _atualizar_status_relatorios_ui(
        f"Relatório atualizado com {len(df_export)} registro(s)",
        tipo="ok",
        total_registros=len(df_export),
    )
    messagebox.showinfo(
        "Exportação",
        "Relatório atualizado com sucesso.\n\n"
        f"Arquivo: {nome_arquivo}\n"
        f"Registros: {len(df_export)}",
    )


# ------------------------
# ATUALIZAR
# ------------------------

# ------------------------
# STATUS
# ------------------------

# ------------------------
# EXCEL
# ------------------------

def gerar_excel(data_inicial=None, data_final=None, exibir_mensagem=True):
    def _falha(msg, erro=False):
        if exibir_mensagem:
            if erro:
                messagebox.showerror("Excel", msg)
            else:
                messagebox.showwarning("Excel", msg)
        return {"ok": False, "mensagem": msg, "arquivo": "", "total_documentos": 0}

    if data_inicial is None or data_final is None:
        try:
            data_inicial, data_final = obter_periodo_relatorios(silencioso=False)
        except ValueError as erro_data:
            return _falha(str(erro_data))

    if data_inicial is None or data_final is None:
        return _falha("Período de Relatórios não configurado.")

    if data_inicial > data_final:
        return _falha("A data inicial não pode ser maior que a data final.")

    df, msg_erro = _obter_dataframe_relatorio_filtrado(
        data_inicial,
        data_final,
        docs_df_base=_obter_documentos_em_memoria(force=False),
    )
    if msg_erro:
        return _falha(msg_erro)
    if df.empty:
        return _falha(
            f"Não há documentos para o período selecionado"
            f" ({data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')})."
        )

    df = df.sort_values(["data_emissao", "numero"], ascending=[True, True])
    df["competencia_excel"] = df["data_competencia"]
    colunas_base = [
        "data_emissao", "competencia_excel", "numero", "numero_original_num",
        "tipo", "frete", "valor_inicial", "valor_final", "status",
    ]
    df_base = df[colunas_base].copy()

    pasta_saida = obter_pasta_saida_relatorios()
    nome = os.path.join(pasta_saida, "Faturamento_AC.xlsx")
    for antigo in glob.glob(os.path.join(pasta_saida, "Faturamento_AC*.xlsx")):
        try:
            if os.path.exists(antigo):
                os.remove(antigo)
        except (OSError, PermissionError):
            try:
                os.rename(antigo, antigo + ".bak")
            except Exception:
                pass

    resultado = escrever_excel_faturamento(df_base, nome)
    if not resultado["ok"]:
        return _falha(f"Erro ao gerar o relatório: {resultado['erro']}", erro=True)

    if exibir_mensagem:
        messagebox.showinfo(
            "Excel",
            f"Relatório gerado com {resultado['total_documentos']} documento(s) no período"
            f" {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}\n\n"
            f"Aba: Faturamento AC\nArquivo: {nome}",
        )
    return {
        "ok": True,
        "mensagem": "",
        "arquivo": nome,
        "total_documentos": resultado["total_documentos"],
        "periodo": (data_inicial.strftime("%d/%m/%Y"), data_final.strftime("%d/%m/%Y")),
    }


# ------------------------
# INTERFACE
# ------------------------

preparar_arquivos_aplicacao()

lock_ok, pid_existente = adquirir_lock_instancia()
if not lock_ok:
    alertar_instancia_em_execucao(pid_existente)
    raise SystemExit(0)

atexit.register(liberar_lock_instancia)

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.screens = {}
        self.current_screen = ""
        self.screen_host = None
        self.title("Sistema de Faturamento - Horizonte Logística")
        self.resizable(True, True)
        aplicar_icone_aplicacao(self)
        self.bind("<Configure>", _on_janela_configure)

    def registrar_tela(self, nome_tela, frame_tela):
        self.screens[nome_tela] = frame_tela

    def mostrar_tela(self, nome_tela):
        if nome_tela not in self.screens:
            return
        for nome, frame in self.screens.items():
            if frame.winfo_exists():
                if nome == nome_tela:
                    frame.pack(fill="both", expand=True, padx=2, pady=2)
                else:
                    frame.pack_forget()
        self.current_screen = nome_tela
        _configurar_nav_tela_ativa(nome_tela)
        # Reset scroll to top so short screens never show blank space above content
        try:
            canvas = ui_refs.get("scroll_canvas")
            if canvas is not None:
                canvas.yview_moveto(0)
        except Exception:
            pass


# DPI awareness por monitor (evita janela borrada ao arrastar entre telas)
try:
    import ctypes as _ctypes
    try:
        _ctypes.windll.shcore.SetProcessDpiAwareness(2)   # Per-monitor DPI aware v2
    except Exception:
        _ctypes.windll.user32.SetProcessDPIAware()        # Fallback
except Exception:
    pass

app = App()

largura_tela = app.winfo_screenwidth()
altura_tela = app.winfo_screenheight()

largura = max(860, min(1120, largura_tela - 80))
altura = max(650, min(860, altura_tela - 90))
largura = min(largura, largura_tela - 10)
altura = min(altura, altura_tela - 10)

centralizar_janela(app, largura, altura)


def _encerrar_aplicacao():
    liberar_lock_instancia()
    app.destroy()


app.protocol("WM_DELETE_WINDOW", _encerrar_aplicacao)

iniciar_banco()
importar_configuracoes_repo_inicial()
importar_estado_operacional_inicial()
criar_backup_automatico_silencioso("inicializacao")
carregar_pasta_saida_relatorios()
carregar_ultimo_relatorio()
inicializar_filtros_dashboard()
inicializar_filtros_relatorios()

APP_THEMES = {
    "light": {
        "app_bg": "#EDF3FA",
        "header_bg": "#E8F1FB",
        "surface": "#FFFFFF",
        "surface_alt": "#F2F7FD",
        "border": "#C9D8EA",
        "divider": "#CFDBEA",
        "text_primary": "#1A2A3B",
        "text_secondary": "#607286",
        "accent": "#2F80D0",
        "accent_hover": "#256FBD",
        "on_accent": "#FFFFFF",
        "success_bg": "#E6F4EB",
        "success_text": "#246B43",
        "danger_bg": "#FCEBED",
        "danger_text": "#A33B4B",
        "scroll_btn": "#BFD1E4",
        "scroll_btn_hover": "#A8BED7",
        "progress_bg": "#D7E4F3",
        "cta_border": "#2A70B8",
        "cta_border_hover": "#205D9E",
        "cta_press": "#1F5FA6",
        "tab_hover": "#E2EDF9",
        "tab_press": "#D4E4F7",
        "tab_active_hover": "#2A77C6",
        "tab_active_press": "#215EA5",
        "metric_initial_bg": "#E1ECFA",
        "metric_initial_title": "#245A94",
        "metric_initial_value": "#113F72",
        "metric_final_bg": "#E2F2E8",
        "metric_final_title": "#2E7950",
        "metric_final_value": "#1D5F3C",
        "metric_nf_bg": "#E5F5EC",
        "metric_nf_title": "#2E7950",
        "metric_nf_value": "#1F6340",
        "metric_docs_bg": "#E5EFFA",
        "metric_docs_title": "#245A94",
        "metric_docs_value": "#164A7E",
        "metric_cte_bg": "#EDF2F9",
        "metric_cte_title": "#415C79",
        "metric_cte_value": "#2E445F",
        "metric_cancelados_bg": "#EDF2F9",
        "metric_cancelados_title": "#415C79",
        "metric_cancelados_value": "#2E445F",
        "metric_default_bg": "#EDF2F9",
        "metric_default_title": "#415C79",
        "metric_default_value": "#2E445F",
        "chart_bg": "#FFFFFF",
        "chart_plot_bg": "#F6F9FD",
        "chart_grid": "#C8D6E7",
        "chart_axis": "#4C637D",
        "chart_bar_primary": "#2F80D0",
        "chart_bar_secondary": "#78AEDD",
        "chart_line": "#2A6EAF",
        "chart_cancelados": "#D56A7A",
    },
    "dark": {
        "app_bg": "#07111B",
        "header_bg": "#0D1A27",
        "surface": "#0E1B29",
        "surface_alt": "#132337",
        "border": "#263C52",
        "divider": "#22364A",
        "text_primary": "#F2F6FA",
        "text_secondary": "#A6B4C5",
        "accent": "#19C79A",
        "accent_hover": "#12AD86",
        "on_accent": "#F7FBFF",
        "success_bg": "#10392E",
        "success_text": "#38DFAE",
        "danger_bg": "#3A2026",
        "danger_text": "#F39AA8",
        "scroll_btn": "#30465E",
        "scroll_btn_hover": "#405A75",
        "progress_bg": "#23364B",
        "cta_border": "#1ED2A5",
        "cta_border_hover": "#43E0B9",
        "cta_press": "#0E9775",
        "tab_hover": "#172C41",
        "tab_press": "#1D3852",
        "tab_active_hover": "#20D4A8",
        "tab_active_press": "#10A47F",
        "metric_initial_bg": "#163150",
        "metric_initial_title": "#9CC5EC",
        "metric_initial_value": "#E6F2FF",
        "metric_final_bg": "#163829",
        "metric_final_title": "#9ED8B6",
        "metric_final_value": "#DDF6EA",
        "metric_nf_bg": "#193D2D",
        "metric_nf_title": "#9ED8B6",
        "metric_nf_value": "#D8F2E6",
        "metric_docs_bg": "#193552",
        "metric_docs_title": "#9CC5EC",
        "metric_docs_value": "#E6F2FF",
        "metric_cte_bg": "#1A2E43",
        "metric_cte_title": "#A9BED6",
        "metric_cte_value": "#E0EAF6",
        "metric_cancelados_bg": "#1A2E43",
        "metric_cancelados_title": "#A9BED6",
        "metric_cancelados_value": "#E0EAF6",
        "metric_default_bg": "#1A2E43",
        "metric_default_title": "#A9BED6",
        "metric_default_value": "#E0EAF6",
        "chart_bg": "#111B27",
        "chart_plot_bg": "#182433",
        "chart_grid": "#38506A",
        "chart_axis": "#A6BCD4",
        "chart_bar_primary": "#20D4B0",
        "chart_bar_secondary": "#0F899E",
        "chart_line": "#22D0AA",
        "chart_cancelados": "#E07D8D",
    },
}

tema_salvo = obter_configuracao("tema_interface", "dark").strip().lower()
current_theme_mode = tema_salvo if tema_salvo in APP_THEMES else "light"
UI_THEME = dict(APP_THEMES[current_theme_mode])
ctk.set_appearance_mode("dark" if current_theme_mode == "dark" else "light")


def _gerar_tab_styles():
    return {
        "normal_fg": UI_THEME["surface_alt"],
        "normal_text": UI_THEME["text_primary"],
        "hover_fg": UI_THEME["tab_hover"],
        "press_fg": UI_THEME["tab_press"],
        "active_fg": UI_THEME["accent"],
        "active_hover_fg": UI_THEME["tab_active_hover"],
        "active_press_fg": UI_THEME["tab_active_press"],
        "active_text": UI_THEME["on_accent"],
    }


TAB_STYLES = _gerar_tab_styles()
app.configure(fg_color=UI_THEME["app_bg"])


def abrir_dialogo_alterar_competencia():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Alterar competencia de documento")
    centralizar_janela(dialog, 620, 270)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Tipo de documento").grid(row=0, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Numero do documento").grid(row=0, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=260)
    tipo_combo.set("NF")
    tipo_combo.grid(row=1, column=0, sticky="ew", padx=6, pady=(0, 10))

    numero_entry = ctk.CTkEntry(form, width=260, placeholder_text="Ex.: 89 ou 2390")
    numero_entry.grid(row=1, column=1, sticky="ew", padx=6, pady=(0, 10))

    ctk.CTkLabel(form, text="Novo mês de competência").grid(row=2, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Ano da competência").grid(row=2, column=1, sticky="w", padx=6, pady=(0, 4))

    mes_combo = ctk.CTkComboBox(form, values=MESES, width=260)
    mes_combo.set(MESES[datetime.now().month - 1])
    mes_combo.grid(row=3, column=0, sticky="ew", padx=6)

    anos_competencia = [str(ano) for ano in range(datetime.now().year - 3, datetime.now().year + 5)]
    ano_combo = ctk.CTkComboBox(form, values=anos_competencia, width=260)
    ano_combo.set(str(datetime.now().year))
    ano_combo.grid(row=3, column=1, sticky="ew", padx=6)

    def salvar_alteracao():
        tipo = tipo_combo.get().strip().upper()
        mes_novo = mes_combo.get().strip().lower()
        ano_novo = ano_combo.get().strip()
        numero_texto = numero_entry.get().strip()

        if tipo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento inválido.")
            return

        if mes_novo not in MESES:
            messagebox.showwarning("Aviso", "Selecione um mês válido.")
            return

        if not numero_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe um número de documento válido.")
            return

        if not ano_novo.isdigit():
            messagebox.showwarning("Aviso", "Informe um ano válido.")
            return

        numero = int(numero_texto)
        alterados = alterar_competencia_documento(tipo, numero, mes_novo, int(ano_novo))
        if alterados == 0:
            messagebox.showwarning("Aviso", "Documento não encontrado para alterar competência.")
            return

        messagebox.showinfo("Sucesso", "Competência atualizada com sucesso.")
        dialog.destroy()

    ctk.CTkButton(form, text="Salvar alteração", command=salvar_alteracao, width=200).grid(row=4, column=0, columnspan=2, pady=(14, 0))


def _abrir_dialogo_declarar_frete(rotulo_frete):
    dialog = ctk.CTkToplevel(app)
    dialog.title(rotulo_frete)
    centralizar_janela(dialog, 620, 220)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Tipo de documento").grid(row=0, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Número do documento").grid(row=0, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=260)
    tipo_combo.set("NF")
    tipo_combo.grid(row=1, column=0, sticky="ew", padx=6)

    numero_entry = ctk.CTkEntry(form, width=260, placeholder_text="Ex.: 92 ou 2390")
    numero_entry.grid(row=1, column=1, sticky="ew", padx=6)

    desfazer_var = ctk.BooleanVar(value=False)
    texto_desfazer = f"Des{rotulo_frete.lower()} (voltar para FRANQUIA)"
    ctk.CTkCheckBox(
        form,
        text=texto_desfazer,
        variable=desfazer_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=2, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 0))

    def confirmar():
        tipo = tipo_combo.get().strip().upper()
        numero_texto = numero_entry.get().strip()

        if tipo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento inválido.")
            return
        if not numero_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe um número de documento válido.")
            return

        numero_doc = int(numero_texto)
        acao = rotulo_frete.upper().strip()

        if desfazer_var.get():
            # Desfaz a modalidade manual e retorna para FRANQUIA.
            resultado = salvar_alteracao_frete_manual(tipo, numero_doc, "FRANQUIA")
            if resultado.get("encontrados", 0) == 0:
                messagebox.showwarning("Aviso", f"Documento não encontrado para des{rotulo_frete.lower()}.")
                return
            messagebox.showinfo(
                "Sucesso",
                f"Des{rotulo_frete.lower()} aplicado com sucesso.\n"
                "A modalidade de frete foi revertida para FRANQUIA.\n\n"
                "A alteração foi salva e será refletida nas próximas consultas e exportações.",
            )
        else:
            # Declaração manual da modalidade.
            if acao == "INTERCOMPANY":
                resultado = declarar_intercompany(tipo, numero_doc)
            elif acao == "DELTA":
                resultado = declarar_delta(tipo, numero_doc)
            elif acao == "SPOT":
                resultado = declarar_spot(tipo, numero_doc)
            else:
                resultado = salvar_alteracao_frete_manual(tipo, numero_doc, acao)

            modalidade = str(resultado.get("modalidade", _normalizar_modalidade_frete(acao))).upper()
            if resultado.get("encontrados", 0) == 0:
                messagebox.showwarning("Aviso", f"Documento não encontrado para {rotulo_frete.lower()}.")
                return
            messagebox.showinfo(
                "Sucesso",
                f"{modalidade} aplicado com sucesso.\n"
                f"A modalidade de frete do documento agora é {modalidade}.\n\n"
                "A alteração foi salva e será refletida nas próximas consultas e exportações.",
            )
        dialog.destroy()

    ctk.CTkButton(form, text="Confirmar", command=confirmar, width=220).grid(row=3, column=0, columnspan=2, pady=(14, 0))


def abrir_dialogo_declarar_intercompany():
    _abrir_dialogo_declarar_frete("Intercompany")


def abrir_dialogo_declarar_delta():
    _abrir_dialogo_declarar_frete("Delta")


def abrir_dialogo_declarar_spot():
    _abrir_dialogo_declarar_frete("Spot")


def abrir_dialogo_cancelar_documento():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Cancelar documento")
    centralizar_janela(dialog, 620, 220)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Tipo de documento").grid(row=0, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Número do documento").grid(row=0, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=260)
    tipo_combo.set("NF")
    tipo_combo.grid(row=1, column=0, sticky="ew", padx=6)

    numero_entry = ctk.CTkEntry(form, width=260, placeholder_text="Ex.: 89 ou 2390")
    numero_entry.grid(row=1, column=1, sticky="ew", padx=6)

    desfazer_var = ctk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        form,
        text="Desfazer cancelamento",
        variable=desfazer_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=2, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 0))

    def confirmar_cancelamento():
        tipo = tipo_combo.get().strip().upper()
        numero_texto = numero_entry.get().strip()

        if tipo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento inválido.")
            return

        if not numero_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe um número de documento válido.")
            return

        numero = int(numero_texto)
        if desfazer_var.get():
            alterados = desfazer_cancelamento_documento(tipo, numero)
            if alterados == 0:
                messagebox.showwarning("Aviso", "Documento não encontrado ou não está cancelado.")
                return
            messagebox.showinfo("Sucesso", "Cancelamento desfeito com sucesso.")
        else:
            alterados = cancelar_documento(tipo, numero)
            if alterados == 0:
                messagebox.showwarning("Aviso", "Documento não encontrado.")
                return
            messagebox.showinfo("Sucesso", "Documento cancelado com sucesso.")
        dialog.destroy()

    ctk.CTkButton(form, text="Confirmar", command=confirmar_cancelamento, width=200).grid(row=3, column=0, columnspan=2, pady=(14, 0))





def abrir_dialogo_substituir_documento():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Substituir documento")
    centralizar_janela(dialog, 700, 320)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=16, pady=12)
    form.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkLabel(form, text="Documento antigo", font=ctk.CTkFont(weight="bold")).grid(
        row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(0, 4)
    )
    ctk.CTkLabel(form, text="Tipo").grid(row=1, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Número").grid(row=1, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_antigo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=320)
    tipo_antigo_combo.set("NF")
    tipo_antigo_combo.grid(row=2, column=0, sticky="ew", padx=6, pady=(0, 10))

    numero_antigo_entry = ctk.CTkEntry(form, width=320, placeholder_text="Ex.: 89 ou 2390")
    numero_antigo_entry.grid(row=2, column=1, sticky="ew", padx=6, pady=(0, 10))

    ctk.CTkLabel(form, text="Documento substituto", font=ctk.CTkFont(weight="bold")).grid(
        row=3, column=0, columnspan=2, sticky="w", padx=6, pady=(2, 4)
    )
    ctk.CTkLabel(form, text="Tipo").grid(row=4, column=0, sticky="w", padx=6, pady=(0, 4))
    ctk.CTkLabel(form, text="Número").grid(row=4, column=1, sticky="w", padx=6, pady=(0, 4))

    tipo_novo_combo = ctk.CTkComboBox(form, values=["NF", "CTE"], width=320)
    tipo_novo_combo.set("NF")
    tipo_novo_combo.grid(row=5, column=0, sticky="ew", padx=6)

    numero_novo_entry = ctk.CTkEntry(form, width=320, placeholder_text="Ex.: 90 ou 2391")
    numero_novo_entry.grid(row=5, column=1, sticky="ew", padx=6)

    desfazer_var = ctk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        form,
        text="Desfazer substituicao",
        variable=desfazer_var,
        onvalue=True,
        offvalue=False,
    ).grid(row=6, column=0, columnspan=2, sticky="w", padx=6, pady=(10, 0))

    def confirmar_substituicao():
        tipo_antigo = tipo_antigo_combo.get().strip().upper()
        tipo_novo = tipo_novo_combo.get().strip().upper()
        numero_antigo_texto = numero_antigo_entry.get().strip()
        numero_novo_texto = numero_novo_entry.get().strip()

        if tipo_antigo not in {"NF", "CTE"} or tipo_novo not in {"NF", "CTE"}:
            messagebox.showwarning("Aviso", "Tipo de documento inválido.")
            return

        if not numero_antigo_texto.isdigit() or not numero_novo_texto.isdigit():
            messagebox.showwarning("Aviso", "Informe números de documento válidos.")
            return

        numero_antigo = int(numero_antigo_texto)
        numero_novo = int(numero_novo_texto)

        if tipo_antigo == tipo_novo and numero_antigo == numero_novo:
            messagebox.showwarning("Aviso", "Documento antigo e substituto não podem ser iguais.")
            return

        if desfazer_var.get():
            antigo_restaurado, novo_restaurado = desfazer_substituicao(
                tipo_antigo, numero_antigo, tipo_novo, numero_novo
            )
            if antigo_restaurado == 0:
                messagebox.showwarning("Aviso", "Documento antigo não encontrado ou não está substituído.")
                return
            if novo_restaurado == 0:
                messagebox.showwarning("Aviso", "Documento substituto não encontrado ou não está como substituto.")
                return
            messagebox.showinfo("Sucesso", "Substituição desfeita com sucesso.")
        else:
            novo_alterado, antigo_alterado = registrar_substituicao(
                tipo_antigo, numero_antigo, tipo_novo, numero_novo
            )
            if novo_alterado == 0:
                messagebox.showwarning("Aviso", "Documento substituto não encontrado.")
                return
            if antigo_alterado == 0:
                messagebox.showwarning("Aviso", "Documento antigo não encontrado.")
                return
            messagebox.showinfo("Sucesso", "Substituição registrada com sucesso.")
        dialog.destroy()

    ctk.CTkButton(form, text="Confirmar", command=confirmar_substituicao, width=220).grid(
        row=7, column=0, columnspan=2, pady=(14, 0)
    )


def abrir_relatorio_cancelados():
    try:
        data_inicial, data_final = obter_periodo_relatorios(silencioso=False)
    except ValueError as exc:
        messagebox.showwarning("Filtro de período", str(exc))
        return

    if data_inicial is None or data_final is None:
        messagebox.showwarning("Filtro de período", "Preencha as datas na aba Relatórios.")
        return

    if data_inicial > data_final:
        messagebox.showwarning("Filtro de período", "A data inicial não pode ser maior que a data final.")
        return

    conn = obter_conexao_banco()
    df = pd.read_sql_query(
        """
        SELECT id, tipo, numero, numero_original, data_emissao, valor_final, status
        FROM documentos
        WHERE UPPER(status) LIKE '%CANCELADO%'
        """,
        conn,
    )
    conn.close()

    df["data_emissao_dt"] = pd.to_datetime(df["data_emissao"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["data_emissao_dt"])
    df = df[(df["data_emissao_dt"] >= data_inicial) & (df["data_emissao_dt"] <= data_final)]
    df["numero_exibicao"] = df.apply(
        lambda r: _numero_documento_exibicao(r["tipo"], r["numero"], r.get("numero_original", ""), r.get("data_emissao", "")),
        axis=1,
    )
    df["chave_documento"] = df.apply(
        lambda r: _chave_documento_compativel(r["tipo"], r["numero"], r.get("numero_original", "")),
        axis=1,
    )
    df = df.sort_values(["id"]).drop_duplicates(subset=["chave_documento"], keep="last")

    if df.empty:
        messagebox.showinfo(
            "Relatório",
            (
                "Nenhum documento cancelado encontrado no período selecionado.\n\n"
                f"Período: {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}"
            ),
        )
        return

    df = df.sort_values(["data_emissao_dt", "numero"], ascending=[True, True])

    janela = ctk.CTkToplevel(app)
    janela.title("Relatório de documentos cancelados")
    centralizar_janela(janela, 860, 520)
    janela.grab_set()

    ctk.CTkLabel(
        janela,
        text=(
            "Documentos Cancelados\n"
            f"{data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}"
        ),
        font=ctk.CTkFont(size=16, weight="bold"),
    ).pack(pady=(12, 8))

    container = ctk.CTkFrame(janela)
    container.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    style = ttk.Style()
    style.configure("Cancelados.Treeview", rowheight=28, borderwidth=1, relief="solid")
    style.configure(
        "Cancelados.Treeview.Heading",
        font=("Segoe UI", 10, "bold"),
        relief="solid",
        borderwidth=1,
    )

    colunas = ("tipo", "numero", "data_emissao", "valor", "status")
    tabela = ttk.Treeview(container, columns=colunas, show="headings", style="Cancelados.Treeview")

    tabela.heading("tipo", text="Tipo Doc")
    tabela.heading("numero", text="Numero Doc")
    tabela.heading("data_emissao", text="Data Emissao")
    tabela.heading("valor", text="Valor")
    tabela.heading("status", text="Status")

    tabela.column("tipo", width=110, anchor="center")
    tabela.column("numero", width=170, anchor="center")
    tabela.column("data_emissao", width=170, anchor="center")
    tabela.column("valor", width=180, anchor="e")
    tabela.column("status", width=240, anchor="center")

    scroll_y = ttk.Scrollbar(container, orient="vertical", command=tabela.yview)
    scroll_x = ttk.Scrollbar(container, orient="horizontal", command=tabela.xview)
    tabela.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    tabela.grid(row=0, column=0, sticky="nsew")
    scroll_y.grid(row=0, column=1, sticky="ns")
    scroll_x.grid(row=1, column=0, sticky="ew")

    container.grid_rowconfigure(0, weight=1)
    container.grid_columnconfigure(0, weight=1)

    for _, linha in df.iterrows():
        data_txt = (
            linha["data_emissao_dt"].strftime("%d/%m/%Y")
            if pd.notna(linha["data_emissao_dt"])
            else str(linha["data_emissao"])
        )
        valor = float(linha["valor_final"]) if pd.notna(linha["valor_final"]) else 0.0
        tabela.insert(
            "",
            "end",
            values=(
                str(linha["tipo"]),
                str(linha["numero_exibicao"]),
                data_txt,
                formatar_moeda_brl(valor),
                str(linha["status"]),
            ),
        )




def abrir_busca_documentos():

    dialog = ctk.CTkToplevel(app)
    dialog.title("Buscar documentos")
    centralizar_janela(dialog, 750, 500)
    dialog.grab_set()

    ctk.CTkLabel(
        dialog,
        text="Buscar documento",
        font=ctk.CTkFont(size=16, weight="bold"),
    ).pack(pady=(12, 8))

    busca_entry = ctk.CTkEntry(dialog, width=300, placeholder_text="Digite o numero original do documento")
    busca_entry.pack(pady=(0, 10))

    container = ctk.CTkFrame(dialog)
    container.pack(fill="both", expand=True, padx=10, pady=10)

    tabela = ttk.Treeview(
        container,
        columns=("tipo","numero","data","valor","status"),
        show="headings"
    )

    tabela.heading("tipo", text="Tipo")
    tabela.heading("numero", text="Numero")
    tabela.heading("data", text="Data")
    tabela.heading("valor", text="Valor")
    tabela.heading("status", text="Status")

    tabela.column("tipo", width=80, anchor="center")
    tabela.column("numero", width=120, anchor="center")
    tabela.column("data", width=120, anchor="center")
    tabela.column("valor", width=120, anchor="e")
    tabela.column("status", width=200, anchor="center")

    scroll = ttk.Scrollbar(container, orient="vertical", command=tabela.yview)
    tabela.configure(yscrollcommand=scroll.set)

    tabela.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    def buscar():

        termo = busca_entry.get().strip()

        for item in tabela.get_children():
            tabela.delete(item)

        conn = obter_conexao_banco()

        df = pd.read_sql_query(
            """
            SELECT id, tipo, numero, numero_original, data_emissao, valor_final, status
            FROM documentos
            WHERE CAST(numero AS TEXT) LIKE ?
               OR COALESCE(numero_original, '') LIKE ?
            ORDER BY data_emissao, id
            """,
            conn,
            params=(f"%{termo}%", f"%{termo}%"),
        )

        conn.close()
        if df.empty:
            return

        df["numero_exibicao"] = df.apply(
            lambda r: _numero_documento_exibicao(r["tipo"], r["numero"], r.get("numero_original", ""), r.get("data_emissao", "")),
            axis=1,
        )
        df["chave_documento"] = df.apply(
            lambda r: _chave_documento_compativel(r["tipo"], r["numero"], r.get("numero_original", "")),
            axis=1,
        )
        df = df.sort_values(["id"]).drop_duplicates(subset=["chave_documento"], keep="last")

        for _, row in df.iterrows():

            tabela.insert(
                "",
                "end",
                values=(
                    row["tipo"],
                    row["numero_exibicao"],
                    row["data_emissao"],
                    formatar_moeda_brl(row["valor_final"]),
                    row["status"],
                ),
            )

    ctk.CTkButton(dialog, text="Buscar", command=buscar, width=200).pack(pady=6)


# ─── _competencia_para_data  →  src/utils.py

def _obter_dataframe_dashboard_filtrado():
    data_inicial, data_final = obter_periodo_dashboard(silencioso=True)
    return obter_dataframe_dashboard(data_inicial, data_final)


def _garantir_matplotlib_dashboard():
    if dashboard_chart_state.get("import_ok") is not None:
        return bool(dashboard_chart_state.get("import_ok"))

    try:
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.ticker import FuncFormatter

        dashboard_chart_state.update(
            {
                "import_ok": True,
                "import_error": "",
                "plt": plt,
                "FigureCanvasTkAgg": FigureCanvasTkAgg,
                "FuncFormatter": FuncFormatter,
            }
        )
    except Exception as exc:
        dashboard_chart_state.update(
            {
                "import_ok": False,
                "import_error": str(exc),
                "plt": None,
                "FigureCanvasTkAgg": None,
                "FuncFormatter": None,
            }
        )
    return bool(dashboard_chart_state.get("import_ok"))


def _liberar_graficos_dashboard():
    plt = dashboard_chart_state.get("plt")
    for cfg in dashboard_chart_widgets.values():
        fig = cfg.get("fig")
        if fig is not None and plt is not None:
            try:
                plt.close(fig)
            except Exception:
                pass
        canvas = cfg.get("canvas")
        if canvas is not None:
            try:
                widget = canvas.get_tk_widget()
                if widget.winfo_exists():
                    widget.destroy()
            except Exception:
                pass
    dashboard_chart_widgets.clear()


def _limpar_host_grafico(host):
    if host is None or not host.winfo_exists():
        return
    for child in host.winfo_children():
        try:
            child.destroy()
        except Exception:
            pass


def _mostrar_placeholder_grafico(chave, mensagem):
    cfg = dashboard_chart_widgets.get(chave)
    if not cfg:
        return

    _limpar_host_grafico(cfg.get("host"))
    lbl = ctk.CTkLabel(
        cfg["host"],
        text=mensagem,
        justify="center",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_secondary"],
    )
    lbl.pack(expand=True, fill="both", padx=8, pady=8)
    cfg["placeholder"] = lbl

    fig_antiga = cfg.get("fig")
    if fig_antiga is not None and dashboard_chart_state.get("plt") is not None:
        try:
            dashboard_chart_state["plt"].close(fig_antiga)
        except Exception:
            pass
    cfg["fig"] = None
    cfg["canvas"] = None


def _renderizar_figura_dashboard(chave, fig):
    cfg = dashboard_chart_widgets.get(chave)
    if not cfg:
        if dashboard_chart_state.get("plt") is not None:
            try:
                dashboard_chart_state["plt"].close(fig)
            except Exception:
                pass
        return

    host = cfg.get("host")
    if host is None or not host.winfo_exists():
        if dashboard_chart_state.get("plt") is not None:
            try:
                dashboard_chart_state["plt"].close(fig)
            except Exception:
                pass
        return

    fig_antiga = cfg.get("fig")
    if fig_antiga is not None and dashboard_chart_state.get("plt") is not None:
        try:
            dashboard_chart_state["plt"].close(fig_antiga)
        except Exception:
            pass

    _limpar_host_grafico(host)
    canvas = dashboard_chart_state["FigureCanvasTkAgg"](fig, master=host)
    canvas.draw()
    widget_canvas = canvas.get_tk_widget()
    try:
        widget_canvas.configure(
            bg=UI_THEME["chart_bg"],
            highlightthickness=0,
            bd=0,
        )
    except Exception:
        pass
    widget_canvas.pack(fill="both", expand=True, padx=8, pady=8)
    cfg["fig"] = fig
    cfg["canvas"] = canvas
    cfg["placeholder"] = None


# ─── criar_figura_faturamento_periodo  →  src/dashboard.py

# ─── criar_figura_comparativo_tipos  →  src/dashboard.py

def _atualizar_graficos_dashboard(df, data_inicial=None, data_final=None):
    if not dashboard_chart_widgets:
        return

    if not _garantir_matplotlib_dashboard():
        erro = dashboard_chart_state.get("import_error") or "Erro desconhecido ao carregar matplotlib."
        _mostrar_placeholder_grafico("faturamento", f"Matplotlib indisponivel.\n{erro}")
        _mostrar_placeholder_grafico("comparativo", f"Matplotlib indisponivel.\n{erro}")
        return

    if data_inicial is not None and data_final is not None and data_inicial > data_final:
        msg = "Período inválido.\nA data inicial precisa ser menor ou igual à data final."
        _mostrar_placeholder_grafico("faturamento", msg)
        _mostrar_placeholder_grafico("comparativo", msg)
        return

    if df is None:
        return

    if df.empty:
        msg = "Sem dados para o período selecionado."
        _mostrar_placeholder_grafico("faturamento", msg)
        _mostrar_placeholder_grafico("comparativo", msg)
        return

    try:
        plt_obj = dashboard_chart_state["plt"]
        fig_faturamento = criar_figura_faturamento_periodo(df, plt_obj, UI_THEME)
        fig_comparativo = criar_figura_comparativo_tipos(df, plt_obj, UI_THEME)
    except Exception as exc:
        _mostrar_placeholder_grafico("faturamento", f"Falha ao desenhar gráfico.\n{exc}")
        _mostrar_placeholder_grafico("comparativo", f"Falha ao desenhar gráfico.\n{exc}")
        return

    _renderizar_figura_dashboard("faturamento", fig_faturamento)
    _renderizar_figura_dashboard("comparativo", fig_comparativo)


def abrir_grafico_faturamento():
    try:
        import matplotlib.pyplot as plt
        from matplotlib.colors import LinearSegmentedColormap
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.patches import FancyBboxPatch, Rectangle
    except Exception as exc:
        messagebox.showerror("Gráfico", f"Falha ao carregar bibliotecas do gráfico.\n\n{exc}")
        return

    conn = obter_conexao_banco()

    df = pd.read_sql_query(
        """
        SELECT data_emissao, valor_inicial
        FROM documentos
        WHERE status NOT LIKE '%CANCELADO%'
        """,
        conn,
    )

    conn.close()

    if df.empty:
        messagebox.showwarning("Gráfico", "Não há dados para gerar gráfico.")
        return

    df["data_emissao"] = pd.to_datetime(df["data_emissao"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["data_emissao"])
    df["valor_inicial"] = pd.to_numeric(df["valor_inicial"], errors="coerce").fillna(0)

    df["mes"] = df["data_emissao"].dt.to_period("M")

    resumo_base = df.groupby("mes")["valor_inicial"].sum().reset_index().sort_values("mes").reset_index(drop=True)
    resumo_base = resumo_base[resumo_base["valor_inicial"].abs() > 0.0001].copy().reset_index(drop=True)
    resumo_base["mes_dt"] = resumo_base["mes"].dt.to_timestamp()
    meses_abrev = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

    def _formatar_mes_ano(valor_dt):
        return f"{meses_abrev[valor_dt.month - 1]}/{valor_dt.strftime('%y')}"

    resumo_base["mes_label"] = resumo_base["mes_dt"].apply(_formatar_mes_ano)
    ano_atual = datetime.now().year

    if resumo_base.empty:
        messagebox.showwarning("Gráfico", "Não há dados para gerar gráfico.")
        return

    janela = ctk.CTkToplevel(app)
    janela.title("Faturamento mensal")
    centralizar_janela(janela, 1020, 620)
    janela.grab_set()

    conteudo = ctk.CTkFrame(janela, fg_color="transparent")
    conteudo.pack(fill="both", expand=True, padx=12, pady=12)

    filtro_card = ctk.CTkFrame(conteudo, fg_color=UI_THEME["surface"], corner_radius=12)
    filtro_card.pack(fill="x", pady=(0, 6))

    topo_filtro = ctk.CTkFrame(filtro_card, fg_color="transparent")
    topo_filtro.pack(fill="x", padx=10, pady=10)

    ctk.CTkLabel(
        topo_filtro,
        text="Filtro",
        font=("Segoe UI", 13, "bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(side="left")

    info_meses_label = ctk.CTkLabel(
        topo_filtro,
        text="",
        font=("Segoe UI", 11),
        text_color=UI_THEME["text_secondary"],
    )
    info_meses_label.pack(side="left", padx=(12, 8))

    botao_meses = ctk.CTkButton(
        topo_filtro,
        text="Meses ▾",
        width=130,
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        text_color=UI_THEME["on_accent"],
        border_width=1,
        border_color=UI_THEME["cta_border"],
    )
    botao_meses.pack(side="right")

    dropdown_meses_frame = ctk.CTkFrame(
        conteudo,
        fg_color=UI_THEME["surface_alt"],
        corner_radius=10,
        border_width=1,
        border_color=UI_THEME["border"],
        width=250,
        height=196,
    )
    dropdown_meses_frame.pack_propagate(False)
    altura_lista_meses = min(112, max(72, len(resumo_base) * 24))
    frame_checkboxes = ctk.CTkScrollableFrame(
        dropdown_meses_frame,
        height=altura_lista_meses,
        fg_color=UI_THEME["surface_alt"],
        corner_radius=10,
    )
    frame_checkboxes.pack(fill="both", expand=True, padx=8, pady=(8, 6))

    check_vars = {}
    periodos_disponiveis = resumo_base["mes"].tolist()
    periodos_ano_atual = [p for p in periodos_disponiveis if p.year == ano_atual]
    periodos_padrao = set(periodos_ano_atual if periodos_ano_atual else periodos_disponiveis)

    for _, linha in resumo_base.iterrows():
        periodo = linha["mes"]
        check_vars[periodo] = tk.BooleanVar(value=(periodo in periodos_padrao))
        ctk.CTkCheckBox(
            frame_checkboxes,
            text=linha["mes_label"],
            variable=check_vars[periodo],
            onvalue=True,
            offvalue=False,
            command=lambda: _desenhar_grafico(fechar_dropdown=False, animar=True),
            font=("Segoe UI", 11),
            text_color=UI_THEME["text_primary"],
            checkbox_height=18,
            checkbox_width=18,
        ).pack(anchor="w", padx=8, pady=4)

    acoes_dropdown = ctk.CTkFrame(dropdown_meses_frame, fg_color="transparent")
    acoes_dropdown.pack(fill="x", padx=8, pady=(0, 8))

    botao_marcar_todos = ctk.CTkButton(
        acoes_dropdown,
        text="Marcar todos",
        width=120,
        fg_color=UI_THEME["tab_hover"],
        hover_color=UI_THEME["tab_press"],
        text_color=UI_THEME["text_primary"],
        border_width=1,
        border_color=UI_THEME["border"],
    )
    botao_marcar_todos.pack(side="left", padx=(0, 8))

    botao_limpar = ctk.CTkButton(
        acoes_dropdown,
        text="Limpar",
        width=90,
        fg_color=UI_THEME["surface"],
        hover_color=UI_THEME["tab_hover"],
        text_color=UI_THEME["text_primary"],
        border_width=1,
        border_color=UI_THEME["border"],
    )
    botao_limpar.pack(side="left")

    dropdown_visivel = False

    frame_grafico = ctk.CTkFrame(conteudo, fg_color="transparent")
    frame_grafico.pack(fill="both", expand=True)

    fig, ax = plt.subplots(figsize=(9.0, 4.8), dpi=110)
    fig.patch.set_facecolor(UI_THEME["chart_bg"])
    canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
    widget_canvas_popup = canvas.get_tk_widget()
    try:
        widget_canvas_popup.configure(
            bg=UI_THEME["chart_bg"],
            highlightthickness=0,
            bd=0,
        )
    except Exception:
        pass
    widget_canvas_popup.pack(fill="both", expand=True)
    animacao_estado = {"job": None}

    def _misturar_cor(cor_a, cor_b, t):
        c1 = _hex_para_rgb(cor_a)
        c2 = _hex_para_rgb(cor_b)
        mix = tuple(int(c1[i] + (c2[i] - c1[i]) * t) for i in range(3))
        return _rgb_para_hex(mix)

    def _fmt_brl_exato(v):
        return formatar_moeda_brl_exata(v)

    def _gerar_cores_gradiente(qtd):
        if qtd <= 0:
            return []
        if qtd == 1:
            return [_misturar_cor(UI_THEME["chart_bar_secondary"], UI_THEME["chart_bar_primary"], 0.6)]
        return [
            _misturar_cor(
                UI_THEME["chart_bar_secondary"],
                UI_THEME["chart_bar_primary"],
                i / (qtd - 1),
            )
            for i in range(qtd)
        ]

    def _calcular_tendencia_linear(valores_seq):
        n = len(valores_seq)
        if n < 2:
            return []
        xs = list(range(n))
        soma_x = sum(xs)
        soma_y = sum(valores_seq)
        soma_x2 = sum(x * x for x in xs)
        soma_xy = sum(x * y for x, y in zip(xs, valores_seq))
        denominador = (n * soma_x2) - (soma_x * soma_x)
        if denominador == 0:
            return []
        m = ((n * soma_xy) - (soma_x * soma_y)) / denominador
        b = (soma_y - (m * soma_x)) / n
        return [(m * x) + b for x in xs]

    def _obter_resumo_filtrado():
        periodos_marcados = [periodo for periodo, var in check_vars.items() if bool(var.get())]
        if not periodos_marcados:
            return pd.DataFrame(columns=list(resumo_base.columns) + ["idx"])
        resumo_filtrado = resumo_base[resumo_base["mes"].isin(periodos_marcados)].copy()
        resumo_filtrado = resumo_filtrado.sort_values("mes").reset_index(drop=True)
        resumo_filtrado["idx"] = range(len(resumo_filtrado))
        return resumo_filtrado

    def _atualizar_resumo_meses():
        meses_selecionados = [
            linha["mes_label"] for _, linha in resumo_base.iterrows() if bool(check_vars[linha["mes"]].get())
        ]
        if not meses_selecionados:
            info_meses_label.configure(text="Nenhum mes selecionado")
            return
        if len(meses_selecionados) <= 3:
            info_meses_label.configure(text=", ".join(meses_selecionados))
            return
        info_meses_label.configure(text=f"{len(meses_selecionados)} meses selecionados")

    def _posicionar_dropdown():
        if not dropdown_visivel:
            return
        try:
            conteudo.update_idletasks()
            filtro_card.update_idletasks()
            y_base = filtro_card.winfo_y() + filtro_card.winfo_height() + 2
            dropdown_meses_frame.place(relx=1.0, x=-10, y=y_base, anchor="ne")
            dropdown_meses_frame.lift()
        except Exception:
            pass

    def _alternar_dropdown_meses():
        nonlocal dropdown_visivel
        if dropdown_visivel:
            dropdown_meses_frame.place_forget()
            dropdown_visivel = False
            botao_meses.configure(text="Meses ▾")
            return
        dropdown_visivel = True
        botao_meses.configure(text="Meses ▴")
        _posicionar_dropdown()

    def _cancelar_animacao():
        job = animacao_estado.get("job")
        if job:
            try:
                janela.after_cancel(job)
            except Exception:
                pass
        animacao_estado["job"] = None

    def _desenhar_estado_vazio():
        ax.clear()
        ax.set_facecolor(UI_THEME["chart_plot_bg"])
        ax.set_title("Faturamento Mensal", fontsize=15, fontweight="bold", color=UI_THEME["text_primary"], pad=12)
        ax.set_xlabel("")
        ax.set_yticks([])
        ax.tick_params(axis="y", left=False, labelleft=False)
        ax.grid(False)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_visible(False)
        ax.spines["bottom"].set_visible(False)
        ax.set_xticks([])
        ax.text(
            0.5,
            0.5,
            "Selecione ao menos um mês para exibir o gráfico.",
            transform=ax.transAxes,
            ha="center",
            va="center",
            fontsize=11,
            color=UI_THEME["text_secondary"],
        )
        fig.tight_layout(pad=1.4)
        canvas.draw()

    def _renderizar_grafico(resumo, progresso=1.0, exibir_rotulos=False):
        ax.clear()
        ax.set_facecolor(UI_THEME["chart_plot_bg"])

        valores_reais = [float(v) for v in resumo["valor_inicial"].tolist()]
        indices = resumo["idx"].tolist()
        labels = resumo["mes_label"].tolist()
        qtd = len(valores_reais)
        if qtd == 0:
            _desenhar_estado_vazio()
            return

        fator = max(0.0, min(1.0, float(progresso)))
        valores_animados = [v * fator for v in valores_reais]
        cores = _gerar_cores_gradiente(qtd)

        valor_max = max(valores_reais) if valores_reais else 0
        limite_superior = max(valor_max * 1.18, 1.0)
        largura_barra = 0.46

        ax.set_title("Faturamento Mensal", fontsize=15, fontweight="bold", color=UI_THEME["text_primary"], pad=12)
        ax.set_xlabel("")
        ax.set_yticks([])
        ax.tick_params(axis="y", left=False, labelleft=False)
        ax.grid(False)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_visible(False)
        ax.spines["bottom"].set_visible(False)
        ax.tick_params(axis="x", colors=UI_THEME["text_secondary"], pad=7)
        ax.set_xticks(indices)
        ax.set_xticklabels(labels, rotation=0, ha="center", fontsize=10.5, fontweight="bold")
        ax.set_xlim(-0.5, qtd - 0.5)
        ax.set_ylim(0, limite_superior)
        ax.margins(x=0.08)

        fundo_topo = _misturar_cor(UI_THEME["chart_plot_bg"], "#FFFFFF", 0.06)
        fundo_base = _misturar_cor(UI_THEME["chart_plot_bg"], "#000000", 0.08)
        gradiente_bg = LinearSegmentedColormap.from_list("popup_bg_grad", [fundo_topo, fundo_base])
        ax.imshow(
            [[0], [1]],
            extent=[-0.6, qtd - 0.4, 0, limite_superior],
            aspect="auto",
            cmap=gradiente_bg,
            interpolation="bicubic",
            alpha=0.54,
            zorder=0,
        )

        alpha_barra = 0.35 + (0.60 * fator)
        for idx_barra, (pos_x, altura, cor) in enumerate(zip(indices, valores_animados, cores)):
            cor_base = cor
            if idx_barra == len(indices) - 1:
                cor_base = _misturar_cor(cor_base, UI_THEME["accent"], 0.34)

            esquerda = pos_x - (largura_barra / 2)
            altura_visivel = max(altura, 0.0001)
            barra = FancyBboxPatch(
                (esquerda, 0),
                largura_barra,
                altura_visivel,
                boxstyle=f"round,pad=0,rounding_size={largura_barra * 0.20}",
                linewidth=0,
                facecolor=cor_base,
                edgecolor=cor_base,
                alpha=max(0.52, alpha_barra),
                zorder=2,
            )
            ax.add_patch(barra)

            brilho_topo = Rectangle(
                (esquerda, max(altura_visivel * 0.42, 0)),
                largura_barra,
                max(altura_visivel * 0.58, 0.0001),
                linewidth=0,
                facecolor=_misturar_cor(cor_base, "#FFFFFF", 0.28),
                alpha=0.36,
                zorder=2.25,
            )
            brilho_topo.set_clip_path(barra)
            ax.add_patch(brilho_topo)

        tendencia = _calcular_tendencia_linear(valores_reais)
        if tendencia:
            ax.plot(
                indices,
                tendencia,
                color=UI_THEME["text_secondary"],
                linewidth=1.15,
                linestyle="--",
                alpha=0.45,
                zorder=3,
            )

        if exibir_rotulos:
            for pos_x, valor in zip(indices, valores_reais):
                if valor <= 0:
                    continue
                y_texto = valor * 0.52
                alinhamento_vertical = "center"
                cor_texto = UI_THEME["on_accent"]
                tamanho_fonte = 10
                if valor < (limite_superior * 0.10):
                    # Evita corte em barras pequenas.
                    y_texto = valor + (limite_superior * 0.025)
                    alinhamento_vertical = "bottom"
                    cor_texto = UI_THEME["text_primary"]
                    tamanho_fonte = 9
                ax.text(
                    pos_x,
                    y_texto,
                    _fmt_brl_exato(valor),
                    ha="center",
                    va=alinhamento_vertical,
                    fontsize=tamanho_fonte + 0.9,
                    fontweight="bold",
                    color=cor_texto,
                    zorder=4,
                )

        periodo_ini = _formatar_mes_ano(resumo["mes_dt"].min())
        periodo_fim = _formatar_mes_ano(resumo["mes_dt"].max())
        resumo_txt = f"Período exibido: {periodo_ini} a {periodo_fim}"
        ax.text(
            0.01,
            1.02,
            resumo_txt,
            transform=ax.transAxes,
            fontsize=9,
            color=UI_THEME["text_secondary"],
            ha="left",
            va="bottom",
        )

        fig.tight_layout(pad=1.4)
        canvas.draw()

    def _animar_grafico(resumo):
        _cancelar_animacao()
        total_frames = 6
        estado = {"frame": 0}

        def _passo():
            frame_atual = estado["frame"]
            t = (frame_atual + 1) / total_frames
            progresso = 1 - ((1 - t) ** 2)
            _renderizar_grafico(
                resumo,
                progresso=progresso,
                exibir_rotulos=(frame_atual >= total_frames - 1),
            )
            if frame_atual >= total_frames - 1:
                animacao_estado["job"] = None
                return
            estado["frame"] += 1
            animacao_estado["job"] = janela.after(16, _passo)

        _passo()

    def _desenhar_grafico(fechar_dropdown=True, animar=False):
        nonlocal dropdown_visivel
        resumo = _obter_resumo_filtrado()
        _atualizar_resumo_meses()

        if resumo.empty:
            _cancelar_animacao()
            _desenhar_estado_vazio()
            return

        if animar:
            _animar_grafico(resumo)
        else:
            _cancelar_animacao()
            _renderizar_grafico(
                resumo,
                progresso=1.0,
                exibir_rotulos=True,
            )

        if fechar_dropdown and dropdown_visivel:
            dropdown_meses_frame.place_forget()
            dropdown_visivel = False
            botao_meses.configure(text="Meses ▾")

    def _marcar_todos():
        for var in check_vars.values():
            var.set(True)
        _desenhar_grafico(fechar_dropdown=False, animar=True)

    def _limpar_todos():
        for var in check_vars.values():
            var.set(False)
        _desenhar_grafico(fechar_dropdown=False, animar=True)

    botao_meses.configure(command=_alternar_dropdown_meses)
    botao_marcar_todos.configure(command=_marcar_todos)
    botao_limpar.configure(command=_limpar_todos)
    _atualizar_resumo_meses()

    janela.bind("<Configure>", lambda _e: _posicionar_dropdown(), add="+")

    _desenhar_grafico(animar=False)

    def _fechar_janela():
        _cancelar_animacao()
        try:
            plt.close(fig)
        except Exception:
            pass
        janela.destroy()

    janela.protocol("WM_DELETE_WINDOW", _fechar_janela)


def atualizar_dashboard():
    global dashboard_update_running
    if dashboard_update_running:
        return
    dashboard_update_running = True

    def _texto_variacao(atual, anterior):
        """Retorna texto de variacao ex: '▲ 12,3%' ou '▼ 5,1%' ou '' se sem dados."""
        if anterior is None or anterior == 0:
            return ""
        pct = (atual - anterior) / abs(anterior) * 100
        seta = "▲" if pct >= 0 else "▼"
        return f"{seta} {abs(pct):.1f}% vs anterior"

    def atualizar_resumo(total_inicial=0.0, total_final=0.0, total_docs=0, total_nf=0, total_cte=0, total_cancelados=0,
                         anterior_inicial=None, anterior_final=None):
        total_inicial_valor_label.configure(text=formatar_moeda_brl(total_inicial))
        total_final_valor_label.configure(text=formatar_moeda_brl(total_final))
        total_documentos_label.configure(text=f"{total_docs}")
        nf_label.configure(text=f"{total_nf}")
        cte_label.configure(text=f"{total_cte}")
        cancelados_label.configure(text=f"{total_cancelados}")

        # Variacao periodo anterior
        txt_ini = _texto_variacao(total_inicial, anterior_inicial)
        cor_ini = UI_THEME["success_text"] if txt_ini.startswith("▲") else (UI_THEME["danger_text"] if txt_ini.startswith("▼") else UI_THEME["text_secondary"])
        variacao_inicial_label.configure(text=txt_ini, text_color=cor_ini)

        txt_fin = _texto_variacao(total_final, anterior_final)
        cor_fin = UI_THEME["success_text"] if txt_fin.startswith("▲") else (UI_THEME["danger_text"] if txt_fin.startswith("▼") else UI_THEME["text_secondary"])
        variacao_final_label.configure(text=txt_fin, text_color=cor_fin)

        diferenca = total_inicial - total_final
        if diferenca > 0:
            diferenca_label.configure(
                text=f"Diferenca referente a Impostos de NFS-e: {formatar_moeda_brl(diferenca)}",
                text_color=UI_THEME["success_text"],
            )
        elif diferenca < 0:
            diferenca_label.configure(
                text=f"Diferenca referente a Impostos de NFS-e: +{formatar_moeda_brl(abs(diferenca))}",
                text_color=UI_THEME["danger_text"],
            )
        else:
            diferenca_label.configure(
                text="Diferenca referente a Impostos de NFS-e: R$ 0,00",
                text_color=UI_THEME["text_secondary"],
            )

        if "cancelados_card" in globals():
            if total_cancelados > 0:
                cancelados_card.configure(fg_color=UI_THEME["danger_bg"])
                cancelados_label.configure(text_color=UI_THEME["danger_text"])
            else:
                cfg_cancel = summary_metric_widgets.get("cancelados", {})
                cancelados_card.configure(fg_color=UI_THEME.get(cfg_cancel.get("bg_key", "metric_default_bg"), UI_THEME["metric_default_bg"]))
                cancelados_label.configure(text_color=UI_THEME.get(cfg_cancel.get("value_key", "metric_default_value"), UI_THEME["metric_default_value"]))

    try:
        df, data_inicial, data_final = _obter_dataframe_dashboard_filtrado()
        if df is None:
            # Enquanto o usuario edita parcialmente a data, nao recalcula.
            return
        if data_inicial is not None and data_final is not None and data_inicial > data_final:
            atualizar_resumo()
            _atualizar_graficos_dashboard(pd.DataFrame(), data_inicial, data_final)
            return
        if df.empty:
            atualizar_resumo()
            _atualizar_graficos_dashboard(df, data_inicial, data_final)
            return

        total_inicial = df["valor_inicial"].fillna(0).sum()
        total_final = df["valor_final"].fillna(0).sum()
        total_docs = len(df)
        total_nf = int((df["tipo"] == "NF").sum())
        total_cte = int((df["tipo"] == "CTE").sum())
        total_cancelados = int(df["status"].str.upper().str.contains("CANCELADO", na=False).sum())

        # Calcula periodo anterior para variacao
        anterior_inicial = None
        anterior_final = None
        try:
            from datetime import timedelta
            duracao = (data_final - data_inicial).days + 1
            prev_fim = data_inicial - timedelta(days=1)
            prev_ini = prev_fim - timedelta(days=duracao - 1)
            df_prev, _, _ = obter_dataframe_dashboard(prev_ini, prev_fim)
            if df_prev is not None and not df_prev.empty:
                anterior_inicial = float(df_prev["valor_inicial"].fillna(0).sum())
                anterior_final = float(df_prev["valor_final"].fillna(0).sum())
        except Exception:
            pass

        atualizar_resumo(
            float(total_inicial),
            float(total_final),
            int(total_docs),
            total_nf,
            total_cte,
            total_cancelados,
            anterior_inicial=anterior_inicial,
            anterior_final=anterior_final,
        )
        _atualizar_graficos_dashboard(df, data_inicial, data_final)
    finally:
        dashboard_update_running = False


tab_buttons = {}
tab_hover_state = {}
tab_active_id = "principal"
tab_indicator = None
tab_indicator_host = None
menu_feedback_label = None
theme_toggle_button = None
ui_refs = {}
summary_metric_widgets = {}
dashboard_chart_widgets = {}
dashboard_chart_state = {
    "import_ok": None,
    "import_error": "",
    "plt": None,
    "FigureCanvasTkAgg": None,
    "FuncFormatter": None,
}
_anim_jobs = {}
screen_nav_buttons = {}
SCREEN_NAV_STYLES = {}
action_buttons = []
SCREEN_NAV_CATALOG = {
    "dashboard": "Dashboard",
    "hoje": "Hoje",
    "relatorios": "Relatórios",
    "alteracoes": "Alterações",
    "seguros": "Seguros",
    "tarefas": "Tarefas",
    "configuracoes": "Configurações",
    "medicao": "Medição",
}
SCREEN_NAV_DEFAULT_ORDER = list(SCREEN_NAV_CATALOG.keys())
ACTION_HOST_SCREENS = ["relatorios", "alteracoes", "configuracoes"]
SCREEN_ORDER_ATUAL = SCREEN_NAV_DEFAULT_ORDER.copy()
ACTION_LAYOUT_ATUAL = {}
seguros_mes_atual = datetime.now().month
seguros_ano_atual = datetime.now().year
seguros_filtro_status = "TODOS"
tarefas_filtro_atual = "TODAS"
tarefas_busca_atual = ""
tarefas_categoria_atual = "Todas"

# Registrar callbacks para src.documentos notificar mudanças na UI
_register_doc_on_change(_atualizar_cache_documentos_pos_alteracao)
_register_doc_on_change(atualizar_dashboard)
_register_doc_on_change(exportar_configuracoes_repo_silencioso)
_register_doc_on_change(lambda: criar_backup_automatico_silencioso("alteracao_documento"))


def _safe_config(widget, **kwargs):
    try:
        if widget is not None and widget.winfo_exists():
            widget.configure(**kwargs)
    except Exception:
        pass


def _texto_botao_tema():
    return "🌙 Tema escuro" if current_theme_mode == "light" else "☀ Tema claro"


def _definir_tema_interface(modo, persistir=True):
    global current_theme_mode, UI_THEME, TAB_STYLES

    novo_modo = "dark" if str(modo).strip().lower() == "dark" else "light"
    current_theme_mode = novo_modo
    UI_THEME = dict(APP_THEMES[novo_modo])
    TAB_STYLES = _gerar_tab_styles()
    ctk.set_appearance_mode("dark" if novo_modo == "dark" else "light")

    if persistir:
        try:
            salvar_configuracao("tema_interface", novo_modo)
        except Exception:
            pass


def aplicar_tema_interface():
    app.configure(fg_color=UI_THEME["app_bg"])

    _safe_config(container_scroll if "container_scroll" in globals() else None, fg_color=UI_THEME["app_bg"])
    _safe_config(ui_refs.get("sidebar"), fg_color=UI_THEME["header_bg"])
    _safe_config(ui_refs.get("content_shell"), fg_color=UI_THEME["app_bg"], border_color=UI_THEME["border"])
    scroll_canvas = ui_refs.get("scroll_canvas")
    if scroll_canvas is not None:
        try:
            scroll_canvas.configure(
                bg=UI_THEME["app_bg"],
                highlightthickness=0,
                bd=0,
            )
        except Exception:
            pass
    scroll_vertical = ui_refs.get("scrollbar_vertical")
    if scroll_vertical is not None:
        try:
            scroll_vertical.configure(
                bg=UI_THEME["scroll_btn"],
                activebackground=UI_THEME["scroll_btn_hover"],
                troughcolor=UI_THEME["app_bg"],
            )
        except Exception:
            pass
    _safe_config(ui_refs.get("main_frame"), fg_color=UI_THEME["app_bg"])
    _safe_config(ui_refs.get("screen_host"), fg_color=UI_THEME["app_bg"])
    if ui_refs.get("logo_watermark_label") is not None:
        try:
            ui_refs["logo_watermark_label"].lower()
        except Exception:
            pass
    for frame_tela in getattr(app, "screens", {}).values():
        _safe_config(frame_tela, fg_color=UI_THEME["app_bg"])

    _safe_config(ui_refs.get("header_card"), fg_color=UI_THEME["header_bg"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("header_title"), text_color=UI_THEME["text_primary"])
    _safe_config(
        theme_toggle_button,
        text=_texto_botao_tema(),
        fg_color=UI_THEME["surface_alt"],
        text_color=UI_THEME["text_primary"],
        hover_color=UI_THEME["tab_hover"],
        border_color=UI_THEME["border"],
    )
    _safe_config(
        ui_refs.get("btn_reordenar_interface"),
        fg_color=UI_THEME["surface_alt"],
        text_color=UI_THEME["text_primary"],
        hover_color=UI_THEME["tab_hover"],
        border_color=UI_THEME["border"],
    )

    _safe_config(ui_refs.get("filtro_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("filtro_titulo"), text_color=UI_THEME["text_primary"])
    _safe_config(ui_refs.get("filtro_subtitulo"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("filtro_ate"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("filtro_icon_inicio"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("filtro_icon_fim"), text_color=UI_THEME["text_secondary"])
    _safe_config(
        dashboard_data_inicio_entry if "dashboard_data_inicio_entry" in globals() else None,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        text_color=UI_THEME["text_primary"],
    )
    _safe_config(
        dashboard_data_fim_entry if "dashboard_data_fim_entry" in globals() else None,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        text_color=UI_THEME["text_primary"],
    )

    buscar_btn = ui_refs.get("buscar_btn")
    _safe_config(
        buscar_btn,
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        text_color=UI_THEME["on_accent"],
        border_color=UI_THEME["cta_border"],
    )
    if buscar_btn is not None:
        _aplicar_microinteracao_cta(
            buscar_btn,
            UI_THEME["accent"],
            UI_THEME["accent_hover"],
            UI_THEME["cta_press"],
            UI_THEME["cta_border"],
            UI_THEME["cta_border_hover"],
        )

    _safe_config(ui_refs.get("info_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("info_divider"), fg_color=UI_THEME["divider"])
    _safe_config(pasta_saida_label if "pasta_saida_label" in globals() else None, text_color=UI_THEME["text_secondary"])
    _safe_config(pasta_label if "pasta_label" in globals() else None, text_color=UI_THEME["text_primary"])
    _safe_config(
        progress if "progress" in globals() else None,
        fg_color=UI_THEME["progress_bg"],
        progress_color=UI_THEME["accent"],
    )
    _safe_config(status_label if "status_label" in globals() else None, text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_status_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("relatorios_status_label"), text_color=UI_THEME["success_text"])
    _safe_config(ui_refs.get("relatorios_arquivo_label"), text_color=UI_THEME["text_primary"])
    _safe_config(ui_refs.get("relatorios_saida_label"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_periodo_label"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_registros_label"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_filtro_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("relatorios_filtro_titulo"), text_color=UI_THEME["text_primary"])
    _safe_config(ui_refs.get("relatorios_filtro_subtitulo"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_filtro_ate"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_filtro_icon_inicio"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("relatorios_filtro_icon_fim"), text_color=UI_THEME["text_secondary"])
    _safe_config(
        relatorio_data_inicio_entry if "relatorio_data_inicio_entry" in globals() else None,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        text_color=UI_THEME["text_primary"],
    )
    _safe_config(
        relatorio_data_fim_entry if "relatorio_data_fim_entry" in globals() else None,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        text_color=UI_THEME["text_primary"],
    )
    relatorios_aplicar_filtro_btn = ui_refs.get("relatorios_aplicar_filtro_btn")
    _safe_config(
        relatorios_aplicar_filtro_btn,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        text_color=UI_THEME["text_primary"],
        border_color=UI_THEME["border"],
    )
    if relatorios_aplicar_filtro_btn is not None:
        _aplicar_microinteracao_botao(relatorios_aplicar_filtro_btn, "secondary")

    _safe_config(ui_refs.get("tabs_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    for tab_id, btn in tab_buttons.items():
        _safe_config(btn, hover_color=TAB_STYLES["hover_fg"], border_color=UI_THEME["border"])
        _aplicar_microinteracao_botao(btn, tab_id)
    _safe_config(tab_indicator, fg_color=TAB_STYLES["active_fg"])
    _safe_config(menu_feedback_label, text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("screen_nav_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    SCREEN_NAV_STYLES.update(
        {
            "normal_fg": UI_THEME["surface_alt"],
            "normal_text": UI_THEME["text_primary"],
            "hover_fg": UI_THEME["tab_hover"],
            "active_fg": UI_THEME["accent"],
            "active_hover_fg": UI_THEME["tab_active_hover"],
            "active_press_fg": UI_THEME["tab_active_press"],
            "active_text": UI_THEME["on_accent"],
        }
    )
    for btn in action_buttons:
        variant = getattr(btn, "_action_variant", "primary")
        if variant == "secondary":
            _safe_config(
                btn,
                fg_color=UI_THEME["surface_alt"],
                hover_color=UI_THEME["tab_hover"],
                text_color=UI_THEME["text_primary"],
                border_color=UI_THEME["border"],
            )
            _aplicar_microinteracao_cta(
                btn,
                UI_THEME["surface_alt"],
                UI_THEME["tab_hover"],
                UI_THEME["tab_press"],
                UI_THEME["border"],
                UI_THEME["border"],
                text_color=UI_THEME["text_primary"],
            )
        else:
            _safe_config(
                btn,
                fg_color=UI_THEME["accent"],
                hover_color=UI_THEME["accent_hover"],
                text_color=UI_THEME["on_accent"],
                border_color=UI_THEME["cta_border"],
            )
            _aplicar_microinteracao_cta(
                btn,
                UI_THEME["accent"],
                UI_THEME["accent_hover"],
                UI_THEME["cta_press"],
                UI_THEME["cta_border"],
                UI_THEME["cta_border_hover"],
                text_color=UI_THEME["on_accent"],
            )
    _configurar_nav_tela_ativa(app.current_screen or "dashboard")

    _safe_config(ui_refs.get("resumo_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("resumo_titulo"), text_color=UI_THEME["text_primary"])
    _safe_config(diferenca_label if "diferenca_label" in globals() else None, text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("graficos_card"), fg_color=UI_THEME["surface"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("graficos_titulo"), text_color=UI_THEME["text_primary"])
    _safe_config(ui_refs.get("graficos_subtitulo"), text_color=UI_THEME["text_secondary"])
    _safe_config(ui_refs.get("grafico_faturamento_card"), fg_color=UI_THEME["surface_alt"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("grafico_comparativo_card"), fg_color=UI_THEME["surface_alt"], border_color=UI_THEME["border"])
    _safe_config(ui_refs.get("grafico_faturamento_titulo"), text_color=UI_THEME["text_primary"])
    _safe_config(ui_refs.get("grafico_comparativo_titulo"), text_color=UI_THEME["text_primary"])
    for cfg in dashboard_chart_widgets.values():
        _safe_config(cfg.get("placeholder"), text_color=UI_THEME["text_secondary"])

    for metric_key, cfg in summary_metric_widgets.items():
        _safe_config(cfg["card"], fg_color=UI_THEME[cfg["bg_key"]], border_color=UI_THEME["border"])
        _safe_config(cfg["title"], text_color=UI_THEME[cfg["title_key"]])
        _safe_config(cfg["value"], text_color=UI_THEME[cfg["value_key"]])

    _configurar_tab_ativo(tab_active_id)
    atualizar_dashboard()
    _atualizar_status_relatorios_ui()


def alternar_tema_interface():
    novo_modo = "dark" if current_theme_mode == "light" else "light"
    _definir_tema_interface(novo_modo, persistir=True)
    aplicar_tema_interface()


def _catalogo_acoes_interface():
    return {
        "selecionar_relatorio": {"titulo": "📁 Selecionar relatório", "comando": selecionar_relatorio},
        "abrir_relatorio": {"titulo": "📄 Abrir relatório", "comando": abrir_relatorio},
        "exportar_relatorio_periodo": {"titulo": "📊 Exportar relatório do período", "comando": exportar_relatorio_filtrado},
        "abrir_pasta_relatorio": {"titulo": "📂 Abrir pasta do arquivo", "comando": abrir_pasta_arquivo_relatorio},
        "pasta_saida": {"titulo": "📂 Pasta de saída", "comando": selecionar_pasta_saida_relatorios},
        "relatorio_cancelados": {"titulo": "⛔ Relatórios cancelados", "comando": abrir_relatorio_cancelados},
        "grafico_faturamento": {"titulo": "📈 Gráfico de faturamento", "comando": abrir_grafico_faturamento},
        "buscar_documento": {"titulo": "🔎 Buscar documento", "comando": abrir_busca_documentos},
        "alterar_competencia": {"titulo": "📅 Alterar competência", "comando": abrir_dialogo_alterar_competencia},
        "substituir_documento": {"titulo": "🔁 Substituir documento", "comando": abrir_dialogo_substituir_documento},
        "cancelar_documento": {"titulo": "⛔ Cancelar documento", "comando": abrir_dialogo_cancelar_documento},
        "declarar_intercompany": {"titulo": "🏢 Declarar intercompany", "comando": abrir_dialogo_declarar_intercompany},
        "declarar_delta": {"titulo": "🔺 Declarar delta", "comando": abrir_dialogo_declarar_delta},
        "declarar_spot": {"titulo": "⚡ Declarar spot", "comando": abrir_dialogo_declarar_spot},
        "alternar_tema": {"titulo": "🌗 Alternar tema", "comando": alternar_tema_interface},
        "exportar_configuracoes": {"titulo": "📤 Exportar configurações", "comando": exportar_configuracoes_ui},
        "importar_configuracoes": {"titulo": "📥 Importar configurações", "comando": importar_configuracoes_ui},
    }


def _layout_botoes_padrao():
    return {
        "relatorios": [
            "selecionar_relatorio",
            "abrir_relatorio",
            "exportar_relatorio_periodo",
            "abrir_pasta_relatorio",
            "pasta_saida",
            "relatorio_cancelados",
            "grafico_faturamento",
            "buscar_documento",
        ],
        "alteracoes": [
            "alterar_competencia",
            "substituir_documento",
            "cancelar_documento",
            "declarar_intercompany",
            "declarar_delta",
            "declarar_spot",
        ],
        "configuracoes": [
            "alternar_tema",
            "exportar_configuracoes",
            "importar_configuracoes",
        ],
    }


def _normalizar_ordem_abas(ordem_raw):
    ordem = []
    for item in ordem_raw or []:
        chave = str(item).strip().lower()
        if chave in SCREEN_NAV_CATALOG and chave not in ordem:
            ordem.append(chave)
    for chave in SCREEN_NAV_DEFAULT_ORDER:
        if chave not in ordem:
            ordem.append(chave)
    return ordem


def _carregar_ordem_abas_interface():
    try:
        raw = obter_configuracao("ui_ordem_abas", "").strip()
        if not raw:
            return SCREEN_NAV_DEFAULT_ORDER.copy()
        dados = json.loads(raw)
        if not isinstance(dados, list):
            return SCREEN_NAV_DEFAULT_ORDER.copy()
        return _normalizar_ordem_abas(dados)
    except Exception:
        return SCREEN_NAV_DEFAULT_ORDER.copy()


def _normalizar_layout_botoes(layout_raw):
    catalogo = _catalogo_acoes_interface()
    layout_padrao = _layout_botoes_padrao()
    validos = set(catalogo.keys())
    usados = set()
    layout = {aba: [] for aba in ACTION_HOST_SCREENS}

    if isinstance(layout_raw, dict):
        for aba in ACTION_HOST_SCREENS:
            ids = layout_raw.get(aba, [])
            if not isinstance(ids, list):
                continue
            for acao_id in ids:
                chave = str(acao_id).strip().lower()
                if chave in validos and chave not in usados:
                    layout[aba].append(chave)
                    usados.add(chave)

    for aba in ACTION_HOST_SCREENS:
        for acao_id in layout_padrao.get(aba, []):
            if acao_id not in usados:
                layout[aba].append(acao_id)
                usados.add(acao_id)

    for acao_id in catalogo.keys():
        if acao_id not in usados:
            layout["configuracoes"].append(acao_id)
            usados.add(acao_id)

    # Mantém fluxo principal sempre disponível na aba Relatórios.
    obrigatorios_relatorios = [
        "selecionar_relatorio",
        "abrir_relatorio",
        "exportar_relatorio_periodo",
        "abrir_pasta_relatorio",
    ]
    for acao_id in obrigatorios_relatorios:
        for aba in ACTION_HOST_SCREENS:
            if acao_id in layout.get(aba, []):
                layout[aba] = [x for x in layout[aba] if x != acao_id]
        layout["relatorios"].insert(0, acao_id)

    # Remove possiveis duplicidades apos reforco dos obrigatorios.
    vistos_rel = set()
    rel_limpo = []
    for acao_id in layout["relatorios"]:
        if acao_id not in vistos_rel:
            rel_limpo.append(acao_id)
            vistos_rel.add(acao_id)
    layout["relatorios"] = rel_limpo

    # Mantém alternância de tema concentrada em Configurações.
    for aba in ACTION_HOST_SCREENS:
        if aba != "configuracoes" and "alternar_tema" in layout.get(aba, []):
            layout[aba] = [x for x in layout[aba] if x != "alternar_tema"]
    if "alternar_tema" not in layout["configuracoes"]:
        layout["configuracoes"].insert(0, "alternar_tema")

    return layout


def _carregar_layout_botoes_interface():
    try:
        raw = obter_configuracao("ui_layout_botoes", "").strip()
        if not raw:
            return _normalizar_layout_botoes(_layout_botoes_padrao())
        dados = json.loads(raw)
        return _normalizar_layout_botoes(dados)
    except Exception:
        return _normalizar_layout_botoes(_layout_botoes_padrao())


def _recarregar_layout_interface():
    global SCREEN_ORDER_ATUAL, ACTION_LAYOUT_ATUAL
    SCREEN_ORDER_ATUAL = _carregar_ordem_abas_interface()
    ACTION_LAYOUT_ATUAL = _carregar_layout_botoes_interface()


def _salvar_layout_interface(ordem_abas, layout_botoes):
    salvar_configuracao(
        "ui_ordem_abas",
        json.dumps(_normalizar_ordem_abas(ordem_abas), ensure_ascii=False),
    )
    salvar_configuracao(
        "ui_layout_botoes",
        json.dumps(_normalizar_layout_botoes(layout_botoes), ensure_ascii=False),
    )


def _renderizar_grade_acoes_por_tela(grade, tela_id):
    catalogo = _catalogo_acoes_interface()
    ordem = ACTION_LAYOUT_ATUAL.get(tela_id, [])
    linha = 0
    coluna = 0
    for acao_id in ordem:
        item = catalogo.get(acao_id)
        if not item:
            continue
        _criar_botao_acao(grade, item["titulo"], item["comando"]).grid(
            row=linha,
            column=coluna,
            padx=6,
            pady=6,
            sticky="ew",
        )
        coluna += 1
        if coluna > 1:
            coluna = 0
            linha += 1
    return bool(ordem)


def abrir_dialogo_reordenar_interface():
    ordem_abas = list(SCREEN_ORDER_ATUAL)
    layout_local = {aba: list(ids) for aba, ids in ACTION_LAYOUT_ATUAL.items()}
    catalogo = _catalogo_acoes_interface()
    labels_abas = {sid: SCREEN_NAV_CATALOG[sid] for sid in SCREEN_NAV_DEFAULT_ORDER}
    id_por_label_aba = {v: k for k, v in labels_abas.items()}
    labels_host = [labels_abas[sid] for sid in ACTION_HOST_SCREENS]

    dialog = ctk.CTkToplevel(app)
    dialog.title("Organizar interface")
    centralizar_janela(dialog, 860, 520)
    dialog.grab_set()

    container = ctk.CTkFrame(dialog, fg_color="transparent")
    container.pack(fill="both", expand=True, padx=14, pady=14)
    container.grid_columnconfigure((0, 1), weight=1)

    card_abas = _criar_card(container, corner_radius=14)
    card_abas.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
    card_abas.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        card_abas,
        text="Ordem das abas",
        font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=12, pady=(10, 8))

    lista_abas = tk.Listbox(
        card_abas,
        height=10,
        activestyle="none",
        bd=0,
        highlightthickness=0,
        bg=UI_THEME["surface_alt"],
        fg=UI_THEME["text_primary"],
        selectbackground=UI_THEME["accent"],
        selectforeground=UI_THEME["on_accent"],
        font=("Segoe UI", 11),
    )
    lista_abas.pack(fill="both", expand=True, padx=12, pady=(0, 10))

    acoes_abas = ctk.CTkFrame(card_abas, fg_color="transparent")
    acoes_abas.pack(fill="x", padx=12, pady=(0, 12))

    card_botoes = _criar_card(container, corner_radius=14)
    card_botoes.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
    card_botoes.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        card_botoes,
        text="Botoes por aba",
        font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=12, pady=(10, 8))

    combo_aba_origem = ctk.CTkComboBox(
        card_botoes,
        values=labels_host,
        width=260,
        state="readonly",
    )
    combo_aba_origem.set(labels_host[0])
    combo_aba_origem.pack(anchor="w", padx=12, pady=(0, 8))

    lista_botoes = tk.Listbox(
        card_botoes,
        height=10,
        activestyle="none",
        bd=0,
        highlightthickness=0,
        bg=UI_THEME["surface_alt"],
        fg=UI_THEME["text_primary"],
        selectbackground=UI_THEME["accent"],
        selectforeground=UI_THEME["on_accent"],
        font=("Segoe UI", 11),
    )
    lista_botoes.pack(fill="both", expand=True, padx=12, pady=(0, 10))
    ctk.CTkLabel(
        card_botoes,
        text="Dica: arraste um botao na lista para reordenar.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).pack(anchor="w", padx=12, pady=(0, 8))

    acoes_botoes = ctk.CTkFrame(card_botoes, fg_color="transparent")
    acoes_botoes.pack(fill="x", padx=12, pady=(0, 6))

    mover_frame = ctk.CTkFrame(card_botoes, fg_color="transparent")
    mover_frame.pack(fill="x", padx=12, pady=(0, 12))

    combo_aba_destino = ctk.CTkComboBox(
        mover_frame,
        values=labels_host,
        width=220,
        state="readonly",
    )
    combo_aba_destino.set(labels_host[1] if len(labels_host) > 1 else labels_host[0])
    combo_aba_destino.pack(side="left", padx=(0, 8))

    def _id_aba_origem():
        return id_por_label_aba.get(combo_aba_origem.get(), "relatorios")

    def _id_aba_destino():
        return id_por_label_aba.get(combo_aba_destino.get(), "alteracoes")

    def _atualizar_lista_abas(selecao_idx=None):
        lista_abas.delete(0, tk.END)
        for sid in ordem_abas:
            lista_abas.insert(tk.END, labels_abas[sid])
        if not ordem_abas:
            return
        idx = 0 if selecao_idx is None else max(0, min(selecao_idx, len(ordem_abas) - 1))
        lista_abas.selection_clear(0, tk.END)
        lista_abas.selection_set(idx)
        lista_abas.activate(idx)

    def _atualizar_lista_botoes(selecao_idx=None):
        lista_botoes.delete(0, tk.END)
        origem = _id_aba_origem()
        ids = layout_local.get(origem, [])
        for acao_id in ids:
            titulo = catalogo.get(acao_id, {}).get("titulo", acao_id)
            lista_botoes.insert(tk.END, titulo)
        if not ids:
            return
        idx = 0 if selecao_idx is None else max(0, min(selecao_idx, len(ids) - 1))
        lista_botoes.selection_clear(0, tk.END)
        lista_botoes.selection_set(idx)
        lista_botoes.activate(idx)

    drag_state = {"indice": None}

    def _iniciar_arraste_botao(event):
        origem = _id_aba_origem()
        ids = layout_local.get(origem, [])
        if not ids:
            drag_state["indice"] = None
            return
        idx = lista_botoes.nearest(event.y)
        idx = max(0, min(idx, len(ids) - 1))
        drag_state["indice"] = idx
        _atualizar_lista_botoes(idx)

    def _arrastar_botao(event):
        origem = _id_aba_origem()
        ids = layout_local.get(origem, [])
        idx_origem = drag_state.get("indice")
        if idx_origem is None or not ids:
            return
        idx_alvo = lista_botoes.nearest(event.y)
        idx_alvo = max(0, min(idx_alvo, len(ids) - 1))
        if idx_alvo == idx_origem:
            return
        item = ids.pop(idx_origem)
        ids.insert(idx_alvo, item)
        drag_state["indice"] = idx_alvo
        _atualizar_lista_botoes(idx_alvo)

    def _finalizar_arraste_botao(_event=None):
        drag_state["indice"] = None

    def _mover_aba(delta):
        selecao = lista_abas.curselection()
        if not selecao:
            return
        idx = int(selecao[0])
        novo = idx + delta
        if novo < 0 or novo >= len(ordem_abas):
            return
        ordem_abas[idx], ordem_abas[novo] = ordem_abas[novo], ordem_abas[idx]
        _atualizar_lista_abas(novo)

    def _mover_botao(delta):
        origem = _id_aba_origem()
        ids = layout_local.get(origem, [])
        selecao = lista_botoes.curselection()
        if not ids or not selecao:
            return
        idx = int(selecao[0])
        novo = idx + delta
        if novo < 0 or novo >= len(ids):
            return
        ids[idx], ids[novo] = ids[novo], ids[idx]
        _atualizar_lista_botoes(novo)

    def _mover_para_outra_aba():
        origem = _id_aba_origem()
        destino = _id_aba_destino()
        if origem == destino:
            return
        ids_origem = layout_local.get(origem, [])
        selecao = lista_botoes.curselection()
        if not ids_origem or not selecao:
            return
        idx = int(selecao[0])
        acao_id = ids_origem.pop(idx)
        layout_local.setdefault(destino, []).append(acao_id)
        _atualizar_lista_botoes(max(0, idx - 1))

    def _restaurar_padrao():
        ordem_abas[:] = SCREEN_NAV_DEFAULT_ORDER.copy()
        layout_padrao = _layout_botoes_padrao()
        layout_local.clear()
        for aba in ACTION_HOST_SCREENS:
            layout_local[aba] = list(layout_padrao.get(aba, []))
        combo_aba_origem.set(labels_abas["relatorios"])
        combo_aba_destino.set(labels_abas["alteracoes"])
        _atualizar_lista_abas(0)
        _atualizar_lista_botoes(0)

    def _salvar_e_aplicar():
        _salvar_layout_interface(ordem_abas, layout_local)
        _recarregar_layout_interface()
        tela_atual = app.current_screen or "dashboard"
        dialog.destroy()
        construir_tela_principal()
        aplicar_tema_interface()
        if tela_atual in app.screens:
            app.mostrar_tela(tela_atual)
        solicitar_atualizacao_dashboard(delay_ms=120)
        messagebox.showinfo("Layout atualizado", "Nova sequencia aplicada com sucesso.")

    ctk.CTkButton(acoes_abas, text="Subir aba", width=120, command=lambda: _mover_aba(-1)).pack(side="left", padx=(0, 8))
    ctk.CTkButton(acoes_abas, text="Descer aba", width=120, command=lambda: _mover_aba(1)).pack(side="left")

    ctk.CTkButton(acoes_botoes, text="Subir botao", width=120, command=lambda: _mover_botao(-1)).pack(side="left", padx=(0, 8))
    ctk.CTkButton(acoes_botoes, text="Descer botao", width=120, command=lambda: _mover_botao(1)).pack(side="left")
    ctk.CTkButton(mover_frame, text="Mover para aba", width=140, command=_mover_para_outra_aba).pack(side="left")

    combo_aba_origem.configure(command=lambda _v: _atualizar_lista_botoes(0))
    lista_botoes.bind("<ButtonPress-1>", _iniciar_arraste_botao, add="+")
    lista_botoes.bind("<B1-Motion>", _arrastar_botao, add="+")
    lista_botoes.bind("<ButtonRelease-1>", _finalizar_arraste_botao, add="+")

    rodape = ctk.CTkFrame(dialog, fg_color="transparent")
    rodape.pack(fill="x", padx=14, pady=(0, 14))
    ctk.CTkButton(rodape, text="Restaurar padrao", width=150, command=_restaurar_padrao).pack(side="left")
    ctk.CTkButton(rodape, text="Cancelar", width=120, command=dialog.destroy).pack(side="right", padx=(8, 0))
    ctk.CTkButton(rodape, text="Salvar", width=140, command=_salvar_e_aplicar).pack(side="right")

    _atualizar_lista_abas(0)
    _atualizar_lista_botoes(0)


def _criar_card(parent, fg_color=None, corner_radius=18, border_width=1):
    return ctk.CTkFrame(
        parent,
        fg_color=fg_color or UI_THEME["surface"],
        corner_radius=corner_radius,
        border_width=border_width,
        border_color=UI_THEME["border"],
    )


def _criar_badge_icone(parent, texto, cor=None, tamanho=46):
    return ctk.CTkLabel(
        parent,
        text=texto,
        width=tamanho,
        height=tamanho,
        corner_radius=14,
        fg_color=cor or UI_THEME["accent"],
        text_color=UI_THEME["on_accent"],
        font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold"),
    )


# ─── _normalizar_hex_cor/_hex_para_rgb/_rgb_para_hex/_interpolar_cor  →  src/utils.py

def _animar_estilo_botao(botao, alvo_fg, alvo_texto, alvo_borda, passos=7, delay_ms=16):
    if ui_animations_paused or scroll_dragging:
        _safe_config(
            botao,
            fg_color=_normalizar_hex_cor(alvo_fg),
            text_color=_normalizar_hex_cor(alvo_texto),
            border_color=_normalizar_hex_cor(alvo_borda),
        )
        return

    chave = f"btn_{id(botao)}"
    job = _anim_jobs.get(chave)
    if job:
        try:
            app.after_cancel(job)
        except Exception:
            pass

    try:
        inicio_fg = _normalizar_hex_cor(botao.cget("fg_color"))
        inicio_texto = _normalizar_hex_cor(botao.cget("text_color"))
        inicio_borda = _normalizar_hex_cor(botao.cget("border_color"))
    except Exception:
        inicio_fg = _normalizar_hex_cor(alvo_fg)
        inicio_texto = _normalizar_hex_cor(alvo_texto)
        inicio_borda = _normalizar_hex_cor(alvo_borda)

    alvo_fg = _normalizar_hex_cor(alvo_fg)
    alvo_texto = _normalizar_hex_cor(alvo_texto)
    alvo_borda = _normalizar_hex_cor(alvo_borda)

    def _passo(idx):
        if not botao.winfo_exists():
            return
        t = idx / float(passos)
        botao.configure(
            fg_color=_interpolar_cor(inicio_fg, alvo_fg, t),
            text_color=_interpolar_cor(inicio_texto, alvo_texto, t),
            border_color=_interpolar_cor(inicio_borda, alvo_borda, t),
        )
        if idx < passos:
            _anim_jobs[chave] = app.after(delay_ms, lambda: _passo(idx + 1))

    _passo(1)


def _animar_indicador_tab(tab_id):
    global tab_indicator
    if not tab_indicator_host or not tab_indicator:
        return
    botao = tab_buttons.get(tab_id)
    if not botao or not botao.winfo_exists():
        return

    tab_indicator_host.update_idletasks()
    x_alvo = botao.winfo_x() + 8
    w_alvo = max(24, botao.winfo_width() - 16)
    y_alvo = botao.winfo_y() + botao.winfo_height() + 6

    if not tab_indicator.winfo_ismapped():
        tab_indicator.configure(width=w_alvo, height=3)
        tab_indicator.place(x=x_alvo, y=y_alvo)
        return

    try:
        x_ini = int(tab_indicator.place_info().get("x", x_alvo))
        w_ini = int(tab_indicator.place_info().get("width", w_alvo))
    except Exception:
        x_ini, w_ini = x_alvo, w_alvo

    passos = 8
    delay_ms = 14

    def _passo(i):
        if not tab_indicator.winfo_exists():
            return
        t = i / float(passos)
        x = int(x_ini + (x_alvo - x_ini) * t)
        w = int(w_ini + (w_alvo - w_ini) * t)
        tab_indicator.configure(width=w, height=3)
        tab_indicator.place(x=x, y=y_alvo)
        if i < passos:
            app.after(delay_ms, lambda: _passo(i + 1))

    _passo(1)


def _aplicar_microinteracao_botao(botao, tab_id):
    tab_hover_state[tab_id] = False

    def _on_enter(_evt):
        tab_hover_state[tab_id] = True
        ativo = tab_active_id == tab_id
        _animar_estilo_botao(
            botao,
            TAB_STYLES["active_hover_fg"] if ativo else TAB_STYLES["hover_fg"],
            TAB_STYLES["active_text"] if ativo else TAB_STYLES["normal_text"],
            TAB_STYLES["active_fg"] if ativo else UI_THEME["border"],
        )

    def _on_leave(_evt):
        tab_hover_state[tab_id] = False
        ativo = tab_active_id == tab_id
        _animar_estilo_botao(
            botao,
            TAB_STYLES["active_fg"] if ativo else TAB_STYLES["normal_fg"],
            TAB_STYLES["active_text"] if ativo else TAB_STYLES["normal_text"],
            TAB_STYLES["active_fg"] if ativo else UI_THEME["border"],
        )

    def _on_press(_evt):
        ativo = tab_active_id == tab_id
        _animar_estilo_botao(
            botao,
            TAB_STYLES["active_press_fg"] if ativo else TAB_STYLES["press_fg"],
            TAB_STYLES["active_text"] if ativo else TAB_STYLES["normal_text"],
            TAB_STYLES["active_fg"] if ativo else UI_THEME["border"],
            passos=4,
            delay_ms=12,
        )

    def _on_release(_evt):
        ativo = tab_active_id == tab_id
        hovered = tab_hover_state.get(tab_id, False)
        _animar_estilo_botao(
            botao,
            TAB_STYLES["active_hover_fg"] if (ativo and hovered) else (
                TAB_STYLES["active_fg"] if ativo else (
                    TAB_STYLES["hover_fg"] if hovered else TAB_STYLES["normal_fg"]
                )
            ),
            TAB_STYLES["active_text"] if ativo else TAB_STYLES["normal_text"],
            TAB_STYLES["active_fg"] if ativo else UI_THEME["border"],
        )

    botao.bind("<Enter>", _on_enter)
    botao.bind("<Leave>", _on_leave)
    botao.bind("<ButtonPress-1>", _on_press)
    botao.bind("<ButtonRelease-1>", _on_release)


def _aplicar_microinteracao_cta(botao, base_fg, hover_fg, press_fg, base_border, hover_border, text_color=None):
    cor_texto = text_color or UI_THEME["on_accent"]

    def _on_enter(_evt):
        _animar_estilo_botao(
            botao,
            hover_fg,
            cor_texto,
            hover_border,
            passos=6,
            delay_ms=14,
        )

    def _on_leave(_evt):
        _animar_estilo_botao(
            botao,
            base_fg,
            cor_texto,
            base_border,
            passos=6,
            delay_ms=14,
        )

    def _on_press(_evt):
        _animar_estilo_botao(
            botao,
            press_fg,
            cor_texto,
            hover_border,
            passos=4,
            delay_ms=10,
        )

    def _on_release(_evt):
        _animar_estilo_botao(
            botao,
            hover_fg,
            cor_texto,
            hover_border,
            passos=5,
            delay_ms=12,
        )

    botao.bind("<Enter>", _on_enter)
    botao.bind("<Leave>", _on_leave)
    botao.bind("<ButtonPress-1>", _on_press)
    botao.bind("<ButtonRelease-1>", _on_release)


def _configurar_nav_tela_ativa(nome_tela):
    if not SCREEN_NAV_STYLES:
        return
    for chave, botao in screen_nav_buttons.items():
        ativo = chave == nome_tela
        alvo_fg = SCREEN_NAV_STYLES["active_fg"] if ativo else SCREEN_NAV_STYLES["normal_fg"]
        alvo_text = SCREEN_NAV_STYLES["active_text"] if ativo else SCREEN_NAV_STYLES["normal_text"]
        alvo_border = SCREEN_NAV_STYLES["active_fg"] if ativo else UI_THEME["border"]
        _safe_config(
            botao,
            hover_color=SCREEN_NAV_STYLES["active_hover_fg"] if ativo else SCREEN_NAV_STYLES["hover_fg"],
        )
        _animar_estilo_botao(
            botao,
            alvo_fg,
            alvo_text,
            alvo_border,
            passos=5,
            delay_ms=12,
        )


def _navegar_tela_com_feedback(nome_tela):
    botao = screen_nav_buttons.get(nome_tela)
    if botao is not None and botao.winfo_exists() and SCREEN_NAV_STYLES:
        _animar_estilo_botao(
            botao,
            SCREEN_NAV_STYLES["active_press_fg"],
            SCREEN_NAV_STYLES["active_text"],
            SCREEN_NAV_STYLES["active_fg"],
            passos=4,
            delay_ms=10,
        )
    app.after(55, lambda: app.mostrar_tela(nome_tela))


def _criar_botao_acao(parent, texto, comando, variant="primary", height=40):
    eh_secundario = str(variant).strip().lower() == "secondary"
    fg_color = UI_THEME["surface_alt"] if eh_secundario else UI_THEME["accent"]
    hover_color = UI_THEME["tab_hover"] if eh_secundario else UI_THEME["accent_hover"]
    press_color = UI_THEME["tab_press"] if eh_secundario else UI_THEME["cta_press"]
    border_color = UI_THEME["border"] if eh_secundario else UI_THEME["cta_border"]
    border_hover = UI_THEME["border"] if eh_secundario else UI_THEME["cta_border_hover"]
    text_color = UI_THEME["text_primary"] if eh_secundario else UI_THEME["on_accent"]

    botao = ctk.CTkButton(
        parent,
        text=texto,
        height=max(height, 42),
        corner_radius=14,
        border_width=1,
        border_color=border_color,
        fg_color=fg_color,
        hover_color=hover_color,
        text_color=text_color,
        font=ctk.CTkFont(family="Segoe UI Semibold", size=13, weight="bold"),
        command=comando,
    )
    botao._action_variant = "secondary" if eh_secundario else "primary"
    _aplicar_microinteracao_cta(
        botao,
        fg_color,
        hover_color,
        press_color,
        border_color,
        border_hover,
        text_color=text_color,
    )
    action_buttons.append(botao)
    return botao


def _configurar_tab_ativo(tab_id):
    global tab_active_id
    tab_active_id = tab_id
    for chave, botao in tab_buttons.items():
        ativo = chave == tab_id
        hovered = tab_hover_state.get(chave, False)
        alvo_fg = (
            TAB_STYLES["active_hover_fg"] if (ativo and hovered) else
            TAB_STYLES["active_fg"] if ativo else
            TAB_STYLES["hover_fg"] if hovered else
            TAB_STYLES["normal_fg"]
        )
        _animar_estilo_botao(
            botao,
            alvo_fg,
            TAB_STYLES["active_text"] if ativo else TAB_STYLES["normal_text"],
            TAB_STYLES["active_fg"] if ativo else UI_THEME["border"],
        )
    _animar_indicador_tab(tab_id)


def _executar_acao_menu(tab_id, titulo_item, comando):
    if menu_feedback_label and menu_feedback_label.winfo_exists():
        menu_feedback_label.configure(text=f"Ação selecionada: {titulo_item}")

    botao = tab_buttons.get(tab_id)
    if botao and botao.winfo_exists():
        _animar_estilo_botao(
            botao,
            TAB_STYLES["active_press_fg"],
            TAB_STYLES["active_text"],
            TAB_STYLES["active_fg"],
            passos=5,
            delay_ms=12,
        )
        app.after(70, lambda: _configurar_tab_ativo(tab_id))

    comando()


def abrir_menu_dropdown(botao, opcoes, tab_id, tab_titulo):
    _configurar_tab_ativo(tab_id)
    if menu_feedback_label and menu_feedback_label.winfo_exists():
        menu_feedback_label.configure(text=f"Menu: {tab_titulo}")

    menu = tk.Menu(
        app,
        tearoff=0,
        bg=UI_THEME["surface"],
        fg=UI_THEME["text_primary"],
        activebackground=UI_THEME["accent"],
        activeforeground=UI_THEME["on_accent"],
        selectcolor=UI_THEME["accent"],
        borderwidth=1,
        relief="solid",
    )
    menu.configure(font=("Segoe UI", 10))
    for item in opcoes:
        if item is None:
            menu.add_separator()
            continue
        titulo, comando = item
        menu.add_command(
            label=f"  {titulo}",
            command=lambda t=titulo, c=comando: _executar_acao_menu(tab_id, t, c),
        )
    try:
        menu.tk_popup(botao.winfo_rootx(), botao.winfo_rooty() + botao.winfo_height() + 4)
    finally:
        menu.grab_release()


def _obter_logo_watermark():
    if not os.path.exists(LOGO_PATH):
        return None

    cache_attr = "_logo_watermark"
    if hasattr(app, cache_attr):
        return getattr(app, cache_attr)

    try:
        with Image.open(LOGO_PATH) as origem:
            base = origem.convert("RGBA")

        def _processar_logo(opacidade):
            img = base.copy()
            novos_pixels = []
            for r, g, b, a in img.getdata():
                # Remove fundo claro para evitar "placa" branca.
                if r > 242 and g > 242 and b > 242:
                    novos_pixels.append((r, g, b, 0))
                    continue
                alpha = int(a * opacidade)
                novos_pixels.append((r, g, b, max(0, min(255, alpha))))
            img.putdata(novos_pixels)
            return img.filter(ImageFilter.GaussianBlur(radius=0.9))

        # Opacidade bem baixa (5% a 10%), com leve ajuste por tema.
        logo_light = _processar_logo(0.05)
        logo_dark = _processar_logo(0.08)

        logo_wm = ctk.CTkImage(
            light_image=logo_light,
            dark_image=logo_dark,
            size=(760, 320),
        )
        setattr(app, cache_attr, logo_wm)
        return logo_wm
    except Exception as e:
        print(f"Aviso: não foi possível preparar a watermark da logo - {e}")
        return None


def _aplicar_logo_watermark(container):
    if container is None or not container.winfo_exists():
        return

    logo_wm = _obter_logo_watermark()
    if logo_wm is None:
        return

    lbl_existente = ui_refs.get("logo_watermark_label")
    if lbl_existente is not None and lbl_existente.winfo_exists():
        try:
            lbl_existente.destroy()
        except Exception:
            pass

    lbl = ctk.CTkLabel(container, image=logo_wm, text="", fg_color="transparent")
    lbl.place(**WATERMARK_POS)
    lbl.lower()
    ui_refs["logo_watermark_label"] = lbl


def _criar_header(parent):
    global theme_toggle_button

    header = _criar_card(parent, fg_color=UI_THEME["header_bg"], corner_radius=22)
    header.pack(fill="x", padx=18, pady=(8, 8))
    ui_refs["header_card"] = header

    header_top = ctk.CTkFrame(header, fg_color="transparent")
    header_top.pack(fill="x", padx=14, pady=(8, 0))
    theme_toggle_button = ctk.CTkButton(
        header_top,
        text=_texto_botao_tema(),
        width=120,
        height=30,
        corner_radius=12,
        border_width=1,
        fg_color=UI_THEME["surface_alt"],
        border_color=UI_THEME["border"],
        hover_color=UI_THEME["tab_hover"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
        command=alternar_tema_interface,
    )
    theme_toggle_button.pack(side="right")

    titulo = ctk.CTkLabel(
        header,
        text="Sistema de Faturamento - Horizonte Logística",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    titulo.pack(pady=(10, 12))
    ui_refs["header_title"] = titulo


def _criar_filtro_periodo(parent):
    global dashboard_data_inicio_entry, dashboard_data_fim_entry

    card = _criar_card(parent, corner_radius=20)
    card.pack(fill="x", padx=18, pady=(0, 10))
    card.grid_columnconfigure(0, weight=0)
    card.grid_columnconfigure((1, 2, 3, 4), weight=1)
    card.grid_columnconfigure((5, 6), weight=0)
    ui_refs["filtro_card"] = card

    icon_box = _criar_badge_icone(card, "📅", cor="#0F7F7C", tamanho=54)
    icon_box.grid(row=0, column=0, rowspan=3, sticky="nw", padx=(20, 14), pady=(18, 0))

    titulo = ctk.CTkLabel(
        card,
        text="Período de emissão",
        font=ctk.CTkFont(family="Segoe UI", size=19, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    titulo.grid(row=0, column=1, columnspan=6, sticky="w", padx=0, pady=(18, 4))
    ui_refs["filtro_titulo"] = titulo

    subtitulo = ctk.CTkLabel(
        card,
        text="Filtre os documentos por intervalo de emissão",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    subtitulo.grid(row=1, column=1, columnspan=6, sticky="w", padx=0, pady=(0, 16))
    ui_refs["filtro_subtitulo"] = subtitulo

    icon_inicio = ctk.CTkLabel(card, text="📅", font=ctk.CTkFont(size=16), text_color=UI_THEME["text_secondary"])
    icon_inicio.grid(
        row=2, column=1, sticky="e", padx=(0, 6), pady=(0, 20)
    )
    ui_refs["filtro_icon_inicio"] = icon_inicio

    dashboard_data_inicio_entry = ctk.CTkEntry(
        card,
        width=150,
        height=44,
        corner_radius=14,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        border_width=1,
        font=ctk.CTkFont(family="Segoe UI", size=13),
    )
    dashboard_data_inicio_entry.grid(row=2, column=2, sticky="ew", padx=(0, 14), pady=(0, 20))

    lbl_ate = ctk.CTkLabel(card, text="até", font=ctk.CTkFont(family="Segoe UI", size=12), text_color=UI_THEME["text_secondary"])
    lbl_ate.grid(
        row=2, column=3, padx=4, pady=(0, 20)
    )
    ui_refs["filtro_ate"] = lbl_ate

    icon_fim = ctk.CTkLabel(card, text="📅", font=ctk.CTkFont(size=16), text_color=UI_THEME["text_secondary"])
    icon_fim.grid(
        row=2, column=4, sticky="e", padx=(8, 6), pady=(0, 20)
    )
    ui_refs["filtro_icon_fim"] = icon_fim

    dashboard_data_fim_entry = ctk.CTkEntry(
        card,
        width=150,
        height=44,
        corner_radius=14,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        border_width=1,
        font=ctk.CTkFont(family="Segoe UI", size=13),
    )
    dashboard_data_fim_entry.grid(row=2, column=5, sticky="ew", padx=(0, 14), pady=(0, 20))

    buscar_btn = ctk.CTkButton(
        card,
        text="Buscar",
        width=126,
        height=44,
        corner_radius=14,
        border_width=1,
        border_color=UI_THEME["cta_border"],
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        text_color=UI_THEME["on_accent"],
        font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
        command=aplicar_filtro_dashboard,
    )
    buscar_btn.grid(row=2, column=6, sticky="e", padx=(0, 20), pady=(0, 20))
    ui_refs["buscar_btn"] = buscar_btn
    _aplicar_microinteracao_cta(
        buscar_btn,
        UI_THEME["accent"],
        UI_THEME["accent_hover"],
        UI_THEME["cta_press"],
        UI_THEME["cta_border"],
        UI_THEME["cta_border_hover"],
    )

    inicializar_filtros_dashboard()

    dashboard_data_inicio_entry.bind("<Button-1>", lambda _e: abrir_seletor_data(dashboard_data_inicio_entry))
    dashboard_data_inicio_entry.bind("<FocusOut>", solicitar_atualizacao_dashboard)
    dashboard_data_inicio_entry.bind("<Return>", solicitar_atualizacao_dashboard)
    dashboard_data_fim_entry.bind("<Button-1>", lambda _e: abrir_seletor_data(dashboard_data_fim_entry))
    dashboard_data_fim_entry.bind("<FocusOut>", solicitar_atualizacao_dashboard)
    dashboard_data_fim_entry.bind("<Return>", solicitar_atualizacao_dashboard)


def _criar_info_relatorio(parent):
    global pasta_saida_label, pasta_label, progress, status_label

    info_card = _criar_card(parent, corner_radius=20)
    info_card.pack(fill="x", padx=18, pady=(0, 10))
    ui_refs["info_card"] = info_card
    ui_refs["relatorios_status_card"] = info_card

    titulo = ctk.CTkLabel(
        info_card,
        text="Informações operacionais",
        anchor="w",
        justify="left",
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
    )
    titulo.pack(fill="x", padx=18, pady=(12, 6))

    lbl_status = ctk.CTkLabel(
        info_card,
        text="Relatório carregado com sucesso",
        anchor="w",
        justify="left",
        text_color=UI_THEME["success_text"],
        font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
    )
    lbl_status.pack(fill="x", padx=18, pady=(0, 4))

    pasta_label = ctk.CTkLabel(
        info_card,
        text="Arquivo: Nenhum selecionado",
        anchor="w",
        justify="left",
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=12),
        wraplength=920,
    )
    pasta_label.pack(fill="x", padx=18, pady=(0, 4))

    pasta_saida_label = ctk.CTkLabel(
        info_card,
        text="Pasta de saída: -",
        anchor="w",
        justify="left",
        text_color=UI_THEME["text_secondary"],
        font=ctk.CTkFont(family="Segoe UI", size=11),
        wraplength=920,
    )
    pasta_saida_label.pack(fill="x", padx=18, pady=(0, 4))

    divisor = ctk.CTkFrame(info_card, height=1, fg_color=UI_THEME["divider"], corner_radius=1)
    divisor.pack(fill="x", padx=18, pady=(4, 8))
    ui_refs["info_divider"] = divisor

    lbl_periodo = ctk.CTkLabel(
        info_card,
        text="Período aplicado: Período indefinido",
        anchor="w",
        justify="left",
        text_color=UI_THEME["text_secondary"],
        font=ctk.CTkFont(family="Segoe UI", size=11),
        wraplength=920,
    )
    lbl_periodo.pack(fill="x", padx=18, pady=(0, 4))

    lbl_registros = ctk.CTkLabel(
        info_card,
        text="Registros no período: -",
        anchor="w",
        justify="left",
        text_color=UI_THEME["text_secondary"],
        font=ctk.CTkFont(family="Segoe UI", size=11),
        wraplength=920,
    )
    lbl_registros.pack(fill="x", padx=18, pady=(0, 8))

    progress = ctk.CTkProgressBar(
        info_card,
        height=10,
        corner_radius=6,
        fg_color=UI_THEME["progress_bg"],
        progress_color=UI_THEME["accent"],
    )
    progress.pack(fill="x", padx=18, pady=(0, 8))
    progress.set(0)

    status_label = ctk.CTkLabel(
        info_card,
        text="Processamento: Relatório 0 página(s) | Docs: 0",
        anchor="w",
        justify="left",
        text_color=UI_THEME["text_secondary"],
        font=ctk.CTkFont(family="Segoe UI", size=11),
    )
    status_label.pack(fill="x", padx=18, pady=(0, 14))

    ui_refs["relatorios_status_label"] = lbl_status
    ui_refs["relatorios_arquivo_label"] = pasta_label
    ui_refs["relatorios_saida_label"] = pasta_saida_label
    ui_refs["relatorios_periodo_label"] = lbl_periodo
    ui_refs["relatorios_registros_label"] = lbl_registros

    atualizar_label_pasta_saida()
    atualizar_label_relatorio()


def _criar_tabs_menu(parent):
    global tab_indicator, tab_indicator_host, menu_feedback_label

    tabs_card = _criar_card(parent, corner_radius=20)
    tabs_card.pack(fill="x", padx=18, pady=(0, 10))
    tabs_card.grid_columnconfigure((0, 1, 2), weight=1)
    tab_indicator_host = tabs_card
    ui_refs["tabs_card"] = tabs_card

    acoes_menu_principal = [
        ("Pasta do relatório", selecionar_pasta_saida_relatorios),
        ("Buscar documento", abrir_busca_documentos),
        ("Gráfico de faturamento", abrir_grafico_faturamento),
    ]
    acoes_menu_relatorios = [
        ("Selecionar relatório", selecionar_relatorio),
        ("Abrir relatório", abrir_relatorio),
        None,
        ("Relatórios cancelados", abrir_relatorio_cancelados),
        ("Exportar configurações", exportar_configuracoes_ui),
        ("Importar configurações", importar_configuracoes_ui),
    ]
    acoes_menu_alteracoes = [
        ("Alterar competência", abrir_dialogo_alterar_competencia),
        ("Substituir documento", abrir_dialogo_substituir_documento),
        ("Cancelar documento", abrir_dialogo_cancelar_documento),
        ("Declarar intercompany", abrir_dialogo_declarar_intercompany),
        ("Declarar delta", abrir_dialogo_declarar_delta),
        ("Declarar spot", abrir_dialogo_declarar_spot),
    ]

    tabs = [
        ("principal", "Principal", acoes_menu_principal),
        ("relatorios", "Relatórios", acoes_menu_relatorios),
        ("alteracoes", "Alterações", acoes_menu_alteracoes),
    ]

    for idx, (tab_id, titulo, opcoes) in enumerate(tabs):
        botao_tab = ctk.CTkButton(
            tabs_card,
            text=titulo,
            height=42,
            corner_radius=14,
            border_width=1,
            fg_color=TAB_STYLES["normal_fg"],
            border_color=UI_THEME["border"],
            hover_color=TAB_STYLES["hover_fg"],
            text_color=TAB_STYLES["normal_text"],
            font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
        )
        botao_tab.configure(command=lambda b=botao_tab, ops=opcoes, t=tab_id, ttl=titulo: abrir_menu_dropdown(b, ops, t, ttl))
        botao_tab.grid(row=0, column=idx, padx=8, pady=12, sticky="ew")
        tab_buttons[tab_id] = botao_tab
        _aplicar_microinteracao_botao(botao_tab, tab_id)

    tab_indicator = ctk.CTkFrame(
        tabs_card,
        fg_color=TAB_STYLES["active_fg"],
        height=3,
        corner_radius=3,
        border_width=0,
    )
    menu_feedback_label = ctk.CTkLabel(
        tabs_card,
        text="Escolha uma aba para acessar as ações",
        text_color=UI_THEME["text_secondary"],
        font=ctk.CTkFont(family="Segoe UI", size=11),
    )
    menu_feedback_label.grid(row=1, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 10))

    _configurar_tab_ativo("principal")


def _criar_bloco_metrica(parent, metric_key, titulo, valor_padrao):
    bg_key = f"metric_{metric_key}_bg" if f"metric_{metric_key}_bg" in UI_THEME else "metric_default_bg"
    title_key = f"metric_{metric_key}_title" if f"metric_{metric_key}_title" in UI_THEME else "metric_default_title"
    value_key = f"metric_{metric_key}_value" if f"metric_{metric_key}_value" in UI_THEME else "metric_default_value"

    card = ctk.CTkFrame(
        parent,
        fg_color=UI_THEME[bg_key],
        corner_radius=16,
        border_width=1,
        border_color=UI_THEME["border"],
    )
    card.configure(height=118 if metric_key in {"initial", "final"} else 106)
    card.pack_propagate(False)

    title_size = 12 if metric_key in {"initial", "final"} else 11
    value_size = 30 if metric_key in {"initial", "final"} else 26
    titulo_lbl = ctk.CTkLabel(
        card,
        text=titulo,
        font=ctk.CTkFont(family="Segoe UI", size=title_size),
        text_color=UI_THEME[title_key],
    )
    titulo_lbl.pack(pady=(12, 4))
    valor_label = ctk.CTkLabel(
        card,
        text=valor_padrao,
        font=ctk.CTkFont(family="Segoe UI", size=value_size, weight="bold"),
        text_color=UI_THEME[value_key],
    )
    valor_label.pack(pady=(0, 12))
    summary_metric_widgets[metric_key] = {
        "card": card,
        "title": titulo_lbl,
        "value": valor_label,
        "bg_key": bg_key,
        "title_key": title_key,
        "value_key": value_key,
    }
    return card, titulo_lbl, valor_label


def _criar_resumo_periodo(parent):
    global total_inicial_valor_label, total_final_valor_label, diferenca_label
    global total_documentos_label, nf_label, cte_label, cancelados_label, cancelados_card
    global variacao_inicial_label, variacao_final_label

    resumo_card = _criar_card(parent, corner_radius=20)
    resumo_card.pack(fill="x", padx=18, pady=(0, 16))
    ui_refs["resumo_card"] = resumo_card

    resumo_titulo = ctk.CTkLabel(
        resumo_card,
        text="Resumo do período",
        font=ctk.CTkFont(family="Segoe UI", size=21, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    resumo_titulo.pack(anchor="w", padx=20, pady=(16, 10))
    ui_refs["resumo_titulo"] = resumo_titulo

    linha_valores = ctk.CTkFrame(resumo_card, fg_color="transparent")
    linha_valores.pack(fill="x", padx=18, pady=(0, 8))
    linha_valores.grid_columnconfigure((0, 1), weight=1)

    inicial_card, _, total_inicial_valor_label = _criar_bloco_metrica(
        linha_valores, "initial", "Valor inicial", "R$ 0,00"
    )
    inicial_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
    variacao_inicial_label = ctk.CTkLabel(
        inicial_card,
        text="",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    variacao_inicial_label.pack(pady=(0, 8))

    final_card, _, total_final_valor_label = _criar_bloco_metrica(
        linha_valores, "final", "Valor final", "R$ 0,00"
    )
    final_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
    variacao_final_label = ctk.CTkLabel(
        final_card,
        text="",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    variacao_final_label.pack(pady=(0, 8))

    # Badge de diferença de impostos
    diferenca_badge_frame = ctk.CTkFrame(
        resumo_card,
        fg_color=UI_THEME["surface_alt"],
        corner_radius=10,
        border_width=1,
        border_color=UI_THEME["border"],
    )
    diferenca_badge_frame.pack(anchor="center", padx=20, pady=(6, 12))
    ui_refs["diferenca_badge_frame"] = diferenca_badge_frame

    diferenca_label = ctk.CTkLabel(
        diferenca_badge_frame,
        text="Diferenca referente a Impostos de NFS-e: R$ 0,00",
        font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
        text_color=UI_THEME["text_secondary"],
    )
    diferenca_label.pack(padx=16, pady=6)

    linha_metricas = ctk.CTkFrame(resumo_card, fg_color="transparent")
    linha_metricas.pack(fill="x", padx=18, pady=(0, 16))
    for c in range(4):
        linha_metricas.grid_columnconfigure(c, weight=1)

    docs_card, _, total_documentos_label = _criar_bloco_metrica(
        linha_metricas, "docs", "Total de documentos", "0"
    )
    docs_card.grid(row=0, column=0, padx=5, sticky="nsew")

    nf_card, _, nf_label = _criar_bloco_metrica(
        linha_metricas, "nf", "NF", "0"
    )
    nf_card.grid(row=0, column=1, padx=5, sticky="nsew")

    cte_card, _, cte_label = _criar_bloco_metrica(
        linha_metricas, "cte", "CTE", "0"
    )
    cte_card.grid(row=0, column=2, padx=5, sticky="nsew")

    cancelados_card, _, cancelados_label = _criar_bloco_metrica(
        linha_metricas, "cancelados", "Cancelados", "0"
    )
    cancelados_card.grid(row=0, column=3, padx=5, sticky="nsew")


def _criar_graficos_dashboard(parent):
    graficos_card = _criar_card(parent, corner_radius=20)
    graficos_card.pack(fill="x", padx=18, pady=(0, 16))
    ui_refs["graficos_card"] = graficos_card

    titulo = ctk.CTkLabel(
        graficos_card,
        text="Análise gráfica",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    titulo.pack(anchor="w", padx=20, pady=(16, 4))
    ui_refs["graficos_titulo"] = titulo

    subtitulo = ctk.CTkLabel(
        graficos_card,
        text="Visão de faturamento por período e comparativo de documentos.",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_secondary"],
    )
    subtitulo.pack(anchor="w", padx=20, pady=(0, 10))
    ui_refs["graficos_subtitulo"] = subtitulo

    grade = ctk.CTkFrame(graficos_card, fg_color="transparent")
    grade.pack(fill="x", padx=16, pady=(0, 16))
    grade.grid_columnconfigure((0, 1), weight=1, uniform="graf")

    card_faturamento = ctk.CTkFrame(
        grade,
        fg_color=UI_THEME["surface_alt"],
        corner_radius=16,
        border_width=1,
        border_color=UI_THEME["border"],
    )
    card_faturamento.grid(row=0, column=0, padx=(0, 8), sticky="nsew")
    ui_refs["grafico_faturamento_card"] = card_faturamento

    titulo_fat = ctk.CTkLabel(
        card_faturamento,
        text="Faturamento por periodo",
        font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    titulo_fat.pack(anchor="w", padx=12, pady=(10, 4))
    ui_refs["grafico_faturamento_titulo"] = titulo_fat

    host_fat = ctk.CTkFrame(card_faturamento, fg_color="transparent", height=230)
    host_fat.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    card_comp = ctk.CTkFrame(
        grade,
        fg_color=UI_THEME["surface_alt"],
        corner_radius=16,
        border_width=1,
        border_color=UI_THEME["border"],
    )
    card_comp.grid(row=0, column=1, padx=(8, 0), sticky="nsew")
    ui_refs["grafico_comparativo_card"] = card_comp

    titulo_comp = ctk.CTkLabel(
        card_comp,
        text="NF x CTE x Cancelados",
        font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    titulo_comp.pack(anchor="w", padx=12, pady=(10, 4))
    ui_refs["grafico_comparativo_titulo"] = titulo_comp

    host_comp = ctk.CTkFrame(card_comp, fg_color="transparent", height=230)
    host_comp.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    _liberar_graficos_dashboard()
    dashboard_chart_widgets["faturamento"] = {
        "host": host_fat,
        "fig": None,
        "canvas": None,
        "placeholder": None,
    }
    dashboard_chart_widgets["comparativo"] = {
        "host": host_comp,
        "fig": None,
        "canvas": None,
        "placeholder": None,
    }

    _mostrar_placeholder_grafico("faturamento", "Aguardando dados para montar o grafico.")
    _mostrar_placeholder_grafico("comparativo", "Aguardando dados para montar o grafico.")


def _criar_navegacao_telas(parent):

    SCREEN_NAV_STYLES.update(
        {
            "normal_fg": UI_THEME["surface_alt"],
            "normal_text": UI_THEME["text_primary"],
            "hover_fg": UI_THEME["tab_hover"],
            "active_fg": UI_THEME["accent"],
            "active_hover_fg": UI_THEME["tab_active_hover"],
            "active_press_fg": UI_THEME["tab_active_press"],
            "active_text": UI_THEME["on_accent"],
        }
    )

    nav_card = ctk.CTkFrame(parent, fg_color="transparent")
    nav_card.pack(fill="both", expand=True, padx=18, pady=18)
    ui_refs["screen_nav_card"] = nav_card

    logo_frame = ctk.CTkFrame(nav_card, fg_color="transparent")
    logo_frame.pack(fill="x", pady=(8, 30))
    if os.path.exists(LOGO_PATH):
        try:
            with Image.open(LOGO_PATH) as logo_img:
                logo_img.thumbnail((190, 78), Image.LANCZOS)
                logo_ctk = ctk.CTkImage(light_image=logo_img.copy(), dark_image=logo_img.copy(), size=logo_img.size)
                app._sidebar_logo = logo_ctk
                ctk.CTkLabel(logo_frame, text="", image=logo_ctk).pack(anchor="w")
        except Exception:
            ctk.CTkLabel(logo_frame, text="Horizonte\nLOGÍSTICA", justify="left",
                         font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
                         text_color=UI_THEME["text_primary"]).pack(anchor="w")
    else:
        ctk.CTkLabel(logo_frame, text="Horizonte\nLOGÍSTICA", justify="left",
                     font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
                     text_color=UI_THEME["text_primary"]).pack(anchor="w")

    telas = [(sid, SCREEN_NAV_CATALOG.get(sid, sid.title())) for sid in SCREEN_ORDER_ATUAL]
    nav_icons = {
        "dashboard": "🏠",
        "hoje": "☀",
        "relatorios": "📋",
        "alteracoes": "🛠",
        "seguros": "🛡",
        "tarefas": "☑",
        "configuracoes": "⚙",
        "medicao": "🔍",
    }

    for id_tela, titulo in telas:
        btn = ctk.CTkButton(
            nav_card,
            text=f"{nav_icons.get(id_tela, '•')}   {titulo}",
            height=48,
            corner_radius=12,
            border_width=1,
            fg_color=SCREEN_NAV_STYLES["normal_fg"],
            border_color=UI_THEME["border"],
            hover_color=SCREEN_NAV_STYLES["hover_fg"],
            text_color=SCREEN_NAV_STYLES["normal_text"],
            anchor="w",
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            command=lambda n=id_tela: _navegar_tela_com_feedback(n),
        )
        btn.pack(fill="x", pady=5)
        screen_nav_buttons[id_tela] = btn

    spacer = ctk.CTkFrame(nav_card, fg_color="transparent")
    spacer.pack(fill="both", expand=True)

    quick = _criar_card(nav_card, corner_radius=18)
    quick.pack(fill="x", pady=(18, 12))
    _criar_badge_icone(quick, "↗", cor="#0F7F7C", tamanho=42).pack(anchor="w", padx=18, pady=(18, 8))
    ctk.CTkLabel(quick, text="Resumo rápido",
                 font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
                 text_color=UI_THEME["text_primary"]).pack(anchor="w", padx=18, pady=(0, 6))
    ctk.CTkLabel(quick, text="Visão geral do faturamento\nem tempo real.", justify="left",
                 font=ctk.CTkFont(family="Segoe UI", size=11),
                 text_color=UI_THEME["text_secondary"]).pack(anchor="w", padx=18, pady=(0, 18))

    btn_reordenar = ctk.CTkButton(
        nav_card,
        text="🧩   Organizar botões",
        height=42,
        corner_radius=12,
        border_width=1,
        fg_color=UI_THEME["surface_alt"],
        border_color=UI_THEME["border"],
        hover_color=UI_THEME["tab_hover"],
        text_color=UI_THEME["text_primary"],
        anchor="w",
        font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
        command=abrir_dialogo_reordenar_interface,
    )
    btn_reordenar.pack(fill="x", pady=(0, 8))
    ui_refs["btn_reordenar_interface"] = btn_reordenar


def _criar_tela_dashboard(parent):
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])
    meses_pt = [
        "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]
    hoje = datetime.now()
    data_hoje = f"{hoje.day:02d} de {meses_pt[hoje.month]} de {hoje.year}"
    hero = ctk.CTkFrame(tela, fg_color="transparent")
    hero.pack(fill="x", padx=18, pady=(14, 12))
    hero.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(
        hero,
        text="Olá, bem-vindo! 👋",
        font=ctk.CTkFont(family="Segoe UI", size=26, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(
        hero,
        text="Aqui está o resumo do seu faturamento.",
        font=ctk.CTkFont(family="Segoe UI", size=13),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=1, column=0, sticky="w", pady=(4, 0))

    data_badge = _criar_card(hero, fg_color=UI_THEME["surface_alt"], corner_radius=14)
    data_badge.grid(row=0, column=1, rowspan=2, sticky="e", padx=(12, 0))
    ctk.CTkLabel(
        data_badge,
        text=f"📅   {data_hoje}",
        font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(padx=18, pady=12)
    _criar_filtro_periodo(tela)
    _criar_graficos_dashboard(tela)
    _criar_resumo_periodo(tela)
    return tela


def _obter_acoes_relatorios_ordenadas():
    catalogo = _catalogo_acoes_interface()
    permitidas = {
        "selecionar_relatorio",
        "abrir_relatorio",
        "exportar_relatorio_periodo",
        "pasta_saida",
        "relatorio_cancelados",
        "buscar_documento",
        "grafico_faturamento",
    }
    ids = [
        acao_id
        for acao_id in ACTION_LAYOUT_ATUAL.get("relatorios", [])
        if acao_id in catalogo and acao_id in permitidas
    ]
    if not ids:
        ids = [
            acao_id
            for acao_id in _layout_botoes_padrao().get("relatorios", [])
            if acao_id in catalogo and acao_id in permitidas
        ]
    for acao_id in ("selecionar_relatorio", "abrir_relatorio", "exportar_relatorio_periodo", "abrir_pasta_relatorio"):
        if acao_id in catalogo and acao_id not in ids:
            ids.append(acao_id)
    return ids


def _criar_filtro_relatorios(parent):
    global relatorio_data_inicio_entry, relatorio_data_fim_entry

    card = _criar_card(parent, corner_radius=20)
    card.pack(fill="x", padx=18, pady=(0, 12))
    card.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)
    ui_refs["relatorios_filtro_card"] = card

    titulo = ctk.CTkLabel(
        card,
        text="Período da aba Relatórios",
        font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
        text_color=UI_THEME["text_primary"],
    )
    titulo.grid(row=0, column=0, columnspan=6, sticky="w", padx=18, pady=(14, 4))
    ui_refs["relatorios_filtro_titulo"] = titulo

    subtitulo = ctk.CTkLabel(
        card,
        text="Esse período afeta a consulta, a abertura e a exportação do relatório.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    subtitulo.grid(row=1, column=0, columnspan=6, sticky="w", padx=18, pady=(0, 10))
    ui_refs["relatorios_filtro_subtitulo"] = subtitulo

    icon_inicio = ctk.CTkLabel(card, text="📅", font=ctk.CTkFont(size=15), text_color=UI_THEME["text_secondary"])
    icon_inicio.grid(row=2, column=0, sticky="e", padx=(18, 6), pady=(0, 14))
    ui_refs["relatorios_filtro_icon_inicio"] = icon_inicio

    relatorio_data_inicio_entry = ctk.CTkEntry(
        card,
        height=42,
        corner_radius=12,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        border_width=1,
        font=ctk.CTkFont(family="Segoe UI", size=13),
    )
    relatorio_data_inicio_entry.grid(row=2, column=1, sticky="ew", padx=(0, 8), pady=(0, 14))

    lbl_ate = ctk.CTkLabel(card, text="até", font=ctk.CTkFont(family="Segoe UI", size=12), text_color=UI_THEME["text_secondary"])
    lbl_ate.grid(row=2, column=2, padx=4, pady=(0, 14))
    ui_refs["relatorios_filtro_ate"] = lbl_ate

    icon_fim = ctk.CTkLabel(card, text="📅", font=ctk.CTkFont(size=15), text_color=UI_THEME["text_secondary"])
    icon_fim.grid(row=2, column=3, sticky="e", padx=(8, 6), pady=(0, 14))
    ui_refs["relatorios_filtro_icon_fim"] = icon_fim

    relatorio_data_fim_entry = ctk.CTkEntry(
        card,
        height=42,
        corner_radius=12,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        border_width=1,
        font=ctk.CTkFont(family="Segoe UI", size=13),
    )
    relatorio_data_fim_entry.grid(row=2, column=4, sticky="ew", padx=(0, 8), pady=(0, 14))

    btn_aplicar = ctk.CTkButton(
        card,
        text="Aplicar filtro",
        width=126,
        height=42,
        corner_radius=12,
        border_width=1,
        border_color=UI_THEME["border"],
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
        command=aplicar_filtro_relatorios,
    )
    btn_aplicar.grid(row=2, column=5, sticky="e", padx=(6, 18), pady=(0, 14))
    ui_refs["relatorios_aplicar_filtro_btn"] = btn_aplicar
    _aplicar_microinteracao_botao(btn_aplicar, "secondary")

    inicializar_filtros_relatorios()

    relatorio_data_inicio_entry.bind("<Button-1>", lambda _e: abrir_seletor_data(relatorio_data_inicio_entry))
    relatorio_data_fim_entry.bind("<Button-1>", lambda _e: abrir_seletor_data(relatorio_data_fim_entry))
    relatorio_data_inicio_entry.bind("<FocusOut>", lambda _e: _atualizar_status_relatorios_ui())
    relatorio_data_fim_entry.bind("<FocusOut>", lambda _e: _atualizar_status_relatorios_ui())
    relatorio_data_inicio_entry.bind("<Return>", lambda _e: aplicar_filtro_relatorios(mensagem_sucesso=False))
    relatorio_data_fim_entry.bind("<Return>", lambda _e: aplicar_filtro_relatorios(mensagem_sucesso=False))


def _criar_bloco_relatorios_principal(parent, acoes_relatorios):
    card = _criar_card(parent, corner_radius=20)
    card.pack(fill="x", padx=18, pady=(0, 12))

    ctk.CTkLabel(
        card,
        text="Relatórios",
        font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(14, 4))

    ctk.CTkLabel(
        card,
        text="Fluxo principal: selecionar, abrir e exportar o relatório filtrado pelo período desta aba.",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_secondary"],
    ).pack(anchor="w", padx=18, pady=(0, 12))

    grade = ctk.CTkFrame(card, fg_color="transparent")
    grade.pack(fill="x", padx=16, pady=(0, 16))
    grade.grid_columnconfigure((0, 1, 2, 3), weight=1)

    catalogo = _catalogo_acoes_interface()
    principal_ordem = ["selecionar_relatorio", "abrir_relatorio", "exportar_relatorio_periodo", "abrir_pasta_relatorio"]
    principal_ids = [aid for aid in principal_ordem if aid in acoes_relatorios]
    for idx, acao_id in enumerate(principal_ids):
        item = catalogo.get(acao_id)
        if not item:
            continue
        _criar_botao_acao(
            grade,
            item["titulo"],
            item["comando"],
            variant="primary",
            height=52,
        ).grid(row=0, column=idx, padx=6, pady=6, sticky="ew")


def _criar_bloco_relatorios_complementar(parent, acoes_relatorios):
    card = _criar_card(parent, corner_radius=18)
    card.pack(fill="x", padx=18, pady=(0, 12))

    ctk.CTkLabel(
        card,
        text="Ações complementares",
        font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(12, 4))

    ctk.CTkLabel(
        card,
        text="Ferramentas auxiliares para consulta e apoio.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).pack(anchor="w", padx=18, pady=(0, 10))

    grade = ctk.CTkFrame(card, fg_color="transparent")
    grade.pack(fill="x", padx=16, pady=(0, 16))
    grade.grid_columnconfigure((0, 1), weight=1)

    catalogo = _catalogo_acoes_interface()
    secundario_ids = [
        aid for aid in acoes_relatorios
        if aid in {"pasta_saida", "relatorio_cancelados", "buscar_documento", "grafico_faturamento"}
    ]

    linha = 0
    coluna = 0
    for acao_id in secundario_ids:
        item = catalogo.get(acao_id)
        if not item:
            continue
        _criar_botao_acao(
            grade,
            item["titulo"],
            item["comando"],
            variant="secondary",
            height=40,
        ).grid(row=linha, column=coluna, padx=6, pady=6, sticky="ew")
        coluna += 1
        if coluna > 1:
            coluna = 0
            linha += 1


def _criar_bloco_status_relatorios(parent):
    card = _criar_card(parent, corner_radius=16)
    card.pack(fill="x", padx=18, pady=(0, 16))
    ui_refs["relatorios_status_card"] = card

    lbl_status = ctk.CTkLabel(
        card,
        text="Relatório carregado com sucesso",
        anchor="w",
        justify="left",
        font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
        text_color=UI_THEME["success_text"],
    )
    lbl_status.pack(fill="x", padx=18, pady=(12, 4))

    lbl_arquivo = ctk.CTkLabel(
        card,
        text="Arquivo: nenhum selecionado",
        anchor="w",
        justify="left",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_primary"],
    )
    lbl_arquivo.pack(fill="x", padx=18, pady=(0, 4))

    lbl_saida = ctk.CTkLabel(
        card,
        text="Pasta de saída: -",
        anchor="w",
        justify="left",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    lbl_saida.pack(fill="x", padx=18, pady=(0, 4))

    lbl_periodo = ctk.CTkLabel(
        card,
        text="Período aplicado: período indefinido",
        anchor="w",
        justify="left",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    lbl_periodo.pack(fill="x", padx=18, pady=(0, 12))

    lbl_registros = ctk.CTkLabel(
        card,
        text="Registros no período: -",
        anchor="w",
        justify="left",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    )
    lbl_registros.pack(fill="x", padx=18, pady=(0, 12))

    ui_refs["relatorios_status_label"] = lbl_status
    ui_refs["relatorios_arquivo_label"] = lbl_arquivo
    ui_refs["relatorios_saida_label"] = lbl_saida
    ui_refs["relatorios_periodo_label"] = lbl_periodo
    ui_refs["relatorios_registros_label"] = lbl_registros


def _criar_tela_relatorios(parent):
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])
    acoes_relatorios = _obter_acoes_relatorios_ordenadas()
    _criar_filtro_relatorios(tela)
    _criar_bloco_relatorios_principal(tela, acoes_relatorios)
    _criar_info_relatorio(tela)
    _criar_bloco_relatorios_complementar(tela, acoes_relatorios)
    _atualizar_status_relatorios_ui()
    return tela


def _criar_tela_faturamento(parent):
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])
    card = _criar_card(tela, corner_radius=20)
    card.pack(fill="x", padx=18, pady=(0, 16))
    ctk.CTkLabel(
        card,
        text="Faturamento",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(14, 10))

    grade = ctk.CTkFrame(card, fg_color="transparent")
    grade.pack(fill="x", padx=16, pady=(0, 12))
    grade.grid_columnconfigure((0, 1), weight=1)
    tem_botoes = _renderizar_grade_acoes_por_tela(grade, "faturamento")

    if not tem_botoes:
        info = ctk.CTkLabel(
            card,
            text="Sem botões nesta aba. Use 'Trocar sequência dos botões' para organizar.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=UI_THEME["text_secondary"],
        )
        info.pack(anchor="w", padx=18, pady=(0, 16))
    return tela


def _consultar_painel_alteracoes(limite=2):
    dados = {
        "contadores": {
            "DELTA": {"qtd": 0, "total": 0.0},
            "SPOT": {"qtd": 0, "total": 0.0},
            "INTERCOMPANY": {"qtd": 0, "total": 0.0},
            "CANCELADOS": {"qtd": 0, "total": 0.0},
            "MANUAIS": {"qtd": 0, "total": 0.0},
        },
        "listas": {"DELTA": [], "SPOT": [], "INTERCOMPANY": [], "CANCELADOS": []},
        "historico": [],
    }
    conn = None
    try:
        conn = obter_conexao_banco()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT UPPER(COALESCE(frete, '')) AS modalidade,
                   COUNT(*) AS qtd,
                   COALESCE(SUM(COALESCE(valor_final, 0)), 0) AS total
            FROM documentos
            WHERE UPPER(COALESCE(frete, '')) IN ('DELTA', 'SPOT', 'INTERCOMPANY')
            GROUP BY UPPER(COALESCE(frete, ''))
            """
        )
        for row in cursor.fetchall():
            modalidade = str(row["modalidade"] or "").upper()
            if modalidade in dados["contadores"]:
                dados["contadores"][modalidade] = {
                    "qtd": int(row["qtd"] or 0),
                    "total": float(row["total"] or 0),
                }

        cursor.execute(
            """
            SELECT COUNT(*) AS qtd, COALESCE(SUM(COALESCE(valor_final_original, valor_final, 0)), 0) AS total
            FROM documentos
            WHERE COALESCE(cancelado_manual, 0)=1 OR UPPER(COALESCE(status, '')) LIKE '%CANCEL%'
            """
        )
        row = cursor.fetchone()
        if row:
            dados["contadores"]["CANCELADOS"] = {"qtd": int(row["qtd"] or 0), "total": float(row["total"] or 0)}

        cursor.execute(
            """
            SELECT COUNT(*) AS qtd
            FROM documentos
            WHERE COALESCE(competencia_manual, 0)=1
               OR COALESCE(frete_manual, 0)=1
               OR COALESCE(frete_revisado_manual, 0)=1
               OR COALESCE(cancelado_manual, 0)=1
            """
        )
        row = cursor.fetchone()
        if row:
            dados["contadores"]["MANUAIS"] = {"qtd": int(row["qtd"] or 0), "total": 0.0}

        def _buscar_lista(where_sql, params=()):
            cursor.execute(
                f"""
                SELECT tipo, numero, numero_original, competencia, data_emissao,
                       valor_final, valor_final_original, frete, status
                FROM documentos
                WHERE {where_sql}
                ORDER BY COALESCE(data_emissao, '') DESC, id DESC
                LIMIT ?
                """,
                (*params, int(limite)),
            )
            return [dict(row) for row in cursor.fetchall()]

        for modalidade in ("DELTA", "SPOT", "INTERCOMPANY"):
            dados["listas"][modalidade] = _buscar_lista("UPPER(COALESCE(frete, ''))=?", (modalidade,))
        dados["listas"]["CANCELADOS"] = _buscar_lista(
            "COALESCE(cancelado_manual, 0)=1 OR UPPER(COALESCE(status, '')) LIKE '%CANCEL%'"
        )
        try:
            cursor.execute(
                """
                SELECT data_hora, acao, tipo, numero, numero_original, campo, valor_anterior, valor_novo
                FROM historico_alteracoes
                ORDER BY id DESC
                LIMIT 6
                """
            )
            dados["historico"] = [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error:
            dados["historico"] = []
    except Exception as exc:
        _logger.warning("Falha ao consultar painel de alteracoes: %s", exc, exc_info=True)
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    return dados


def _formatar_doc_curto(row):
    tipo = str(row.get("tipo") or "-").upper()
    numero = row.get("numero_original") or row.get("numero") or "-"
    return f"{tipo} {numero}"


def _formatar_competencia_curta(valor):
    texto = str(valor or "").strip()
    return texto if texto else "-"


def _formatar_status_curto(valor):
    texto = str(valor or "").strip()
    return texto if texto else "-"


def _desfazer_alteracao_painel(row, tipo_acao):
    tipo = str(row.get("tipo") or "").upper().strip()
    numero = row.get("numero")
    try:
        numero = int(numero)
    except (TypeError, ValueError):
        messagebox.showwarning("Desfazer alteração", "Não foi possível identificar o número do documento.")
        return

    if tipo not in {"NF", "CTE"}:
        messagebox.showwarning("Desfazer alteração", "Tipo de documento inválido.")
        return

    if tipo_acao == "frete":
        confirmar = messagebox.askyesno(
            "Desfazer modalidade",
            f"Deseja voltar {tipo} {row.get('numero_original') or numero} para FRANQUIA?",
        )
        if not confirmar:
            return
        resultado = salvar_alteracao_frete_manual(tipo, numero, "FRANQUIA")
        if int(resultado.get("encontrados", 0) or 0) == 0:
            messagebox.showwarning("Desfazer modalidade", "Documento não encontrado.")
            return
        messagebox.showinfo("Desfazer modalidade", "Modalidade revertida para FRANQUIA.")
    elif tipo_acao == "cancelamento":
        confirmar = messagebox.askyesno(
            "Desfazer cancelamento",
            f"Deseja restaurar {tipo} {row.get('numero_original') or numero}?",
        )
        if not confirmar:
            return
        alterados = desfazer_cancelamento_documento(tipo, numero)
        if alterados == 0:
            messagebox.showwarning("Desfazer cancelamento", "Documento não estava cancelado manualmente.")
            return
        messagebox.showinfo("Desfazer cancelamento", "Cancelamento desfeito com sucesso.")

    _recriar_tela_alteracoes()


def _criar_tela_alteracoes(parent):
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])
    dados = _consultar_painel_alteracoes()

    header = ctk.CTkFrame(tela, fg_color="transparent")
    header.pack(fill="x", padx=18, pady=(0, 10))
    header.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        header,
        text="Alterações manuais",
        font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(
        header,
        text="Resumo das intervenções no faturamento, com foco em Delta, Spot e revisões manuais.",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=1, column=0, sticky="w", pady=(4, 0))
    _criar_botao_acao(
        header,
        "🔄 Atualizar painel",
        lambda: _recriar_tela_alteracoes(),
        variant="secondary",
        height=36,
    ).grid(row=0, column=1, rowspan=2, sticky="e", padx=(14, 0))

    resumo_grid = ctk.CTkFrame(tela, fg_color="transparent")
    resumo_grid.pack(fill="x", padx=18, pady=(0, 12))
    resumo_grid.grid_columnconfigure((0, 1, 2, 3), weight=1)

    def _card_resumo(coluna, titulo, chave, icone, cor, detalhe):
        info = dados["contadores"].get(chave, {"qtd": 0, "total": 0.0})
        card_resumo = _criar_card(resumo_grid, corner_radius=16)
        card_resumo.grid(row=0, column=coluna, sticky="nsew", padx=(0 if coluna == 0 else 6, 0 if coluna == 3 else 6))
        topo = ctk.CTkFrame(card_resumo, fg_color="transparent")
        topo.pack(fill="x", padx=14, pady=(14, 6))
        _criar_badge_icone(topo, icone, cor=cor, tamanho=38).pack(side="left")
        ctk.CTkLabel(
            topo,
            text=titulo,
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=UI_THEME["text_secondary"],
        ).pack(side="left", padx=(10, 0))
        ctk.CTkLabel(
            card_resumo,
            text=str(info["qtd"]),
            font=ctk.CTkFont(family="Segoe UI", size=28, weight="bold"),
            text_color=UI_THEME["text_primary"],
        ).pack(anchor="w", padx=14)
        ctk.CTkLabel(
            card_resumo,
            text=detalhe(info),
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w", padx=14, pady=(0, 14))

    _card_resumo(0, "Notas Delta", "DELTA", "🔺", "#1AAE8F", lambda i: f"Total: {formatar_moeda_brl(i['total'])}")
    _card_resumo(1, "Notas Spot", "SPOT", "⚡", "#2386D1", lambda i: f"Total: {formatar_moeda_brl(i['total'])}")
    _card_resumo(2, "Intercompany", "INTERCOMPANY", "🏢", "#8C5DE8", lambda i: f"Total: {formatar_moeda_brl(i['total'])}")
    _card_resumo(3, "Revisões manuais", "MANUAIS", "🛠", "#E0A422", lambda _i: "Competência, frete ou cancelamento")

    def _criar_tabela(parent_grid, linha_grid, coluna, titulo, subtitulo, linhas, cor, desfazer_tipo=None):
        card_tabela = _criar_card(parent_grid, corner_radius=16)
        card_tabela.grid(row=linha_grid, column=coluna, sticky="nsew", padx=(0 if coluna == 0 else 6, 0 if coluna == 1 else 6), pady=(0, 12))
        cab = ctk.CTkFrame(card_tabela, fg_color="transparent")
        cab.pack(fill="x", padx=14, pady=(14, 8))
        _criar_badge_icone(cab, "▦", cor=cor, tamanho=34).pack(side="left")
        textos = ctk.CTkFrame(cab, fg_color="transparent")
        textos.pack(side="left", padx=(10, 0), fill="x", expand=True)
        ctk.CTkLabel(
            textos,
            text=titulo,
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            text_color=UI_THEME["text_primary"],
        ).pack(anchor="w")
        ctk.CTkLabel(
            textos,
            text=subtitulo,
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w")

        grid = ctk.CTkFrame(card_tabela, fg_color="transparent")
        grid.pack(fill="x", padx=14, pady=(0, 14))
        pesos = (2, 1, 1, 1, 1) if desfazer_tipo else (2, 1, 1, 1)
        for idx, peso in enumerate(pesos):
            grid.grid_columnconfigure(idx, weight=peso)
        headers = ("Documento", "Competência", "Valor", "Status", "") if desfazer_tipo else ("Documento", "Competência", "Valor", "Status")
        for idx, texto in enumerate(headers):
            ctk.CTkLabel(
                grid,
                text=texto,
                font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
                text_color=UI_THEME["text_secondary"],
                anchor="w",
            ).grid(row=0, column=idx, sticky="ew", padx=(0 if idx == 0 else 8, 0), pady=(0, 6))

        if not linhas:
            ctk.CTkLabel(
                grid,
                text="Nenhuma nota encontrada nessa modalidade.",
                font=ctk.CTkFont(family="Segoe UI", size=12),
                text_color=UI_THEME["text_secondary"],
                anchor="w",
            ).grid(row=1, column=0, columnspan=4, sticky="ew", pady=(8, 2))
            return

        for row_idx, row in enumerate(linhas[:2], start=1):
            valores = (
                _formatar_doc_curto(row),
                _formatar_competencia_curta(row.get("competencia")),
                formatar_moeda_brl(row.get("valor_final") or 0),
                _formatar_status_curto(row.get("status")),
            )
            for col_idx, valor in enumerate(valores):
                ctk.CTkLabel(
                    grid,
                    text=valor,
                    font=ctk.CTkFont(family="Segoe UI", size=11),
                    text_color=UI_THEME["text_primary"] if col_idx == 0 else UI_THEME["text_secondary"],
                    anchor="w",
                ).grid(row=row_idx, column=col_idx, sticky="ew", padx=(0 if col_idx == 0 else 8, 0), pady=3)
            if desfazer_tipo:
                ctk.CTkButton(
                    grid,
                    text="Desfazer",
                    height=28,
                    corner_radius=9,
                    fg_color=UI_THEME["surface_alt"],
                    hover_color=UI_THEME["tab_hover"],
                    border_width=1,
                    border_color=UI_THEME["border"],
                    text_color=UI_THEME["text_primary"],
                    font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
                    command=lambda r=dict(row), t=desfazer_tipo: _desfazer_alteracao_painel(r, t),
                ).grid(row=row_idx, column=4, sticky="ew", padx=(8, 0), pady=2)

    tabelas_grid = ctk.CTkFrame(tela, fg_color="transparent")
    tabelas_grid.pack(fill="x", padx=18, pady=(0, 0))
    tabelas_grid.grid_columnconfigure((0, 1), weight=1)
    _criar_tabela(tabelas_grid, 0, 0, "Notas declaradas Delta", "Últimos documentos com frete revisado para Delta.", dados["listas"].get("DELTA", []), "#1AAE8F", "frete")
    _criar_tabela(tabelas_grid, 0, 1, "Notas declaradas Spot", "Últimos documentos com frete revisado para Spot.", dados["listas"].get("SPOT", []), "#2386D1", "frete")
    _criar_tabela(tabelas_grid, 1, 0, "Intercompany recentes", "Itens separados para acompanhamento interno.", dados["listas"].get("INTERCOMPANY", []), "#8C5DE8", "frete")
    _criar_tabela(tabelas_grid, 1, 1, "Cancelamentos recentes", "Documentos marcados como cancelados ou zerados manualmente.", dados["listas"].get("CANCELADOS", []), "#D85B6A", "cancelamento")

    historico_card = _criar_card(tela, corner_radius=18)
    historico_card.pack(fill="x", padx=18, pady=(0, 12))
    ctk.CTkLabel(
        historico_card,
        text="Histórico recente",
        font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(12, 4))
    ctk.CTkLabel(
        historico_card,
        text="Últimas alterações manuais registradas para conferência.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).pack(anchor="w", padx=18, pady=(0, 10))
    hist_grid = ctk.CTkFrame(historico_card, fg_color="transparent")
    hist_grid.pack(fill="x", padx=18, pady=(0, 14))
    for idx, peso in enumerate((1, 1, 1, 2, 2)):
        hist_grid.grid_columnconfigure(idx, weight=peso)
    for idx, texto in enumerate(("Data", "Ação", "Documento", "Antes", "Depois")):
        ctk.CTkLabel(
            hist_grid,
            text=texto,
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
            text_color=UI_THEME["text_secondary"],
            anchor="w",
        ).grid(row=0, column=idx, sticky="ew", padx=(0 if idx == 0 else 8, 0), pady=(0, 6))
    if not dados.get("historico"):
        ctk.CTkLabel(
            hist_grid,
            text="Nenhuma alteração registrada ainda. As próximas ações aparecerão aqui.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=UI_THEME["text_secondary"],
            anchor="w",
        ).grid(row=1, column=0, columnspan=5, sticky="ew", pady=(4, 0))
    else:
        for row_idx, item in enumerate(dados["historico"], start=1):
            data_txt = str(item.get("data_hora") or "")[5:16].replace("-", "/")
            doc_txt = f"{item.get('tipo') or '-'} {item.get('numero_original') or item.get('numero') or '-'}"
            valores = (
                data_txt,
                str(item.get("acao") or "").replace("_", " ").title(),
                doc_txt,
                str(item.get("valor_anterior") or "-"),
                str(item.get("valor_novo") or "-"),
            )
            for col_idx, valor in enumerate(valores):
                ctk.CTkLabel(
                    hist_grid,
                    text=valor,
                    font=ctk.CTkFont(family="Segoe UI", size=11),
                    text_color=UI_THEME["text_primary"] if col_idx == 2 else UI_THEME["text_secondary"],
                    anchor="w",
                ).grid(row=row_idx, column=col_idx, sticky="ew", padx=(0 if col_idx == 0 else 8, 0), pady=3)

    card = _criar_card(tela, corner_radius=20)
    card.pack(fill="x", padx=18, pady=(0, 16))
    ctk.CTkLabel(
        card,
        text="Ações rápidas",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(14, 2))
    ctk.CTkLabel(
        card,
        text="Corrija competência, substitua, cancele ou declare modalidades sem precisar sair da tela.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).pack(anchor="w", padx=18, pady=(0, 10))

    grade = ctk.CTkFrame(card, fg_color="transparent")
    grade.pack(fill="x", padx=16, pady=(0, 16))
    grade.grid_columnconfigure((0, 1), weight=1)
    tem_botoes = _renderizar_grade_acoes_por_tela(grade, "alteracoes")
    if not tem_botoes:
        ctk.CTkLabel(
            card,
            text="Sem botões nesta aba. Use 'Trocar sequência dos botões' para organizar.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w", padx=18, pady=(0, 16))
    return tela


def _recriar_tela_alteracoes():
    try:
        host = getattr(app, "screen_host", None)
        if host is None or not host.winfo_exists():
            return

        tela_antiga = app.screens.get("alteracoes")
        if tela_antiga is not None and tela_antiga.winfo_exists():
            tela_antiga.destroy()

        nova_tela = _criar_tela_alteracoes(host)
        app.registrar_tela("alteracoes", nova_tela)
        if getattr(app, "current_screen", "") == "alteracoes":
            app.mostrar_tela("alteracoes")
    except Exception as exc:
        _logger.warning("Falha ao atualizar painel de alteracoes: %s", exc, exc_info=True)


def _agendar_atualizacao_painel_alteracoes():
    try:
        app.after(0, _recriar_tela_alteracoes)
    except Exception:
        _recriar_tela_alteracoes()


def _status_seguro_label(status):
    status_norm = str(status or "PENDENTE").upper().strip()
    return {
        "PENDENTE": "Pendente",
        "RECEBIDO": "Recebido",
        "ENVIADO": "Enviado",
    }.get(status_norm, "Pendente")


def _status_seguro_cores(status):
    status_norm = str(status or "PENDENTE").upper().strip()
    if status_norm == "ENVIADO":
        return "#1FAE66", "#DDF8E9"
    if status_norm == "RECEBIDO":
        return "#2F80D0", "#E6F2FF"
    return "#E0A422", "#FFF4D9"


def _recriar_tela_seguros():
    try:
        host = getattr(app, "screen_host", None)
        if host is None or not host.winfo_exists():
            return
        tela_antiga = app.screens.get("seguros")
        if tela_antiga is not None and tela_antiga.winfo_exists():
            tela_antiga.destroy()
        nova_tela = _criar_tela_seguros(host)
        app.registrar_tela("seguros", nova_tela)
        if getattr(app, "current_screen", "") == "seguros":
            app.mostrar_tela("seguros")
    except Exception as exc:
        _logger.warning("Falha ao atualizar tela de seguros: %s", exc, exc_info=True)


def _recriar_tela_hoje():
    try:
        host = getattr(app, "screen_host", None)
        if host is None or not host.winfo_exists():
            return
        tela_antiga = app.screens.get("hoje")
        if tela_antiga is not None and tela_antiga.winfo_exists():
            tela_antiga.destroy()
        nova_tela = _criar_tela_hoje(host)
        app.registrar_tela("hoje", nova_tela)
        if getattr(app, "current_screen", "") == "hoje":
            app.mostrar_tela("hoje")
    except Exception as exc:
        _logger.warning("Falha ao atualizar tela Hoje: %s", exc, exc_info=True)


def _persistir_e_recriar_seguros():
    exportar_estado_operacional_silencioso()
    criar_backup_automatico_silencioso("seguros")
    _recriar_tela_hoje()
    _recriar_tela_seguros()


def _abrir_dialogo_adicionar_seguro():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Adicionar Seguro")
    centralizar_janela(dialog, 420, 190)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=18, pady=16)
    ctk.CTkLabel(
        form,
        text="Nome do seguro",
        font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
    ).pack(anchor="w", pady=(0, 6))
    nome_entry = ctk.CTkEntry(form, placeholder_text="Ex.: Fator, HDI", height=38)
    nome_entry.pack(fill="x", pady=(0, 14))
    nome_entry.focus_set()

    def _salvar():
        try:
            adicionar_seguro(nome_entry.get())
        except Exception as exc:
            messagebox.showwarning("Adicionar Seguro", str(exc))
            return
        dialog.destroy()
        _persistir_e_recriar_seguros()

    ctk.CTkButton(
        form,
        text="Salvar seguro",
        height=38,
        corner_radius=11,
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        command=_salvar,
    ).pack(fill="x")
    nome_entry.bind("<Return>", lambda _e: _salvar())


def _abrir_dialogo_observacao_seguro(item):
    dialog = ctk.CTkToplevel(app)
    dialog.title("Observação do Seguro")
    centralizar_janela(dialog, 540, 330)
    dialog.grab_set()

    form = ctk.CTkFrame(dialog, fg_color="transparent")
    form.pack(fill="both", expand=True, padx=18, pady=16)
    ctk.CTkLabel(
        form,
        text=f"Observação — {item.get('nome', 'Seguro')}",
        font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", pady=(0, 8))
    texto = ctk.CTkTextbox(form, height=170, corner_radius=12)
    texto.pack(fill="both", expand=True, pady=(0, 12))
    texto.insert("1.0", str(item.get("observacao") or ""))

    def _salvar():
        obs = texto.get("1.0", "end").strip()
        atualizar_observacao_seguro(item["seguro_id"], seguros_mes_atual, seguros_ano_atual, obs)
        dialog.destroy()
        _persistir_e_recriar_seguros()

    ctk.CTkButton(
        form,
        text="Salvar observação",
        height=38,
        corner_radius=11,
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        command=_salvar,
    ).pack(fill="x")


def _definir_status_seguro_ui(seguro_id, status):
    atualizar_status_seguro(seguro_id, seguros_mes_atual, seguros_ano_atual, status)
    _persistir_e_recriar_seguros()


def _inativar_seguro_ui(item):
    if not messagebox.askyesno(
        "Inativar Seguro",
        f"Deseja inativar o seguro {item.get('nome')}?\n\nEle não aparecerá em novas competências, mas o histórico registrado será mantido.",
    ):
        return
    inativar_seguro(item["seguro_id"])
    _persistir_e_recriar_seguros()


def _criar_tela_seguros(parent):
    global seguros_mes_atual, seguros_ano_atual, seguros_filtro_status

    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])
    registros = listar_controle_competencia(seguros_mes_atual, seguros_ano_atual, seguros_filtro_status)
    resumo = resumo_competencia(seguros_mes_atual, seguros_ano_atual)

    header = ctk.CTkFrame(tela, fg_color="transparent")
    header.pack(fill="x", padx=18, pady=(0, 12))
    header.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        header,
        text="Controle de Seguros",
        font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(
        header,
        text="Acompanhe comprovantes pendentes, recebidos e enviados por competência.",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=1, column=0, sticky="w", pady=(4, 0))
    ctk.CTkButton(
        header,
        text="➕ Adicionar Seguro",
        height=40,
        corner_radius=12,
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
        command=_abrir_dialogo_adicionar_seguro,
    ).grid(row=0, column=1, rowspan=2, sticky="e", padx=(12, 0))

    filtros_card = _criar_card(tela, corner_radius=18)
    filtros_card.pack(fill="x", padx=18, pady=(0, 12))
    filtros_card.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)
    ctk.CTkLabel(
        filtros_card,
        text="Competência",
        font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w", padx=16, pady=(14, 6))

    mes_combo = ctk.CTkComboBox(filtros_card, values=[m.capitalize() for m in MESES], height=36)
    mes_combo.set(MESES[seguros_mes_atual - 1].capitalize())
    mes_combo.grid(row=1, column=0, columnspan=2, sticky="ew", padx=(16, 8), pady=(0, 14))

    anos = [str(ano) for ano in range(datetime.now().year - 2, datetime.now().year + 5)]
    ano_combo = ctk.CTkComboBox(filtros_card, values=anos, height=36)
    ano_combo.set(str(seguros_ano_atual))
    ano_combo.grid(row=1, column=2, sticky="ew", padx=(0, 8), pady=(0, 14))

    def _aplicar_competencia(_event=None):
        global seguros_mes_atual, seguros_ano_atual
        mes_nome = mes_combo.get().strip().lower()
        if mes_nome in MESES:
            seguros_mes_atual = MESES.index(mes_nome) + 1
        try:
            seguros_ano_atual = int(ano_combo.get())
        except ValueError:
            seguros_ano_atual = datetime.now().year
        _recriar_tela_seguros()

    ctk.CTkButton(
        filtros_card,
        text="Aplicar",
        height=36,
        corner_radius=10,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        command=_aplicar_competencia,
    ).grid(row=1, column=3, sticky="ew", padx=(0, 8), pady=(0, 14))
    mes_combo.configure(command=_aplicar_competencia)
    ano_combo.configure(command=_aplicar_competencia)

    resumo_grid = ctk.CTkFrame(tela, fg_color="transparent")
    resumo_grid.pack(fill="x", padx=18, pady=(0, 12))
    resumo_grid.grid_columnconfigure((0, 1, 2, 3), weight=1)

    def _card_resumo(col, titulo, valor, cor, icone):
        card = _criar_card(resumo_grid, corner_radius=16)
        card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 6, 0 if col == 3 else 6))
        _criar_badge_icone(card, icone, cor=cor, tamanho=36).pack(anchor="w", padx=14, pady=(14, 6))
        ctk.CTkLabel(card, text=titulo, font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), text_color=UI_THEME["text_secondary"]).pack(anchor="w", padx=14)
        ctk.CTkLabel(card, text=str(valor), font=ctk.CTkFont(family="Segoe UI", size=28, weight="bold"), text_color=UI_THEME["text_primary"]).pack(anchor="w", padx=14, pady=(0, 12))

    _card_resumo(0, "Total cadastrados", resumo["total"], "#65758B", "▦")
    _card_resumo(1, "Pendentes", resumo["pendente"], "#E0A422", "!")
    _card_resumo(2, "Recebidos", resumo["recebido"], "#2F80D0", "↓")
    _card_resumo(3, "Enviados", resumo["enviado"], "#1FAE66", "✓")

    filtro_card = _criar_card(tela, corner_radius=18)
    filtro_card.pack(fill="x", padx=18, pady=(0, 12))
    filtro_row = ctk.CTkFrame(filtro_card, fg_color="transparent")
    filtro_row.pack(fill="x", padx=14, pady=12)
    opcoes = [("TODOS", "Todos"), ("PENDENTE", "Pendentes"), ("RECEBIDO", "Recebidos"), ("ENVIADO", "Enviados")]
    filtro_row.grid_columnconfigure(tuple(range(len(opcoes))), weight=1)

    def _set_filtro(status):
        global seguros_filtro_status
        seguros_filtro_status = status
        _recriar_tela_seguros()

    for idx, (status, label) in enumerate(opcoes):
        ativo = seguros_filtro_status == status
        ctk.CTkButton(
            filtro_row,
            text=label,
            height=34,
            corner_radius=10,
            fg_color=UI_THEME["accent"] if ativo else UI_THEME["surface_alt"],
            hover_color=UI_THEME["accent_hover"] if ativo else UI_THEME["tab_hover"],
            text_color=UI_THEME["on_accent"] if ativo else UI_THEME["text_primary"],
            border_width=1,
            border_color=UI_THEME["cta_border"] if ativo else UI_THEME["border"],
            command=lambda s=status: _set_filtro(s),
        ).grid(row=0, column=idx, sticky="ew", padx=5)

    lista_card = _criar_card(tela, corner_radius=20)
    lista_card.pack(fill="x", padx=18, pady=(0, 16))
    ctk.CTkLabel(
        lista_card,
        text=f"Comprovantes — {MESES[seguros_mes_atual - 1].capitalize()}/{seguros_ano_atual}",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(14, 8))

    if not registros:
        ctk.CTkLabel(
            lista_card,
            text="Nenhum seguro para este filtro. Cadastre um seguro ou altere o filtro.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w", padx=18, pady=(0, 16))
        return tela

    for item in registros:
        status = str(item.get("status") or "PENDENTE").upper()
        cor, _ = _status_seguro_cores(status)
        row = ctk.CTkFrame(lista_card, fg_color=UI_THEME["surface_alt"], corner_radius=14, border_width=1, border_color=UI_THEME["border"])
        row.pack(fill="x", padx=16, pady=6)
        row.grid_columnconfigure(0, weight=2)
        row.grid_columnconfigure(1, weight=1)
        row.grid_columnconfigure(2, weight=3)

        nome = item.get("nome") or "Seguro"
        nome_txt = f"{nome}" if int(item.get("ativo") or 0) == 1 else f"{nome} (inativo)"
        ctk.CTkLabel(row, text=nome_txt, font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"), text_color=UI_THEME["text_primary"], anchor="w").grid(row=0, column=0, sticky="w", padx=14, pady=(12, 2))
        ctk.CTkLabel(row, text=_status_seguro_label(status), font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"), fg_color=cor, corner_radius=12, text_color="#FFFFFF", width=86).grid(row=0, column=1, sticky="w", padx=8, pady=(12, 2))

        obs = str(item.get("observacao") or "").strip()
        obs_txt = obs if obs else "Sem observação."
        ctk.CTkLabel(row, text=obs_txt, font=ctk.CTkFont(family="Segoe UI", size=11), text_color=UI_THEME["text_secondary"], anchor="w").grid(row=1, column=0, columnspan=2, sticky="ew", padx=14, pady=(0, 12))

        acoes = ctk.CTkFrame(row, fg_color="transparent")
        acoes.grid(row=0, column=2, rowspan=2, sticky="e", padx=12, pady=10)
        for status_btn, label in (("PENDENTE", "Pendente"), ("RECEBIDO", "Recebido"), ("ENVIADO", "Enviado")):
            bcor, _ = _status_seguro_cores(status_btn)
            ctk.CTkButton(
                acoes,
                text=label,
                width=82,
                height=30,
                corner_radius=9,
                fg_color=bcor if status == status_btn else UI_THEME["surface"],
                hover_color=bcor,
                text_color="#FFFFFF" if status == status_btn else UI_THEME["text_primary"],
                border_width=1,
                border_color=bcor,
                font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
                command=lambda sid=item["seguro_id"], st=status_btn: _definir_status_seguro_ui(sid, st),
            ).pack(side="left", padx=3)
        ctk.CTkButton(
            acoes,
            text="Obs.",
            width=58,
            height=30,
            corner_radius=9,
            fg_color=UI_THEME["surface_alt"],
            hover_color=UI_THEME["tab_hover"],
            border_width=1,
            border_color=UI_THEME["border"],
            text_color=UI_THEME["text_primary"],
            command=lambda i=dict(item): _abrir_dialogo_observacao_seguro(i),
        ).pack(side="left", padx=3)
        if int(item.get("ativo") or 0) == 1:
            ctk.CTkButton(
                acoes,
                text="Inativar",
                width=70,
                height=30,
                corner_radius=9,
                fg_color=UI_THEME["surface_alt"],
                hover_color=UI_THEME["danger_bg"],
                border_width=1,
                border_color=UI_THEME["border"],
                text_color=UI_THEME["text_primary"],
                command=lambda i=dict(item): _inativar_seguro_ui(i),
            ).pack(side="left", padx=3)

    return tela


def _prioridade_tarefa_cor(prioridade):
    prioridade = str(prioridade or "MEDIA").upper()
    return {
        "BAIXA": "#64748B",
        "MEDIA": "#2F80D0",
        "ALTA": "#E0A422",
        "URGENTE": "#D85B6A",
    }.get(prioridade, "#2F80D0")


def _prazo_tarefa_cor(tarefa):
    classe = classificar_prazo(tarefa.get("prazo"), tarefa.get("status"))
    if classe == "atrasada":
        return "#D85B6A"
    if classe == "hoje":
        return "#E0A422"
    if classe == "concluido":
        return "#1FAE66"
    return UI_THEME["text_secondary"]


def _recriar_tela_tarefas():
    try:
        host = getattr(app, "screen_host", None)
        if host is None or not host.winfo_exists():
            return
        tela_antiga = app.screens.get("tarefas")
        if tela_antiga is not None and tela_antiga.winfo_exists():
            tela_antiga.destroy()
        nova_tela = _criar_tela_tarefas(host)
        app.registrar_tela("tarefas", nova_tela)
        if getattr(app, "current_screen", "") == "tarefas":
            app.mostrar_tela("tarefas")
    except Exception as exc:
        _logger.warning("Falha ao atualizar tela de tarefas: %s", exc, exc_info=True)


def _persistir_e_recriar_tarefas():
    exportar_estado_operacional_silencioso()
    criar_backup_automatico_silencioso("tarefas")
    _recriar_tela_hoje()
    _recriar_tela_tarefas()


def _abrir_dialogo_categoria_tarefa():
    dialog = ctk.CTkToplevel(app)
    dialog.title("Nova categoria")
    centralizar_janela(dialog, 380, 170)
    dialog.grab_set()
    frame = ctk.CTkFrame(dialog, fg_color="transparent")
    frame.pack(fill="both", expand=True, padx=16, pady=14)
    ctk.CTkLabel(frame, text="Nome da categoria", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold")).pack(anchor="w", pady=(0, 6))
    entry = ctk.CTkEntry(frame, height=36, placeholder_text="Ex.: Financeiro")
    entry.pack(fill="x", pady=(0, 12))

    def _salvar():
        try:
            adicionar_categoria(entry.get())
        except Exception as exc:
            messagebox.showwarning("Categoria", str(exc))
            return
        dialog.destroy()
        _persistir_e_recriar_tarefas()

    ctk.CTkButton(frame, text="Salvar categoria", height=36, command=_salvar).pack(fill="x")
    entry.focus_set()
    entry.bind("<Return>", lambda _e: _salvar())


def _abrir_dialogo_tarefa(tarefa=None):
    editando = bool(tarefa)
    dialog = ctk.CTkToplevel(app)
    dialog.title("Editar Tarefa" if editando else "Nova Tarefa")
    centralizar_janela(dialog, 720, 620)
    dialog.grab_set()

    frame = ctk.CTkFrame(dialog, fg_color="transparent")
    frame.pack(fill="both", expand=True, padx=18, pady=16)
    frame.grid_columnconfigure((0, 1), weight=1)

    def _label(txt, row, col):
        ctk.CTkLabel(frame, text=txt, font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold")).grid(row=row, column=col, sticky="w", padx=6, pady=(0, 4))

    _label("Título", 0, 0)
    titulo = ctk.CTkEntry(frame, height=36)
    titulo.grid(row=1, column=0, columnspan=2, sticky="ew", padx=6, pady=(0, 10))
    titulo.insert(0, tarefa.get("titulo", "") if tarefa else "")

    _label("Descrição/observação", 2, 0)
    descricao = ctk.CTkTextbox(frame, height=115, corner_radius=10)
    descricao.grid(row=3, column=0, columnspan=2, sticky="ew", padx=6, pady=(0, 10))
    descricao.insert("1.0", tarefa.get("descricao", "") if tarefa else "")

    categorias = [c["nome"] for c in listar_categorias()]
    _label("Projeto/categoria", 4, 0)
    categoria = ctk.CTkComboBox(frame, values=categorias or ["Outros"], height=36)
    categoria.set(tarefa.get("categoria", "Outros") if tarefa else "Outros")
    categoria.grid(row=5, column=0, sticky="ew", padx=6, pady=(0, 10))

    _label("Responsável", 4, 1)
    responsavel = ctk.CTkEntry(frame, height=36, placeholder_text="Ex.: Ismael, Financeiro")
    responsavel.grid(row=5, column=1, sticky="ew", padx=6, pady=(0, 10))
    responsavel.insert(0, tarefa.get("responsavel", "") if tarefa else "")

    _label("Prazo (dd/mm/aaaa)", 6, 0)
    prazo = ctk.CTkEntry(frame, height=36, placeholder_text="Sem prazo")
    prazo.grid(row=7, column=0, sticky="ew", padx=6, pady=(0, 10))
    prazo_val = formatar_prazo_br(tarefa.get("prazo", "")) if tarefa else ""
    prazo.insert(0, "" if prazo_val == "Sem prazo" else prazo_val)

    _label("Tags", 6, 1)
    tags = ctk.CTkEntry(frame, height=36, placeholder_text="Ex.: mensal, cliente")
    tags.grid(row=7, column=1, sticky="ew", padx=6, pady=(0, 10))
    tags.insert(0, tarefa.get("tags", "") if tarefa else "")

    prioridade_map = {v: k for k, v in TAREFA_PRIORIDADE_LABELS.items()}
    status_map = {v: k for k, v in TAREFA_STATUS_LABELS.items()}
    _label("Prioridade", 8, 0)
    prioridade = ctk.CTkComboBox(frame, values=list(prioridade_map.keys()), height=36)
    prioridade.set(TAREFA_PRIORIDADE_LABELS.get(tarefa.get("prioridade", "MEDIA"), "Média") if tarefa else "Média")
    prioridade.grid(row=9, column=0, sticky="ew", padx=6, pady=(0, 12))

    _label("Status", 8, 1)
    status = ctk.CTkComboBox(frame, values=list(status_map.keys()), height=36)
    status.set(TAREFA_STATUS_LABELS.get(tarefa.get("status", "A_FAZER"), "A Fazer") if tarefa else "A Fazer")
    status.grid(row=9, column=1, sticky="ew", padx=6, pady=(0, 12))

    actions = ctk.CTkFrame(frame, fg_color="transparent")
    actions.grid(row=10, column=0, columnspan=2, sticky="ew", padx=6, pady=(4, 0))
    actions.grid_columnconfigure((0, 1, 2), weight=1)

    def _salvar():
        try:
            dados = {
                "titulo": titulo.get(),
                "descricao": descricao.get("1.0", "end").strip(),
                "categoria": categoria.get(),
                "responsavel": responsavel.get(),
                "prazo": prazo.get(),
                "prioridade": prioridade_map.get(prioridade.get(), "MEDIA"),
                "status": status_map.get(status.get(), "A_FAZER"),
                "tags": tags.get(),
            }
            if editando:
                atualizar_tarefa(tarefa["id"], **dados)
            else:
                criar_tarefa(**dados)
        except Exception as exc:
            messagebox.showwarning("Tarefa", str(exc))
            return
        dialog.destroy()
        _persistir_e_recriar_tarefas()

    ctk.CTkButton(actions, text="Salvar", height=38, fg_color=UI_THEME["accent"], hover_color=UI_THEME["accent_hover"], command=_salvar).grid(row=0, column=0, sticky="ew", padx=(0, 6))
    ctk.CTkButton(actions, text="Nova categoria", height=38, fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], command=_abrir_dialogo_categoria_tarefa).grid(row=0, column=1, sticky="ew", padx=6)
    if editando:
        def _excluir():
            if messagebox.askyesno("Excluir tarefa", "Deseja excluir esta tarefa?"):
                excluir_tarefa(tarefa["id"])
                dialog.destroy()
                _persistir_e_recriar_tarefas()
        ctk.CTkButton(actions, text="Excluir", height=38, fg_color="#D85B6A", hover_color="#B94755", command=_excluir).grid(row=0, column=2, sticky="ew", padx=(6, 0))


def _mover_tarefa_ui(tarefa_id, status):
    mover_tarefa(tarefa_id, status)
    _persistir_e_recriar_tarefas()


def _criar_tela_hoje(parent):
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])

    hoje = datetime.now()
    tarefas_atrasadas = listar_tarefas("ATRASADAS")
    tarefas_hoje = listar_tarefas("HOJE")
    tarefas_urgentes = listar_tarefas("URGENTES")
    seguros_pendentes = listar_controle_competencia(hoje.month, hoje.year, "PENDENTE")
    alteracoes = _consultar_painel_alteracoes(limite=2)
    total_manuais = alteracoes.get("contadores", {}).get("MANUAIS", {}).get("qtd", 0)
    backups = listar_backups(limite=3)
    ultimo_bkp = backups[0] if backups else None

    header = ctk.CTkFrame(tela, fg_color="transparent")
    header.pack(fill="x", padx=18, pady=(0, 12))
    header.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        header,
        text="Hoje",
        font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(
        header,
        text="Um painel curto para ver atrasos, vencimentos, seguros e alteracoes recentes.",
        font=ctk.CTkFont(family="Segoe UI", size=12),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=1, column=0, sticky="w", pady=(4, 0))
    ctk.CTkLabel(
        header,
        text=hoje.strftime("%d/%m/%Y"),
        font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
        text_color=UI_THEME["text_primary"],
        fg_color=UI_THEME["surface_alt"],
        corner_radius=12,
        padx=14,
        pady=8,
    ).grid(row=0, column=1, rowspan=2, sticky="e", padx=(12, 0))

    resumo_grid = ctk.CTkFrame(tela, fg_color="transparent")
    resumo_grid.pack(fill="x", padx=18, pady=(0, 12))
    resumo_grid.grid_columnconfigure((0, 1, 2, 3), weight=1)

    def _resumo_card(col, titulo, valor, detalhe, cor, icone, destino):
        card = _criar_card(resumo_grid, corner_radius=16)
        card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 6, 0 if col == 3 else 6))
        topo = ctk.CTkFrame(card, fg_color="transparent")
        topo.pack(fill="x", padx=14, pady=(14, 6))
        _criar_badge_icone(topo, icone, cor=cor, tamanho=38).pack(side="left")
        ctk.CTkLabel(
            topo,
            text=str(valor),
            font=ctk.CTkFont(family="Segoe UI", size=28, weight="bold"),
            text_color=cor,
        ).pack(side="right")
        ctk.CTkLabel(
            card,
            text=titulo,
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=UI_THEME["text_primary"],
        ).pack(anchor="w", padx=14)
        ctk.CTkLabel(
            card,
            text=detalhe,
            font=ctk.CTkFont(family="Segoe UI", size=10),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w", padx=14, pady=(2, 12))
        card.bind("<Button-1>", lambda _e, d=destino: app.mostrar_tela(d))
        for child in card.winfo_children():
            child.bind("<Button-1>", lambda _e, d=destino: app.mostrar_tela(d))

    _resumo_card(0, "Tarefas atrasadas", len(tarefas_atrasadas), "Precisam de decisao ou replanejamento.", "#D85B6A", "!", "tarefas")
    _resumo_card(1, "Vencem hoje", len(tarefas_hoje), "Prazo para resolver ainda hoje.", "#E0A422", "H", "tarefas")
    _resumo_card(2, "Seguros pendentes", len(seguros_pendentes), f"{MESES[hoje.month - 1].capitalize()}/{hoje.year}", "#2F80D0", "S", "seguros")
    _resumo_card(3, "Alteracoes manuais", total_manuais, "Delta, Spot, intercompany e cancelados.", "#1AAE8F", "A", "alteracoes")

    central = _criar_card(tela, corner_radius=18)
    central.pack(fill="x", padx=18, pady=(0, 12))
    central.grid_columnconfigure(0, weight=1)
    cab_alertas = ctk.CTkFrame(central, fg_color="transparent")
    cab_alertas.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 8))
    cab_alertas.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        cab_alertas,
        text="Central de alertas",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkButton(
        cab_alertas,
        text="Backup agora",
        height=32,
        corner_radius=10,
        fg_color=UI_THEME["accent"],
        hover_color=UI_THEME["accent_hover"],
        font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
        command=criar_backup_manual_ui,
    ).grid(row=0, column=1, sticky="e", padx=(10, 0))
    ctk.CTkButton(
        cab_alertas,
        text="Buscar",
        height=32,
        width=86,
        corner_radius=10,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
        command=_abrir_dialogo_busca_global,
    ).grid(row=0, column=2, sticky="e", padx=(8, 0))
    ctk.CTkButton(
        cab_alertas,
        text="Consistencia",
        height=32,
        width=118,
        corner_radius=10,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
        command=_abrir_dialogo_auditoria_consistencia,
    ).grid(row=0, column=3, sticky="e", padx=(8, 0))

    def _abrir_pasta_backup():
        alvo = ultimo_bkp.get("pasta") if ultimo_bkp else getattr(_cfg, "BACKUP_DIR", "")
        if alvo:
            os.makedirs(alvo, exist_ok=True)
            os.startfile(alvo)

    if ultimo_bkp:
        data_bkp = str(ultimo_bkp.get("criado_em") or "").replace("T", " ")
        backup_txt = f"Ultimo backup: {data_bkp}"
    else:
        backup_txt = "Nenhum backup local registrado ainda."

    alertas = []
    if tarefas_atrasadas:
        alertas.append(("critico", f"{len(tarefas_atrasadas)} tarefa(s) atrasada(s)", "Abrir Tarefas", "tarefas"))
    if tarefas_hoje:
        alertas.append(("atencao", f"{len(tarefas_hoje)} tarefa(s) vencem hoje", "Abrir Tarefas", "tarefas"))
    if seguros_pendentes:
        alertas.append(("atencao", f"{len(seguros_pendentes)} seguro(s) pendente(s) nesta competencia", "Abrir Seguros", "seguros"))
    if not ultimo_bkp:
        alertas.append(("info", "Backup local ainda nao foi criado", "Backup agora", "backup"))
    if not alertas:
        alertas.append(("ok", "Sem alertas importantes agora. O painel esta limpo.", "Atualizar", "hoje"))

    cores_alerta = {"critico": "#D85B6A", "atencao": "#E0A422", "info": "#2F80D0", "ok": "#1FAE66"}
    for row_idx, (nivel, texto, acao, destino) in enumerate(alertas, start=1):
        linha = ctk.CTkFrame(central, fg_color=UI_THEME["surface_alt"], corner_radius=12, border_width=1, border_color=UI_THEME["border"])
        linha.grid(row=row_idx, column=0, sticky="ew", padx=14, pady=4)
        linha.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(
            linha,
            text="!",
            fg_color=cores_alerta.get(nivel, "#2F80D0"),
            corner_radius=10,
            text_color="#FFFFFF",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            width=28,
        ).grid(row=0, column=0, padx=(10, 8), pady=9)
        ctk.CTkLabel(
            linha,
            text=texto,
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
            text_color=UI_THEME["text_primary"],
            anchor="w",
        ).grid(row=0, column=1, sticky="ew", pady=9)
        comando = criar_backup_manual_ui if destino == "backup" else (lambda d=destino: app.mostrar_tela(d))
        ctk.CTkButton(
            linha,
            text=acao,
            width=104,
            height=28,
            corner_radius=9,
            fg_color=UI_THEME["surface"],
            hover_color=UI_THEME["tab_hover"],
            border_width=1,
            border_color=UI_THEME["border"],
            text_color=UI_THEME["text_primary"],
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
            command=comando,
        ).grid(row=0, column=2, padx=10, pady=8)

    rodape_backup = ctk.CTkFrame(central, fg_color="transparent")
    rodape_backup.grid(row=len(alertas) + 1, column=0, sticky="ew", padx=16, pady=(8, 14))
    rodape_backup.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(
        rodape_backup,
        text=backup_txt,
        font=ctk.CTkFont(family="Segoe UI", size=10),
        text_color=UI_THEME["text_secondary"],
    ).grid(row=0, column=0, sticky="w")
    ctk.CTkButton(
        rodape_backup,
        text="Abrir backup",
        height=28,
        width=96,
        corner_radius=9,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
        command=_abrir_pasta_backup,
    ).grid(row=0, column=1, sticky="e", padx=(8, 0))
    ctk.CTkButton(
        rodape_backup,
        text="Restaurar",
        height=28,
        width=92,
        corner_radius=9,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
        command=_abrir_dialogo_restaurar_backup,
    ).grid(row=0, column=2, sticky="e", padx=(8, 0))
    ctk.CTkButton(
        rodape_backup,
        text="Ver logs",
        height=28,
        width=82,
        corner_radius=9,
        fg_color=UI_THEME["surface_alt"],
        hover_color=UI_THEME["tab_hover"],
        border_width=1,
        border_color=UI_THEME["border"],
        text_color=UI_THEME["text_primary"],
        font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
        command=_abrir_dialogo_logs,
    ).grid(row=0, column=3, sticky="e", padx=(8, 0))

    corpo = ctk.CTkFrame(tela, fg_color="transparent")
    corpo.pack(fill="x", padx=18, pady=(0, 16))
    corpo.grid_columnconfigure((0, 1), weight=1)

    def _lista_card(row, col, titulo, subtitulo, itens, vazio, tipo, destino):
        card = _criar_card(corpo, corner_radius=18)
        card.grid(row=row, column=col, sticky="nsew", padx=(0 if col == 0 else 6, 0 if col == 1 else 6), pady=6)
        card.grid_columnconfigure(0, weight=1)
        head = ctk.CTkFrame(card, fg_color="transparent")
        head.pack(fill="x", padx=14, pady=(14, 6))
        ctk.CTkLabel(
            head,
            text=titulo,
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
            text_color=UI_THEME["text_primary"],
        ).pack(side="left")
        ctk.CTkButton(
            head,
            text="Abrir",
            width=64,
            height=28,
            corner_radius=9,
            fg_color=UI_THEME["surface_alt"],
            hover_color=UI_THEME["tab_hover"],
            border_width=1,
            border_color=UI_THEME["border"],
            text_color=UI_THEME["text_primary"],
            font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
            command=lambda d=destino: app.mostrar_tela(d),
        ).pack(side="right")
        ctk.CTkLabel(
            card,
            text=subtitulo,
            font=ctk.CTkFont(family="Segoe UI", size=10),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w", padx=14, pady=(0, 8))
        if not itens:
            ctk.CTkLabel(
                card,
                text=vazio,
                font=ctk.CTkFont(family="Segoe UI", size=12),
                text_color=UI_THEME["text_secondary"],
            ).pack(anchor="w", padx=14, pady=(8, 16))
            return
        for item in itens[:4]:
            linha = ctk.CTkFrame(card, fg_color=UI_THEME["surface_alt"], corner_radius=12, border_width=1, border_color=UI_THEME["border"])
            linha.pack(fill="x", padx=12, pady=4)
            if tipo == "seguro":
                nome = item.get("nome") or "Seguro"
                detalhe = str(item.get("observacao") or "").strip() or "Sem observacao."
                destaque = "Pendente"
                cor = "#E0A422"
            elif tipo == "alteracao":
                nome = _formatar_doc_curto(item)
                detalhe = f"{_formatar_competencia_curta(item.get('competencia'))} | {_formatar_status_curto(item.get('status'))}"
                destaque = str(item.get("frete") or "Manual")
                cor = "#1AAE8F"
            else:
                nome = item.get("titulo") or "Tarefa"
                detalhe = f"{item.get('categoria') or 'Sem categoria'} | {item.get('responsavel') or 'Sem responsavel'} | {formatar_prazo_br(item.get('prazo'))}"
                destaque = TAREFA_PRIORIDADE_LABELS.get(str(item.get("prioridade") or "MEDIA").upper(), "Media")
                cor = _prioridade_tarefa_cor(item.get("prioridade"))
            linha.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                linha,
                text=nome,
                font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
                text_color=UI_THEME["text_primary"],
                anchor="w",
            ).grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 0))
            ctk.CTkLabel(
                linha,
                text=detalhe,
                font=ctk.CTkFont(family="Segoe UI", size=10),
                text_color=UI_THEME["text_secondary"],
                anchor="w",
            ).grid(row=1, column=0, sticky="ew", padx=10, pady=(1, 8))
            ctk.CTkLabel(
                linha,
                text=destaque,
                fg_color=cor,
                corner_radius=10,
                text_color="#FFFFFF",
                font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
                width=74,
            ).grid(row=0, column=1, rowspan=2, sticky="e", padx=10, pady=8)

    _lista_card(0, 0, "Tarefas atrasadas", "Itens que passaram do prazo e ainda nao foram concluidos.", tarefas_atrasadas, "Nada atrasado por aqui.", "tarefa", "tarefas")
    _lista_card(0, 1, "Vencendo hoje", "O que merece prioridade no dia.", tarefas_hoje, "Nenhuma tarefa vence hoje.", "tarefa", "tarefas")
    _lista_card(1, 0, "Urgentes", "Prioridade alta do quadro Kanban.", tarefas_urgentes, "Nenhuma tarefa urgente cadastrada.", "tarefa", "tarefas")
    _lista_card(1, 1, "Seguros pendentes", "Comprovantes ainda pendentes na competencia atual.", seguros_pendentes, "Nenhum seguro pendente nesta competencia.", "seguro", "seguros")

    recentes = []
    for modalidade in ("DELTA", "SPOT", "INTERCOMPANY", "CANCELADOS"):
        recentes.extend(alteracoes.get("listas", {}).get(modalidade, [])[:1])
    if recentes:
        _lista_card(2, 0, "Alteracoes recentes", "Ultimos documentos marcados manualmente.", recentes[:4], "", "alteracao", "alteracoes")

    return tela


def _criar_tela_tarefas(parent):
    global tarefas_filtro_atual, tarefas_busca_atual, tarefas_categoria_atual
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])

    header = ctk.CTkFrame(tela, fg_color="transparent")
    header.pack(fill="x", padx=18, pady=(0, 12))
    header.grid_columnconfigure(0, weight=1)
    ctk.CTkLabel(header, text="Tarefas", font=ctk.CTkFont(family="Segoe UI", size=24, weight="bold"), text_color=UI_THEME["text_primary"]).grid(row=0, column=0, sticky="w")
    ctk.CTkLabel(header, text="Quadro Kanban para organizar afazeres, responsáveis, prioridades e prazos.", font=ctk.CTkFont(family="Segoe UI", size=12), text_color=UI_THEME["text_secondary"]).grid(row=1, column=0, sticky="w", pady=(4, 0))
    ctk.CTkButton(header, text="➕ Nova Tarefa", height=40, corner_radius=12, fg_color=UI_THEME["accent"], hover_color=UI_THEME["accent_hover"], font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"), command=lambda: _abrir_dialogo_tarefa()).grid(row=0, column=1, rowspan=2, sticky="e", padx=(12, 0))

    resumo = resumo_tarefas()
    resumo_grid = ctk.CTkFrame(tela, fg_color="transparent")
    resumo_grid.pack(fill="x", padx=18, pady=(0, 12))
    resumo_grid.grid_columnconfigure(tuple(range(7)), weight=1)
    cards = [
        ("Total", resumo["total"], "#65758B"),
        ("A fazer", resumo["a_fazer"], "#2F80D0"),
        ("Andamento", resumo["em_andamento"], "#1AAE8F"),
        ("Aguardando", resumo["aguardando"], "#E0A422"),
        ("Concluídas", resumo["concluido"], "#1FAE66"),
        ("Urgentes", resumo["urgentes"], "#D85B6A"),
        ("Atrasadas", resumo["atrasadas"], "#B94755"),
    ]
    for idx, (titulo, valor, cor) in enumerate(cards):
        card = _criar_card(resumo_grid, corner_radius=14)
        card.grid(row=0, column=idx, sticky="nsew", padx=4)
        ctk.CTkLabel(card, text=titulo, font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), text_color=UI_THEME["text_secondary"]).pack(anchor="w", padx=10, pady=(10, 2))
        ctk.CTkLabel(card, text=str(valor), font=ctk.CTkFont(family="Segoe UI", size=22, weight="bold"), text_color=cor).pack(anchor="w", padx=10, pady=(0, 10))

    filtro_card = _criar_card(tela, corner_radius=18)
    filtro_card.pack(fill="x", padx=18, pady=(0, 12))
    filtro_card.grid_columnconfigure((0, 1, 2), weight=1)
    busca = ctk.CTkEntry(filtro_card, height=36, placeholder_text="Buscar por título, descrição, responsável, categoria...")
    busca.insert(0, tarefas_busca_atual)
    busca.grid(row=0, column=0, sticky="ew", padx=(14, 8), pady=12)
    categorias = ["Todas"] + [c["nome"] for c in listar_categorias()]
    cat_combo = ctk.CTkComboBox(filtro_card, values=categorias, height=36)
    cat_combo.set(tarefas_categoria_atual if tarefas_categoria_atual in categorias else "Todas")
    cat_combo.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=12)

    def _aplicar_busca(_event=None):
        global tarefas_busca_atual, tarefas_categoria_atual
        tarefas_busca_atual = busca.get().strip()
        tarefas_categoria_atual = cat_combo.get().strip() or "Todas"
        _recriar_tela_tarefas()

    ctk.CTkButton(filtro_card, text="Aplicar filtros", height=36, fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], command=_aplicar_busca).grid(row=0, column=2, sticky="ew", padx=(0, 14), pady=12)
    busca.bind("<Return>", _aplicar_busca)
    cat_combo.configure(command=lambda _v: _aplicar_busca())

    filtros = [
        ("TODAS", "Todas"), ("A_FAZER", "A Fazer"), ("EM_ANDAMENTO", "Em andamento"),
        ("AGUARDANDO", "Aguardando"), ("CONCLUIDO", "Concluídas"), ("URGENTES", "Urgentes"),
        ("ATRASADAS", "Atrasadas"), ("HOJE", "Vencendo hoje"),
    ]
    filtros_row = ctk.CTkFrame(tela, fg_color="transparent")
    filtros_row.pack(fill="x", padx=18, pady=(0, 12))
    filtros_row.grid_columnconfigure(tuple(range(len(filtros))), weight=1)
    def _set_filtro(filtro):
        global tarefas_filtro_atual
        tarefas_filtro_atual = filtro
        _recriar_tela_tarefas()
    for idx, (filtro, label) in enumerate(filtros):
        ativo = tarefas_filtro_atual == filtro
        ctk.CTkButton(filtros_row, text=label, height=32, corner_radius=10, fg_color=UI_THEME["accent"] if ativo else UI_THEME["surface_alt"], hover_color=UI_THEME["accent_hover"] if ativo else UI_THEME["tab_hover"], text_color=UI_THEME["on_accent"] if ativo else UI_THEME["text_primary"], border_width=1, border_color=UI_THEME["cta_border"] if ativo else UI_THEME["border"], font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), command=lambda f=filtro: _set_filtro(f)).grid(row=0, column=idx, sticky="ew", padx=3)

    tarefas = listar_tarefas(tarefas_filtro_atual, tarefas_busca_atual, tarefas_categoria_atual)
    por_status = {status: [] for status in STATUS_TAREFA}
    for tarefa in tarefas:
        por_status.setdefault(tarefa.get("status"), []).append(tarefa)

    board = ctk.CTkFrame(tela, fg_color="transparent")
    board.pack(fill="x", padx=18, pady=(0, 16))
    board.grid_columnconfigure((0, 1, 2, 3), weight=1, uniform="kanban")

    for col_idx, status in enumerate(STATUS_TAREFA):
        coluna = _criar_card(board, corner_radius=18)
        coluna.grid(row=0, column=col_idx, sticky="nsew", padx=5)
        ctk.CTkLabel(coluna, text=f"{TAREFA_STATUS_LABELS[status]} ({len(por_status.get(status, []))})", font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"), text_color=UI_THEME["text_primary"]).pack(anchor="w", padx=12, pady=(12, 8))
        lista = ctk.CTkScrollableFrame(coluna, height=430, fg_color="transparent")
        lista.pack(fill="both", expand=True, padx=8, pady=(0, 10))
        if not por_status.get(status):
            ctk.CTkLabel(lista, text="Sem tarefas.", font=ctk.CTkFont(family="Segoe UI", size=11), text_color=UI_THEME["text_secondary"]).pack(anchor="w", padx=8, pady=8)
        for tarefa in por_status.get(status, []):
            _criar_card_tarefa(lista, tarefa)
    return tela


def _criar_card_tarefa(parent, tarefa):
    prioridade = str(tarefa.get("prioridade") or "MEDIA").upper()
    cor_prio = _prioridade_tarefa_cor(prioridade)
    card = ctk.CTkFrame(parent, fg_color=UI_THEME["surface_alt"], corner_radius=14, border_width=1, border_color=cor_prio)
    card.pack(fill="x", padx=4, pady=6)
    card.bind("<Button-1>", lambda _e, t=dict(tarefa): _abrir_dialogo_tarefa(t))
    titulo = str(tarefa.get("titulo") or "Sem título")
    ctk.CTkLabel(card, text=titulo, wraplength=210, justify="left", font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"), text_color=UI_THEME["text_primary"], anchor="w").pack(fill="x", padx=10, pady=(10, 4))
    meta = f"{tarefa.get('categoria') or 'Sem categoria'} • {tarefa.get('responsavel') or 'Sem responsável'}"
    ctk.CTkLabel(card, text=meta, wraplength=210, justify="left", font=ctk.CTkFont(family="Segoe UI", size=10), text_color=UI_THEME["text_secondary"], anchor="w").pack(fill="x", padx=10)
    row = ctk.CTkFrame(card, fg_color="transparent")
    row.pack(fill="x", padx=10, pady=(8, 6))
    ctk.CTkLabel(row, text=TAREFA_PRIORIDADE_LABELS.get(prioridade, "Média"), fg_color=cor_prio, corner_radius=10, text_color="#FFFFFF", font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"), width=64).pack(side="left")
    ctk.CTkLabel(row, text=formatar_prazo_br(tarefa.get("prazo")), text_color=_prazo_tarefa_cor(tarefa), font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold")).pack(side="right")
    mover = ctk.CTkFrame(card, fg_color="transparent")
    mover.pack(fill="x", padx=8, pady=(0, 8))
    outros = [s for s in STATUS_TAREFA if s != tarefa.get("status")]
    for status in outros[:3]:
        ctk.CTkButton(mover, text=TAREFA_STATUS_LABELS[status].split()[0], height=24, width=58, corner_radius=8, fg_color=UI_THEME["surface"], hover_color=UI_THEME["tab_hover"], border_width=1, border_color=UI_THEME["border"], text_color=UI_THEME["text_primary"], font=ctk.CTkFont(family="Segoe UI", size=9), command=lambda tid=tarefa["id"], st=status: _mover_tarefa_ui(tid, st)).pack(side="left", padx=2)


def _criar_tela_medicao(parent):
    import threading
    import traceback
    from tkinter import simpledialog

    _LOG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs", "medicao_audit.log")

    def _log(msg):
        try:
            os.makedirs(os.path.dirname(_LOG), exist_ok=True)
            with open(_LOG, "a", encoding="utf-8") as _f:
                _f.write(f"{msg}\n")
        except Exception:
            pass

    def _abrir_relatorio_file(path):
        try:
            os.startfile(path)
        except Exception:
            import webbrowser
            webbrowser.open(f"file:///{path.replace(os.sep, '/')}")

    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])

    # Shared state — use lists so closures in nested functions can mutate them
    _state = {"folder": "", "report_path": None, "diagnostic_path": None}

    # ── Header ──────────────────────────────────────────────────────────
    hdr = _criar_card(tela, corner_radius=20)
    hdr.pack(fill="x", padx=18, pady=(0, 10))
    ctk.CTkLabel(hdr, text="🔍  Auditoria de Medição — Energisa",
                 font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
                 text_color=UI_THEME["text_primary"]).pack(anchor="w", padx=18, pady=(14, 2))
    ctk.CTkLabel(hdr, text="Horizonte Logística — Verificação de Documentação Mensal",
                 font=ctk.CTkFont(family="Segoe UI", size=11),
                 text_color=UI_THEME["text_secondary"]).pack(anchor="w", padx=18, pady=(0, 14))

    # ── Folder selection ─────────────────────────────────────────────────
    sel_card = _criar_card(tela, corner_radius=20)
    sel_card.pack(fill="x", padx=18, pady=(0, 10))
    ctk.CTkLabel(sel_card, text="Pasta da Competência:",
                 font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold"),
                 text_color=UI_THEME["text_primary"], anchor="w").pack(fill="x", padx=18, pady=(14, 4))

    folder_var = tk.StringVar()
    sel_row = ctk.CTkFrame(sel_card, fg_color="transparent")
    sel_row.pack(fill="x", padx=18, pady=(0, 14))
    sel_row.grid_columnconfigure(0, weight=1)
    ctk.CTkEntry(sel_row, textvariable=folder_var,
                 font=ctk.CTkFont(family="Segoe UI", size=11), height=36,
                 corner_radius=10).grid(row=0, column=0, sticky="ew", padx=(0, 8))

    def _selecionar_pasta():
        last = folder_var.get() or "C:/Users/ismae/OneDrive/Área de Trabalho/MEDIÇÕES HORIZONTRE X ENERGISA/2026"
        d = filedialog.askdirectory(title="Selecione a pasta da competência", initialdir=last)
        if d:
            folder_var.set(d)
            _state["folder"] = d
            _state["report_path"] = None
            _state["diagnostic_path"] = None
            _set_status(f"Pasta: {os.path.basename(d)}", UI_THEME["text_secondary"])
            btn_rel.configure(state="disabled")
            btn_diag.configure(state="disabled")

    ctk.CTkButton(sel_row, text="Procurar…", width=110, height=36, corner_radius=10,
                  command=_selecionar_pasta,
                  font=ctk.CTkFont(family="Segoe UI", size=11)).grid(row=0, column=1)

    # ── Audit card ───────────────────────────────────────────────────────
    act_card = _criar_card(tela, corner_radius=20)
    act_card.pack(fill="x", padx=18, pady=(0, 10))

    btn_run = ctk.CTkButton(act_card, text="▶  INICIAR AUDITORIA", height=46, corner_radius=13,
                            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
                            fg_color="#27ae60", hover_color="#1e8449",
                            command=lambda: _iniciar())
    btn_run.pack(fill="x", padx=18, pady=(14, 8))

    ocr_var = tk.BooleanVar(value=False)
    ctk.CTkCheckBox(
        act_card,
        text="Análise profunda com OCR para PDFs escaneados",
        variable=ocr_var,
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(0, 8))

    # ttk.Progressbar — identical to what worked in the standalone app
    progress_style = ttk.Style()
    progress_style.configure("Medicao.Horizontal.TProgressbar",
                              troughcolor=UI_THEME.get("surface_alt", "#e0e0e0"),
                              background=UI_THEME.get("accent", "#2980b9"), thickness=8)
    progress = ttk.Progressbar(act_card, mode="indeterminate", style="Medicao.Horizontal.TProgressbar")
    progress.pack(fill="x", padx=18, pady=(0, 6))

    status_var = tk.StringVar(value="Selecione a pasta da competência para iniciar.")
    status_lbl = ctk.CTkLabel(act_card, textvariable=status_var,
                               font=ctk.CTkFont(family="Segoe UI", size=11),
                               text_color=UI_THEME["text_secondary"], wraplength=560)
    status_lbl.pack(padx=18, pady=(0, 8))

    btn_rel = ctk.CTkButton(act_card, text="📄  Ver Relatório no Navegador", height=38,
                             corner_radius=11, state="disabled",
                             font=ctk.CTkFont(family="Segoe UI", size=12),
                             command=lambda: _ver_relatorio())
    btn_rel.pack(fill="x", padx=18, pady=(0, 8))

    btn_diag = ctk.CTkButton(act_card, text="🧪  Abrir Diagnóstico da Auditoria", height=34,
                             corner_radius=11, state="disabled",
                             fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"],
                             text_color=UI_THEME["text_primary"],
                             font=ctk.CTkFont(family="Segoe UI", size=11),
                             command=lambda: _ver_diagnostico())
    btn_diag.pack(fill="x", padx=18, pady=(0, 14))

    # ── Secondary buttons ────────────────────────────────────────────────
    sec_card = _criar_card(tela, corner_radius=20)
    sec_card.pack(fill="x", padx=18, pady=(0, 10))
    sec_row = ctk.CTkFrame(sec_card, fg_color="transparent")
    sec_row.pack(fill="x", padx=18, pady=12)
    sec_row.grid_columnconfigure((0, 1), weight=1)

    ctk.CTkButton(sec_row, text="📁  Criar Pastas Modelo", height=38, corner_radius=11,
                  font=ctk.CTkFont(family="Segoe UI", size=11),
                  fg_color=UI_THEME["surface_alt"], hover_color=UI_THEME["tab_hover"],
                  text_color=UI_THEME["text_primary"],
                  command=lambda: _criar_pastas()).grid(row=0, column=0, sticky="ew", padx=(0, 4))

    ctk.CTkButton(sec_row, text="🗂️  Organizar Escaneados", height=38, corner_radius=11,
                  font=ctk.CTkFont(family="Segoe UI", size=11),
                  fg_color="#7d5c9e", hover_color="#6a4d88",
                  command=lambda: _MedicaoOrganizerDialog(tela, folder_var.get())
                  ).grid(row=0, column=1, sticky="ew", padx=(4, 0))

    # ── Helpers ──────────────────────────────────────────────────────────
    def _set_status(msg, color):
        status_var.set(msg)
        try:
            status_lbl.configure(text_color=color)
        except Exception:
            pass

    def _audit_done(msg, color, report_path):
        try:
            progress.stop()
        except Exception:
            pass
        try:
            _set_status(msg, color)
            btn_run.configure(state="normal")
            btn_rel.configure(state="normal")
            btn_diag.configure(state="normal")
        except Exception as exc:
            _log(f"[_audit_done] widget error: {exc}")
        if report_path and os.path.isfile(report_path):
            _abrir_relatorio_file(report_path)

    def _audit_error(err_msg):
        try:
            progress.stop()
        except Exception:
            pass
        try:
            _set_status(f"Erro: {err_msg}", "#e74c3c")
            btn_run.configure(state="normal")
            btn_diag.configure(state="normal" if _state.get("diagnostic_path") else "disabled")
        except Exception:
            pass
        messagebox.showerror("Erro na auditoria", err_msg)

    def _iniciar():
        folder = folder_var.get().strip()
        if not folder:
            messagebox.showwarning("Atenção", "Selecione uma pasta antes de iniciar.")
            return
        if not os.path.isdir(folder):
            messagebox.showerror("Erro", f"Pasta não encontrada:\n{folder}")
            return
        enable_ocr = bool(ocr_var.get())
        if enable_ocr:
            ok = messagebox.askyesno(
                "Análise profunda com OCR",
                "O OCR pode demorar em PDFs grandes. Deseja continuar com a análise profunda?"
            )
            if not ok:
                return

        btn_run.configure(state="disabled")
        btn_rel.configure(state="disabled")
        btn_diag.configure(state="disabled")
        progress.start(12)
        _set_status("Auditando em modo profundo com OCR..." if enable_ocr else "Auditando em modo rápido...", UI_THEME.get("accent", "#2980b9"))

        # Clear previous log for this run
        try:
            os.makedirs(os.path.dirname(_LOG), exist_ok=True)
            with open(_LOG, "w", encoding="utf-8") as _f:
                _f.write(f"[audit] folder={folder}\n")
                _f.write(f"[audit] mode={'ocr' if enable_ocr else 'safe'}\n")
        except Exception:
            pass

        def _progress_msg(msg):
            _log(f"[progress] {msg}")
            app.after(0, lambda m=msg: _set_status(m, UI_THEME.get("accent", "#2980b9")))

        def _write_diagnostic(result):
            diagnostics = result.get("diagnostics", {})
            path = os.path.join(os.path.dirname(_LOG), "medicao_diagnostico.txt")
            try:
                lines = [
                    "Diagnóstico da Auditoria de Medição",
                    f"Pasta: {folder}",
                    f"Modo: {diagnostics.get('mode', '-')}",
                    f"PDFs encontrados: {diagnostics.get('pdf_files', 0)}",
                    f"PDFs analisados agora: {diagnostics.get('analyzed_files', 0)}",
                    f"Cache reutilizado: {diagnostics.get('cache_hits', 0)}",
                    f"Páginas vistas: {diagnostics.get('pages_seen', 0)}",
                    f"Páginas processadas: {diagnostics.get('pages_processed', 0)}",
                    f"Páginas com OCR: {diagnostics.get('ocr_pages', 0)}",
                    "",
                    "Documentos encontrados:",
                ]
                for doc in diagnostics.get("documents_found", [])[:80]:
                    pages = doc.get("pages") or []
                    page_txt = f"{min(pages)}-{max(pages)}" if pages else "?"
                    lines.append(
                        f"- {doc.get('documentName')} | {doc.get('fileName')} | páginas {page_txt} | confiança {doc.get('confidence')} | {doc.get('method')}"
                    )
                if diagnostics.get("errors"):
                    lines.append("")
                    lines.append("Avisos/erros:")
                    lines.extend(f"- {err}" for err in diagnostics.get("errors", [])[:80])
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, "w", encoding="utf-8") as f:
                    f.write("\n".join(lines))
                return path
            except Exception as exc:
                _log(f"[diagnostic] write failed: {exc}")
                return None

        def _worker():
            try:
                _log("[audit] run_audit starting")
                result = _medicao_run_audit(
                    folder,
                    enable_ocr=enable_ocr,
                    analyze_pdf_content=enable_ocr,
                    progress_cb=_progress_msg,
                )
                _log("[audit] run_audit done, generating report")
                rpath = _medicao_generate_report(result, folder)
                _log(f"[audit] report written: {rpath}")
                _state["report_path"] = rpath
                _state["diagnostic_path"] = _write_diagnostic(result)
                n = len(result.get("all_issues", []))
                ov = result.get("overall_status", "ok")
                if ov == "ok":
                    msg = "✅ Auditoria concluída — Nenhuma pendência."
                    color = "#27ae60"
                elif ov == "warning":
                    msg = f"⚠️ Auditoria concluída — {n} ponto(s) de atenção."
                    color = "#f39c12"
                else:
                    msg = f"❌ Auditoria concluída — {n} pendência(s)."
                    color = "#e74c3c"
                _log(f"[audit] scheduling _audit_done: {msg}")
                app.after(0, lambda m=msg, c=color, p=rpath: _audit_done(m, c, p))
            except Exception as exc:
                tb = traceback.format_exc()
                _log(f"[audit] EXCEPTION:\n{tb}")
                err = str(exc)
                app.after(0, lambda e=err: _audit_error(e))

        threading.Thread(target=_worker, daemon=True).start()

    def _ver_relatorio():
        p = _state.get("report_path")
        if p and os.path.isfile(p):
            _abrir_relatorio_file(p)
        else:
            messagebox.showwarning("Atenção", "Execute a auditoria primeiro.")

    def _ver_diagnostico():
        p = _state.get("diagnostic_path")
        if p and os.path.isfile(p):
            _abrir_relatorio_file(p)
        else:
            messagebox.showwarning("Atenção", "Execute a auditoria primeiro.")

    def _criar_pastas():
        default = folder_var.get().strip()
        parent_dir = os.path.dirname(default) if default and os.path.isdir(default) \
            else "C:/Users/ismae/OneDrive/Área de Trabalho/MEDIÇÕES HORIZONTRE X ENERGISA/2026"
        parent_dir = filedialog.askdirectory(title="Pasta do ano (onde criar a competência)", initialdir=parent_dir)
        if not parent_dir:
            return
        comp_name = simpledialog.askstring("Nova Competência", "Nome da competência (ex: ABRIL):", parent=tela)
        if not comp_name:
            return
        try:
            comp_path, created = _medicao_scaffold(parent_dir, comp_name.strip().upper())
            folder_var.set(comp_path)
            _state["folder"] = comp_path
            _set_status(f"✅ Pastas criadas: {comp_name.upper()} ({len(created)} subpastas)", "#27ae60")
            messagebox.showinfo("Pronto", f'Competência "{comp_name.upper()}" criada com {len(created)} subpastas.')
        except Exception as exc:
            messagebox.showerror("Erro", str(exc))

    return tela


def _criar_tela_configuracoes(parent):
    tela = ctk.CTkFrame(parent, fg_color=UI_THEME["app_bg"])
    card = _criar_card(tela, corner_radius=20)
    card.pack(fill="x", padx=18, pady=(0, 16))
    ctk.CTkLabel(
        card,
        text="Configurações do sistema",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(14, 10))

    grade = ctk.CTkFrame(card, fg_color="transparent")
    grade.pack(fill="x", padx=16, pady=(0, 16))
    grade.grid_columnconfigure((0, 1), weight=1)
    tem_botoes = _renderizar_grade_acoes_por_tela(grade, "configuracoes")
    if not tem_botoes:
        ctk.CTkLabel(
            card,
            text="Sem botões nesta aba. Use 'Trocar sequência dos botões' para organizar.",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=UI_THEME["text_secondary"],
        ).pack(anchor="w", padx=18, pady=(0, 16))
    diag = _criar_card(tela, corner_radius=20)
    diag.pack(fill="x", padx=18, pady=(0, 16))
    ctk.CTkLabel(
        diag,
        text="Seguranca e diagnostico",
        font=ctk.CTkFont(family="Segoe UI", size=18, weight="bold"),
        text_color=UI_THEME["text_primary"],
    ).pack(anchor="w", padx=18, pady=(14, 2))
    ctk.CTkLabel(
        diag,
        text="Backups locais, restauracao e leitura dos ultimos logs do sistema.",
        font=ctk.CTkFont(family="Segoe UI", size=11),
        text_color=UI_THEME["text_secondary"],
    ).pack(anchor="w", padx=18, pady=(0, 10))
    diag_grid = ctk.CTkFrame(diag, fg_color="transparent")
    diag_grid.pack(fill="x", padx=16, pady=(0, 16))
    diag_grid.grid_columnconfigure((0, 1, 2, 3, 4, 5), weight=1)
    diag_botoes = [
        ("Backup agora", criar_backup_manual_ui, UI_THEME["accent"]),
        ("Restaurar backup", _abrir_dialogo_restaurar_backup, "#D85B6A"),
        ("Ver logs", _abrir_dialogo_logs, UI_THEME["surface_alt"]),
        ("Pasta de backups", lambda: _abrir_pasta_no_explorer(_cfg.BACKUP_DIR), UI_THEME["surface_alt"]),
        ("Busca global", _abrir_dialogo_busca_global, UI_THEME["surface_alt"]),
        ("Consistencia", _abrir_dialogo_auditoria_consistencia, UI_THEME["surface_alt"]),
    ]
    for idx, (texto, comando, cor) in enumerate(diag_botoes):
        destaque = cor not in {UI_THEME["surface_alt"], UI_THEME["surface"]}
        ctk.CTkButton(
            diag_grid,
            text=texto,
            height=38,
            corner_radius=11,
            fg_color=cor,
            hover_color=UI_THEME["accent_hover"] if cor == UI_THEME["accent"] else ("#B94755" if cor == "#D85B6A" else UI_THEME["tab_hover"]),
            border_width=0 if destaque else 1,
            border_color=UI_THEME["border"],
            text_color=UI_THEME["on_accent"] if destaque else UI_THEME["text_primary"],
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            command=comando,
        ).grid(row=0, column=idx, sticky="ew", padx=5)
    return tela


def construir_tela_principal():
    global container_scroll, main_frame

    app.screens = {}
    app.current_screen = ""
    screen_nav_buttons.clear()
    action_buttons.clear()
    tab_buttons.clear()
    tab_hover_state.clear()
    summary_metric_widgets.clear()
    _liberar_graficos_dashboard()
    ui_refs.clear()
    _recarregar_layout_interface()

    container_scroll = ctk.CTkFrame(app, fg_color=UI_THEME["app_bg"], corner_radius=0)
    container_scroll.pack(fill="both", expand=True)
    ui_refs["container_shell"] = container_scroll

    container_scroll.grid_columnconfigure(0, weight=0)
    container_scroll.grid_columnconfigure(1, weight=1)
    container_scroll.grid_rowconfigure(0, weight=1)

    sidebar = ctk.CTkFrame(
        container_scroll,
        width=300,
        fg_color=UI_THEME["header_bg"],
        corner_radius=0,
        border_width=0,
    )
    sidebar.grid(row=0, column=0, sticky="ns")
    sidebar.grid_propagate(False)
    ui_refs["sidebar"] = sidebar

    content_shell = ctk.CTkFrame(
        container_scroll,
        fg_color=UI_THEME["app_bg"],
        corner_radius=24,
        border_width=1,
        border_color=UI_THEME["border"],
    )
    content_shell.grid(row=0, column=1, sticky="nsew", padx=(0, 10), pady=10)
    content_shell.grid_columnconfigure(0, weight=1)
    content_shell.grid_rowconfigure(0, weight=1)
    ui_refs["content_shell"] = content_shell

    scroll_canvas = tk.Canvas(
        content_shell,
        highlightthickness=0,
        bd=0,
        bg=UI_THEME["app_bg"],
    )
    scroll_vertical = tk.Scrollbar(
        content_shell,
        orient="vertical",
        jump=1,
        relief="flat",
        bd=0,
        highlightthickness=0,
        width=12,
        bg=UI_THEME["scroll_btn"],
        activebackground=UI_THEME["scroll_btn_hover"],
        troughcolor=UI_THEME["app_bg"],
    )
    scroll_state = {"job": None, "last_args": None, "dragging": False}

    def _executar_scroll(*args):
        try:
            if args and args[0] == "moveto":
                scroll_canvas.yview_moveto(float(args[1]))
            else:
                scroll_canvas.yview(*args)
        except Exception:
            return
        _forcar_redesenho_pos_scroll()

    def _flush_scroll():
        scroll_state["job"] = None
        args = scroll_state.get("last_args")
        if args:
            _executar_scroll(*args)

    def _comando_scroll(*args):
        scroll_state["last_args"] = args
        if scroll_state.get("dragging"):
            if scroll_state.get("job") is None:
                # Limita frequencia de redraw no arraste rapido para evitar ghosting.
                scroll_state["job"] = app.after(24, _flush_scroll)
            return
        _executar_scroll(*args)

    scroll_vertical.configure(command=_comando_scroll)
    scroll_canvas.configure(yscrollcommand=scroll_vertical.set)

    scroll_canvas.grid(row=0, column=0, sticky="nsew")
    scroll_vertical.grid(row=0, column=1, sticky="ns")
    ui_refs["scroll_canvas"] = scroll_canvas
    ui_refs["scrollbar_vertical"] = scroll_vertical

    main_frame = ctk.CTkFrame(scroll_canvas, fg_color=UI_THEME["app_bg"], corner_radius=0)
    scroll_window_id = scroll_canvas.create_window((0, 0), window=main_frame, anchor="nw")
    ui_refs["scroll_window_id"] = scroll_window_id

    def _atualizar_scrollregion(_event=None):
        try:
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))
        except Exception:
            pass

    def _ajustar_largura_frame(event):
        try:
            scroll_canvas.itemconfigure(scroll_window_id, width=event.width)
        except Exception:
            pass

    def _widget_no_mainframe(widget):
        atual = widget
        while atual is not None:
            if atual == main_frame:
                return True
            atual = getattr(atual, "master", None)
        return False

    def _rolar_mouse(event):
        if not _widget_no_mainframe(event.widget):
            return
        delta = int(getattr(event, "delta", 0))
        if delta == 0:
            return "break"
        passos = -int(delta / 120)
        if passos == 0:
            passos = -1 if delta > 0 else 1
        try:
            scroll_canvas.yview_scroll(passos, "units")
            _forcar_redesenho_pos_scroll()
        except Exception:
            pass
        return "break"

    def _press_scrollbar(_event=None):
        scroll_state["dragging"] = True

    def _release_scrollbar(_event=None):
        scroll_state["dragging"] = False
        if scroll_state.get("job") is not None:
            try:
                app.after_cancel(scroll_state["job"])
            except Exception:
                pass
            scroll_state["job"] = None
        _flush_scroll()

    main_frame.bind("<Configure>", _atualizar_scrollregion, add="+")
    scroll_canvas.bind("<Configure>", _ajustar_largura_frame, add="+")
    scroll_vertical.bind("<ButtonPress-1>", _press_scrollbar, add="+")
    scroll_vertical.bind("<ButtonRelease-1>", _release_scrollbar, add="+")
    if not getattr(app, "_bind_rolagem_global_principal", False):
        app.bind_all("<MouseWheel>", _rolar_mouse, add="+")
        setattr(app, "_bind_rolagem_global_principal", True)

    ui_refs["main_frame"] = main_frame

    _aplicar_logo_watermark(main_frame)
    _criar_navegacao_telas(sidebar)

    screen_host = ctk.CTkFrame(main_frame, fg_color=UI_THEME["app_bg"])
    screen_host.pack(fill="both", expand=True)
    app.screen_host = screen_host
    ui_refs["screen_host"] = screen_host

    dashboard_frame = _criar_tela_dashboard(screen_host)
    hoje_frame = _criar_tela_hoje(screen_host)
    relatorios_frame = _criar_tela_relatorios(screen_host)
    alteracoes_frame = _criar_tela_alteracoes(screen_host)
    seguros_frame = _criar_tela_seguros(screen_host)
    tarefas_frame = _criar_tela_tarefas(screen_host)
    configuracoes_frame = _criar_tela_configuracoes(screen_host)

    medicao_frame = _criar_tela_medicao(screen_host)

    app.registrar_tela("dashboard", dashboard_frame)
    app.registrar_tela("hoje", hoje_frame)
    app.registrar_tela("relatorios", relatorios_frame)
    app.registrar_tela("alteracoes", alteracoes_frame)
    app.registrar_tela("seguros", seguros_frame)
    app.registrar_tela("tarefas", tarefas_frame)
    app.registrar_tela("configuracoes", configuracoes_frame)
    app.registrar_tela("medicao", medicao_frame)
    app.mostrar_tela("dashboard")


construir_tela_principal()
_register_doc_on_change(_agendar_atualizacao_painel_alteracoes)
_register_doc_on_change(lambda: app.after(0, _recriar_tela_hoje))
aplicar_tema_interface()

# Mantém abertura previsível; telas largas podem ser acessadas pelo scroll e redimensionamento manual.
app.update_idletasks()
largura_ideal = min(1180, largura_tela - 10)
altura_ideal = min(760, altura_tela - 10)
centralizar_janela(app, largura_ideal, altura_ideal)

app.after(120, atualizar_dashboard)

try:
    app.mainloop()
finally:
    liberar_lock_instancia()
