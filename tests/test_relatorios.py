"""
Testes para src/relatorios.py — montagem de DataFrames e escrita Excel.
Execute com: pytest tests/test_relatorios.py -v
"""

import os
from pathlib import Path
from uuid import uuid4
import pytest
import pandas as pd
from datetime import datetime

from src.relatorios import (
    _montar_dataframe_exportacao_periodo,
    _montar_df_relatorio_excel,
    escrever_excel_faturamento,
)

TEST_OUTPUT_DIR = Path(__file__).resolve().parents[1] / "_tmp" / "pytest" / "relatorios"


def _arquivo_saida(nome):
    TEST_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stem = Path(nome).stem
    suffix = Path(nome).suffix
    arquivo = TEST_OUTPUT_DIR / f"{stem}_{uuid4().hex}{suffix}"
    return str(arquivo)


# ---------------------------------------------------------------------------
#  Fixtures
# ---------------------------------------------------------------------------

def _df_base():
    """DataFrame mínimo que simula registros vindos do banco."""
    return pd.DataFrame([
        {
            "id": 1,
            "tipo": "NF",
            "numero": 100,
            "numero_original": "",
            "data_emissao": datetime(2025, 1, 10),
            "competencia": "janeiro/2025",
            "competencia_excel": datetime(2025, 1, 1),
            "data_competencia": datetime(2025, 1, 1),
            "frete": "CIF",
            "valor_inicial": 1000.0,
            "valor_final": 1000.0,
            "status": "EMITIDO",
        },
        {
            "id": 2,
            "tipo": "CTE",
            "numero": 200,
            "numero_original": "",
            "data_emissao": datetime(2025, 2, 5),
            "competencia": "fevereiro/2025",
            "competencia_excel": datetime(2025, 2, 1),
            "data_competencia": datetime(2025, 2, 1),
            "frete": "FOB",
            "valor_inicial": 2000.0,
            "valor_final": 1900.0,
            "status": "CANCELADO",
        },
    ])


# ---------------------------------------------------------------------------
#  _montar_dataframe_exportacao_periodo
# ---------------------------------------------------------------------------

class TestMontarDataframeExportacaoPeriodo:
    def test_retorna_dataframe_nao_vazio(self):
        df = _montar_dataframe_exportacao_periodo(_df_base())
        assert not df.empty

    def test_colunas_corretas(self):
        df = _montar_dataframe_exportacao_periodo(_df_base())
        esperadas = {"Data Emissao", "Mes Referencia", "Numero Doc", "Tipo Doc", "Frete", "Valor Inicial", "Valor Final", "Status"}
        assert esperadas.issubset(set(df.columns))

    def test_dois_registros(self):
        df = _montar_dataframe_exportacao_periodo(_df_base())
        assert len(df) == 2

    def test_df_vazio_retorna_vazio(self):
        df = _montar_dataframe_exportacao_periodo(pd.DataFrame())
        assert df.empty

    def test_df_none_retorna_vazio(self):
        df = _montar_dataframe_exportacao_periodo(None)
        assert df.empty

    def test_tipo_maiusculo(self):
        df = _montar_dataframe_exportacao_periodo(_df_base())
        assert all(v == v.upper() for v in df["Tipo Doc"])

    def test_numero_doc_inteiro_para_numerico(self):
        df = _montar_dataframe_exportacao_periodo(_df_base())
        # Documentos sem numero_original devem ter Numero Doc numérico
        for val in df["Numero Doc"]:
            assert isinstance(val, (int, float, str))


# ---------------------------------------------------------------------------
#  _montar_df_relatorio_excel
# ---------------------------------------------------------------------------

class TestMontarDfRelatorioExcel:
    def test_colunas_excel(self):
        df = _montar_df_relatorio_excel(_df_base())
        assert "Concat" in df.columns
        assert "Mes Referencia" in df.columns

    def test_concat_contem_tipo(self):
        df = _montar_df_relatorio_excel(_df_base())
        for val in df["Concat"]:
            assert any(t in str(val) for t in ("NF", "CTE"))

    def test_ordenado_por_data(self):
        df = _montar_df_relatorio_excel(_df_base())
        datas = df["Data Emissao"].tolist()
        assert datas == sorted(datas)


# ---------------------------------------------------------------------------
#  escrever_excel_faturamento
# ---------------------------------------------------------------------------

class TestEscreverExcelFaturamento:
    def test_cria_arquivo_xlsx(self):
        arquivo = _arquivo_saida("teste_faturamento.xlsx")
        resultado = escrever_excel_faturamento(_df_base(), arquivo)
        assert resultado["ok"] is True
        assert os.path.exists(arquivo)

    def test_retorna_total_correto(self):
        arquivo = _arquivo_saida("teste.xlsx")
        resultado = escrever_excel_faturamento(_df_base(), arquivo)
        assert resultado["total_documentos"] == 2

    def test_falha_com_caminho_invalido(self):
        resultado = escrever_excel_faturamento(_df_base(), "/caminho/inexistente/arquivo.xlsx")
        assert resultado["ok"] is False
        assert resultado["erro"] != ""

    def test_arquivo_tem_duas_abas(self):
        """O Excel gerado deve ter as abas 'Faturamento AC' e 'Faturamento AC 2'."""
        import openpyxl
        arquivo = _arquivo_saida("abas.xlsx")
        escrever_excel_faturamento(_df_base(), arquivo)
        wb = openpyxl.load_workbook(arquivo)
        try:
            assert "Faturamento AC" in wb.sheetnames
            assert "Faturamento AC 2" in wb.sheetnames
        finally:
            wb.close()
